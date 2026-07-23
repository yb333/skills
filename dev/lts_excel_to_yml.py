#!/usr/bin/env python3
"""LTS 调度任务 Excel → yml 转换脚本。

把调度任务 Excel（schedule_tasks.xlsx）转成代码仓 yml 目录结构：
  LTS/项目/任务组/任务名称.yml

转换逻辑和术加执行平台 Excel→yml 一致：
  - 一个 yml 文件 = 一个调度任务
  - key 用中文表头（和代码仓现有 yml 风格一致）
  - 三个 Sheet（tasks/jobs/taskParams）按"任务名称"合并

用法:
    python3 lts_excel_to_yml.py schedule_tasks.xlsx --output LTS/
    python3 lts_excel_to_yml.py schedule_tasks.xlsx --output LTS/ --repo-root /path/to/repo
"""

import argparse
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml


def _read_sheet_as_dicts(ws):
    """读一个 Sheet，返回 list[dict]，key 是表头。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = val if val is not None else ""
        result.append(d)
    return result


def convert_lts_excel_to_yml(excel_path, output_dir, repo_root=None):
    """转换主函数。

    Args:
        excel_path: schedule_tasks.xlsx 路径
        output_dir: yml 输出根目录（如 LTS/）
        repo_root: 代码仓根（如果 output_dir 是相对路径，基于此解析）

    产出目录结构：
        output_dir/项目名称/任务组名称/任务名称.yml
    """
    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    if not output_dir.is_absolute() and repo_root:
        output_dir = Path(repo_root) / output_dir

    print(f"=== LTS 调度任务 Excel → yml ===")
    print(f"输入: {excel_path}")
    print(f"输出: {output_dir}")
    print()

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    # 读三个 Sheet
    tasks = _read_sheet_as_dicts(wb["tasks"]) if "tasks" in wb.sheetnames else []
    jobs = _read_sheet_as_dicts(wb["jobs"]) if "jobs" in wb.sheetnames else []
    task_params = _read_sheet_as_dicts(wb["taskParams"]) if "taskParams" in wb.sheetnames else []

    print(f"读取: {len(tasks)} 任务, {len(jobs)} job, {len(task_params)} 参数")

    # 按"项目/任务组/任务名"分组 jobs 和 params
    jobs_by_task = defaultdict(list)
    for j in jobs:
        key = (j.get("项目名称", ""), j.get("任务组名称", ""), j.get("任务名称", ""))
        jobs_by_task[key].append(j)

    params_by_task = defaultdict(list)
    for p in task_params:
        key = (p.get("项目名称", ""), p.get("任务组名称", ""), p.get("任务名称", ""))
        params_by_task[key].append(p)

    # 生成 yml
    count = 0
    for task in tasks:
        project = str(task.get("项目名称", "")).strip() or "UNKNOWN_PROJECT"
        task_group = str(task.get("任务组名称", "")).strip() or "UNKNOWN_GROUP"
        task_name = str(task.get("任务名称", "")).strip() or "UNKNOWN_TASK"

        # 安全目录名
        safe_project = _safe_dir_name(project)
        safe_group = _safe_dir_name(task_group)
        safe_task = _safe_dir_name(task_name)

        task_key = (project, task_group, task_name)
        task_jobs = jobs_by_task.get(task_key, [])
        task_params_list = params_by_task.get(task_key, [])

        # 构建 yml 数据（中文 key，和代码仓风格一致）
        yml_data = {}

        # 顶层：任务调度信息（tasks sheet 的字段）
        for k, v in task.items():
            if k in ("项目名称", "任务组名称", "任务名称"):
                continue  # 这三个用于目录，不进 yml
            yml_data[k] = _clean_value(v)

        # jobs（子表）
        if task_jobs:
            yml_data["Jobs"] = [
                {k: _clean_value(v) for k, v in j.items()
                 if k not in ("项目名称", "任务组名称", "任务名称") and v != ""}
                for j in task_jobs
            ]

        # 任务参数（子表）
        if task_params_list:
            yml_data["任务参数"] = [
                {"参数名称": p.get("参数名称", ""), "参数值": _clean_value(p.get("参数值", ""))}
                for p in task_params_list
            ]

        # 写文件
        yml_dir = output_dir / safe_project / safe_group
        yml_dir.mkdir(parents=True, exist_ok=True)
        yml_path = yml_dir / f"{safe_task}.yml"

        with open(yml_path, "w", encoding="utf-8") as f:
            yaml.dump(yml_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

        count += 1
        # 关键关联信息（V_GROUP_CODE）
        group_codes = [p.get("参数值") for p in task_params_list
                       if p.get("参数名称") == "V_GROUP_CODE" and p.get("参数值")]
        gc_str = f" → 规则组 {','.join(str(gc) for gc in group_codes)}" if group_codes else ""
        print(f"  + {safe_project}/{safe_group}/{safe_task}.yml{gc_str}")

    print(f"\n完成: 生成 {count} 个任务 yml")
    print(f"输出目录: {output_dir}")
    return count


def _safe_dir_name(name):
    """清洗目录名（去掉非法字符）。"""
    import re
    return re.sub(r'[<>:"/\\|?*\s]', "_", str(name)).strip("_") or "unknown"


def _clean_value(v):
    """清理单元格值：None→空串，datetime→字符串。"""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return v


def main():
    parser = argparse.ArgumentParser(
        description="LTS 调度任务 Excel → yml 转换"
    )
    parser.add_argument("input", help="schedule_tasks.xlsx 路径")
    parser.add_argument("--output", default="LTS/", help="yml 输出根目录（默认 LTS/）")
    parser.add_argument("--repo-root", default=None, help="代码仓根（output 相对路径基于此）")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"错误: 文件不存在: {args.input}", file=sys.stderr)
        sys.exit(1)

    convert_lts_excel_to_yml(args.input, args.output, args.repo_root)


if __name__ == "__main__":
    main()
