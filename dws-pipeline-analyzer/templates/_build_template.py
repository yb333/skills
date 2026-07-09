#!/usr/bin/env python3
"""生成用户可用的变更清单模板 Excel。

产出两个文件：
  1. 变更清单_模板.xlsx      — 空模板（只有列头，用户往里填）
  2. 变更清单_示例.xlsx      — 带示例数据（教用户怎么填）

运行:
    python _build_template.py
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
NOTE_FONT = Font(italic=True, color="808080", size=10)

SHEET1_NAME = "表级变更"
SHEET1_HEADERS = ["切换前表名", "切换后表名", "是否平切", "切换说明", "表级变化类型"]
SHEET1_NOTE = "表级变化类型：表/视图下线、表/视图替换、表/视图主键变化、表/视图名称或者schema变化、平切、表/视图数据初始化（刷时间戳）、表/视图初始化（不刷时间戳）、表/视图数据归档、表/视图取消权限、表/视图数据硬删除"

SHEET2_NAME = "字段级变更"
SHEET2_HEADERS = [
    "序号", "切换前数据库", "切换前表schema", "切换前表名",
    "切换前表字段名", "切换前表字段中文名", "切换前表字段类型",
    "切换后数据库", "切换后表schema", "切换后表名",
    "切换后表字段名", "切换后表字段中文名", "切换后表字段类型",
    "字段变化类型", "是否可还原(Y/N)", "还原方案详细说明",
    "源端IT责任人", "源端业务责任人",
]
SHEET2_NOTE = "字段变化类型：字段类型及长度变化、字段下线/删除、字段值语义变化、新增字段、字段名称变化、字段数据初始化（刷时间戳）、字段数据初始化（不刷时间戳）"

SHEET3_NAME = "变动类型说明"
SHEET3_HEADERS = ["类型", "说明"]


def _set_headers(ws, headers, note=""):
    """写入表头行 + 格式化"""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    if note:
        ws.append([])
        ws.append([note])
        ws.cell(row=ws.max_row, column=1).font = NOTE_FONT
    ws.freeze_panes = "A2"


def build_empty_template(path: str):
    """空模板：只有列头"""
    wb = Workbook()

    ws1 = wb.active
    ws1.title = SHEET1_NAME
    _set_headers(ws1, SHEET1_HEADERS, SHEET1_NOTE)

    ws2 = wb.create_sheet(SHEET2_NAME)
    _set_headers(ws2, SHEET2_HEADERS, SHEET2_NOTE)

    ws3 = wb.create_sheet(SHEET3_NAME)
    _set_headers(ws3, SHEET3_HEADERS, "类型说明（可选）：补充每个变化类型的详细定义，用于报告翻译")

    wb.save(path)
    wb.close()
    print(f"[OK] 空模板: {path}")


def build_sample_template(path: str):
    """示例模板：带示例数据，教用户怎么填"""
    wb = Workbook()

    # ── Sheet1: 表级变更示例 ──
    ws1 = wb.active
    ws1.title = SHEET1_NAME
    _set_headers(ws1, SHEET1_HEADERS, SHEET1_NOTE)
    ws1.append(["ods.user_info_src", "ods_new.user_info", "Y", "用户表平切", "平切"])
    ws1.append(["ods.order_log_src", "", "N", "日志表整表下线", "表/视图下线"])
    ws1.append(["ods.product_src", "ods_new.product", "N", "表名变更", "表/视图名称或者schema变化"])
    ws1.append(["ods.config_src", "ods_new.config", "N", "数据初始化", "表/视图初始化（不刷时间戳）"])

    # ── Sheet2: 字段级变更示例 ──
    ws2 = wb.create_sheet(SHEET2_NAME)
    _set_headers(ws2, SHEET2_HEADERS, SHEET2_NOTE)
    samples = [
        [1, "ods", "public", "user_info_src", "user_id", "用户ID", "varchar(20)",
         "ods_new", "public", "user_info", "user_id", "用户ID", "varchar(50)",
         "字段类型及长度变化", "Y", "CAST兼容，无需还原", "张三", "李四"],
        [2, "ods", "public", "user_info_src", "old_status", "旧状态", "int",
         "", "", "", "", "", "",
         "字段下线/删除", "N", "", "张三", "李四"],
        [3, "ods", "public", "user_info_src", "status", "状态", "int",
         "ods_new", "public", "user_info", "status", "状态", "int",
         "字段值语义变化", "N", "状态码从0/1改为A/B/C", "张三", "李四"],
        [4, "ods", "public", "user_info_src", "phone", "手机号", "varchar(20)",
         "ods_new", "public", "user_info", "mobile", "手机号", "varchar(20)",
         "字段名称变化", "Y", "引用处改字段名即可", "张三", "李四"],
    ]
    for row in samples:
        ws2.append(row)

    # ── Sheet3: 类型说明示例 ──
    ws3 = wb.create_sheet(SHEET3_NAME)
    _set_headers(ws3, SHEET3_HEADERS, "类型说明（可选）：补充每个变化类型的详细定义")
    type_docs = [
        ["字段类型及长度变化", "字段的数据类型或长度发生变化，需检查下游cast/转换是否兼容"],
        ["字段下线/删除", "字段被废弃删除，下游取不到该数据"],
        ["字段值语义变化", "字段名和类型不变，但值的含义变了（如状态码含义改变）"],
        ["新增字段", "源端新增字段，下游一般无影响（未引用）"],
        ["字段名称变化", "字段重命名，数据不变但引用处需改名"],
        ["字段数据初始化（刷时间戳）", "数据被重新初始化并更新时间戳，增量下游会重新拉取"],
        ["字段数据初始化（不刷时间戳）", "数据被重新初始化但不更新时间戳，增量下游可能漏拉"],
        ["表/视图下线", "整表下线，来源消失"],
        ["表/视图替换", "表被替换为另一张表"],
        ["表/视图主键变化", "主键改变，影响JOIN/去重逻辑"],
        ["表/视图名称或者schema变化", "表名/schema变了，需改引用+术+规则+调度"],
        ["平切", "字段完全一致仅表名变化"],
        ["表/视图数据初始化（刷时间戳）", "整表数据初始化并更新时间戳"],
        ["表/视图初始化（不刷时间戳）", "整表数据初始化但不更新时间戳"],
        ["表/视图数据归档", "历史数据归档，需确认是否影响查询"],
        ["表/视图取消权限", "权限取消，下游无法访问"],
        ["表/视图数据硬删除", "数据被硬删除"],
    ]
    for t, d in type_docs:
        ws3.append([t, d])

    wb.save(path)
    wb.close()
    print(f"[OK] 示例模板: {path}")


if __name__ == "__main__":
    out_dir = Path(__file__).resolve().parent
    build_empty_template(str(out_dir / "变更清单_模板.xlsx"))
    build_sample_template(str(out_dir / "变更清单_示例.xlsx"))
