"""
端到端用例 — 真实 Excel → 完整 analyzer + view_generator → 三视图产物验证

验证最终产物质量，不只是中间数据结构。
用真实结构的 Excel（docs/execution_tasks.xlsx 和 docs/multi_scenario_test.xlsx）。

运行:
    pytest tests/test_end_to_end.py -v
"""

import sys
import json
import shutil
import tempfile
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).parent.parent
ANALYZER_SKILL = PROJECT_ROOT / ".opencode" / "skills" / "dws-pipeline-analyzer" / "references"
sys.path.insert(0, str(ANALYZER_SKILL))

from analyzer import read_excel, detect_dialect, parse_single_sql, build_topology, build_data_flow, build_field_mappings, analyze_quality
from view_generator import generate_mapping, generate_asset_report, generate_tech_design


def run_full_analysis(xlsx_path: str, output_dir: str, ddl_dir: str = None):
    """完整分析流程: Excel → knowledge → 三视图"""
    from analyzer import parse_ddl_for_metadata, detect_patterns, build_source, generate_step_description
    from datetime import datetime

    raw = read_excel(xlsx_path)
    rules = raw["rules"]
    if not rules:
        return None, "无有效规则"

    dialect = detect_dialect([r.query_sql for r in rules if r.query_sql])
    parsed_map = {}
    for rule in rules:
        parsed_map[rule.rule_code] = parse_single_sql(rule.query_sql, dialect)

    topology = build_topology(rules, parsed_map)
    data_flow = build_data_flow(rules, parsed_map)
    field_mappings = build_field_mappings(rules, parsed_map, raw["target_fields"])
    quality = analyze_quality(topology, data_flow, field_mappings, parsed_map)

    # DDL 元数据
    target_metadata = {}
    if ddl_dir:
        target_name = rules[0].target_table or "unknown"
        target_metadata = parse_ddl_for_metadata(ddl_dir, target_name)

    patterns = detect_patterns(parsed_map, topology)

    # 兜底描述
    scenarios = topology.get("scenarios", [])
    auto_step_desc = []
    for rule in rules:
        parsed = parsed_map.get(rule.rule_code)
        desc = generate_step_description(rule, parsed, scenarios, rules)
        step = next((s for s in topology["steps"] if s["rule_code"] == rule.rule_code), None)
        auto_step_desc.append({
            "step_id": step["step_id"] if step else "",
            "rule_code": rule.rule_code,
            "purpose": desc["purpose"],
            "logic": desc["logic"],
        })

    knowledge = {
        "meta": {
            "source_type": "execution_tasks.xlsx",
            "analysis_time": datetime.now().isoformat(),
            "dialect": dialect,
            "total_rules": len(rules),
            "target_table": rules[0].target_table or "",
            "patterns": patterns,
            "target_field_types": {k: v["type"] for k, v in target_metadata.items() if v.get("type")},
            "target_field_comments": {k: v["comment"] for k, v in target_metadata.items() if v.get("comment")},
        },
        "topology": topology,
        "data_flow": data_flow,
        "field_mappings": field_mappings,
        "quality": quality,
        "business_logic": {
            "summary": "",
            "step_descriptions": auto_step_desc,
            "key_transforms": [],
        },
        "source": build_source(rules, raw["target_fields"], raw["group_variables"], parsed_map),
    }

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    (out / "knowledge_final.json").write_text(json.dumps(knowledge, ensure_ascii=False, indent=2))

    # 生成三视图
    results = {}
    results["mapping"] = generate_mapping(knowledge, str(out))
    results["asset"] = generate_asset_report(knowledge, str(out))
    results["techspec"] = generate_tech_design(knowledge, str(out))

    return knowledge, results


# ═══════════════════════════════════════════════════════════════
# 端到端测试
# ═══════════════════════════════════════════════════════════════

class TestEndToEnd:
    """真实 Excel → 三视图产物验证。"""

    @pytest.fixture
    def tmp_output(self):
        d = tempfile.mkdtemp()
        yield d
        shutil.rmtree(d, ignore_errors=True)

    def test_real_sample_single_rule(self, tmp_output):
        """真实样本（单规则 CTE + 行转列）：三视图全部生成成功。"""
        xlsx = str(PROJECT_ROOT / "docs" / "execution_tasks.xlsx")
        if not Path(xlsx).exists():
            pytest.skip("真实样本文件不存在")

        ddl = str(PROJECT_ROOT / "docs" / "output" / "dwl_con_pu_any_f" / "04_ddl")
        ddl_arg = ddl if Path(ddl).exists() else None

        knowledge, results = run_full_analysis(xlsx, tmp_output, ddl_arg)

        assert knowledge is not None, "分析失败"
        assert results["mapping"] == True, "mapping.xlsx 生成失败"
        assert results["asset"] == True, "asset_report.html 生成失败"
        assert results["techspec"] == True, "tech_design.md 生成失败"

        # 验证产物文件存在
        assert (Path(tmp_output) / "mapping.xlsx").exists()
        assert (Path(tmp_output) / "asset_report.html").exists()
        assert (Path(tmp_output) / "tech_design.md").exists()

        # 验证 HTML 非空且含关键内容
        html = (Path(tmp_output) / "asset_report.html").read_text()
        assert "REPORT_DATA" in html
        assert len(html) > 5000, f"HTML 内容过少: {len(html)} 字符"

        # 验证 knowledge 关键结构
        assert len(knowledge["field_mappings"]["fields"]) > 0
        assert len(knowledge["topology"]["steps"]) > 0
        # CTE 穿透验证（真实样本有 CTE）
        # 检查有 CTE 的规则
        for rule in knowledge.get("source", {}).get("rule_sheet_raw", []):
            rule_sql = ""
            if isinstance(rule, dict):
                rule_sql = rule.get("query_sql", "")
            elif isinstance(rule, str):
                rule_sql = rule
            if rule_sql and "WITH" in rule_sql.upper():
                parsed = parse_single_sql(rule_sql, "dws")
                if parsed.ctes:
                    agg_fields = [f for f in knowledge["field_mappings"]["fields"]
                                 if f.get("transform_type") in ("aggregate",)]
                    assert len(agg_fields) > 0, "有CTE但无聚合穿透字段"
                break

    def test_multi_scenario_sample(self, tmp_output):
        """多场景样本（2场景+公共步骤）：场景分组正确，三视图生成。"""
        xlsx = str(PROJECT_ROOT / "docs" / "multi_scenario_test.xlsx")
        if not Path(xlsx).exists():
            pytest.skip("多场景样本文件不存在")

        knowledge, results = run_full_analysis(xlsx, tmp_output)

        assert knowledge is not None
        assert results["mapping"] == True
        assert results["asset"] == True
        assert results["techspec"] == True

        # 验证场景分组
        scenarios = knowledge["topology"].get("scenarios", [])
        non_common = [s for s in scenarios if not s.get("is_common")]
        assert len(non_common) >= 2, f"多场景应有2+个分区场景，实际 {len(non_common)}"

        # 验证 HTML 含场景标签
        html = (Path(tmp_output) / "asset_report.html").read_text()
        assert "is_multi_scenario" in html

    def test_mapping_xlsx_structure(self, tmp_output):
        """mapping.xlsx 结构验证：两个 sheet 都有数据。"""
        xlsx = str(PROJECT_ROOT / "docs" / "execution_tasks.xlsx")
        if not Path(xlsx).exists():
            pytest.skip("真实样本文件不存在")

        knowledge, results = run_full_analysis(xlsx, tmp_output)
        assert results["mapping"] == True

        # 读 mapping.xlsx 验证
        import openpyxl
        wb = openpyxl.load_workbook(Path(tmp_output) / "mapping.xlsx", read_only=True)
        assert "实体级mapping" in wb.sheetnames
        assert "属性级mapping" in wb.sheetnames

        entity_ws = wb["实体级mapping"]
        attr_ws = wb["属性级mapping"]
        # 实体级至少有表头+1行数据
        assert entity_ws.max_row >= 2, f"实体级mapping行数过少: {entity_ws.max_row}"
        # 属性级至少有表头+1行数据
        assert attr_ws.max_row >= 2, f"属性级mapping行数过少: {attr_ws.max_row}"
        wb.close()

    def test_tech_design_structure(self, tmp_output):
        """tech_design.md 结构验证：包含关键章节。"""
        xlsx = str(PROJECT_ROOT / "docs" / "execution_tasks.xlsx")
        if not Path(xlsx).exists():
            pytest.skip("真实样本文件不存在")

        knowledge, results = run_full_analysis(xlsx, tmp_output)
        assert results["techspec"] == True

        md = (Path(tmp_output) / "tech_design.md").read_text()
        # 验证关键章节
        assert "## 1." in md or "# " in md, "缺少标题"
        assert len(md) > 500, f"tech_design.md 内容过少: {len(md)} 字符"


def rules_sql(knowledge):
    """从 knowledge 提取第一条 SQL（用于辅助验证）"""
    sources = knowledge.get("source", {})
    sqls = sources.get("raw_sql", [])
    return sqls[0] if sqls else ""
