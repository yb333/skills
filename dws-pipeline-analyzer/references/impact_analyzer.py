"""关联影响分析（单资产 MVP）

定位：影响清单 + 定位器，不是权威影响报告。
核心价值 = 自动化两端（确定有/无影响），把中间（判不了）留给人。

架构分层（见 architecture.md §6.2）：
    ① 输入解析层   read_changes(xlsx) → 三 Sheet 解析 + 容错
    ② 传播层      filter + propagate（确定性图遍历，与判定分离）
    ③ 判定层      assess_severity（映射表驱动，可独立替换/AI化）
    ④ 渲染层      render_excel（三 Sheet：影响清单/表级影响/过滤摘要）

铁律：传播（确定性）与影响判定（带规则）必须分离。
  - cast 吸收型变化会"传播"到下游，但"影响"是 nil。
  - 不分离会把所有传播都报成有影响，假阳性爆炸。

用法:
    python impact_analyzer.py --changes 变更清单.xlsx --knowledge knowledge.json --output impact.xlsx
"""

import sys
import json
import argparse
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl.styles import Alignment

# ═══════════════════════════════════════════════════════════════
# 数据类
# ═══════════════════════════════════════════════════════════════


@dataclass
class TableChange:
    """Sheet1: 源系统切换前后表级 mapping（整表维度）"""
    before_table: str = ""       # 切换前表名（资产 SQL 当前引用的）
    after_table: str = ""        # 切换后表名
    is_ping_cut: bool = False    # 是否平切（字段完全一致仅名称变化）
    note: str = ""               # 切换说明
    change_type: str = ""        # 表级变化类型（人工标注，优先于推断）

    @property
    def is_table_dropped(self) -> bool:
        """整表下线：切换前有表名，切换后为空"""
        return bool(self.before_table) and not self.after_table

    @property
    def derived_table_type(self) -> str:
        """从标志位推导表级变化类型（change_type 为空时用）"""
        if self.change_type:
            return self.change_type
        if self.is_ping_cut:
            return "平切"
        if self.is_table_dropped:
            return "表/视图下线"
        return "表级变更"


@dataclass
class ChangeItem:
    """Sheet2: 源系统切换前后字段级 mapping（字段维度）

    每行描述一个字段的变化，前后对照自动推导变化类型。
    """
    # 切换前
    before_db: str = ""
    before_schema: str = ""
    before_table: str = ""
    before_field: str = ""
    before_field_cn: str = ""
    before_type: str = ""
    # 切换后
    after_db: str = ""
    after_schema: str = ""
    after_table: str = ""
    after_field: str = ""
    after_field_cn: str = ""
    after_type: str = ""
    # 人工标注
    change_type: str = ""        # 字段变化类型（人工填，优先于推导）
    recoverable: str = ""        # 是否可还原 Y/N
    recovery_plan: str = ""      # 还原方案详细说明
    source_it_owner: str = ""    # 源端IT责任人
    source_biz_owner: str = ""   # 源端业务责任人

    @property
    def table(self) -> str:
        """资产 SQL 引用的是切换前表名"""
        return self.before_table

    @property
    def field(self) -> str:
        return self.before_field

    @property
    def derived_change_type(self) -> str:
        """从前后对照推导变化类型（change_type 为空时用）"""
        if self.change_type:
            return self.change_type
        # 推导逻辑（草案，待变化类型专题讨论后完善）
        if self.before_field and not self.after_field:
            return "1:0废弃字段"
        if not self.before_field and self.after_field:
            return "0:1新增字段"
        if self.before_field and self.after_field:
            if self.before_type != self.after_type:
                return "1:1数据类型/长度变化"
            if self.before_field != self.after_field:
                return "字段名称变化"
            return "1:1数据内容变化"
        return ""


@dataclass
class Hop:
    """传播路径中的单跳"""
    step: str = ""               # step_id
    source_table: str = ""       # 真实表名（从别名解析）
    source_field: str = ""       # 输入字段
    expression: str = ""         # 该跳表达式（raw_sql）
    output: str = ""             # 输出别名
    rule_code: str = ""          # 规则编码


@dataclass
class ImpactPath:
    """一条完整的影响路径（一个目标字段被一个源变更命中）"""
    target_table: str = ""
    target_field: str = ""
    change: Optional[ChangeItem] = None
    hops: list = field(default_factory=list)   # list[Hop]
    status: str = ""             # 🔴有影响 / 🟡待确认 / 🟢无影响
    severity: str = ""           # high / low / none / unknown
    reason: str = ""             # 一句话说明
    uncertain_reason: str = ""   # 待确认原因（SELECT */断链等）


@dataclass
class AnalysisResult:
    """影响分析完整结果"""
    table_level_impacts: list = field(default_factory=list)   # Sheet2: 表级影响（含平切/下线）
    field_level_impacts: list = field(default_factory=list)   # Sheet1: 字段级影响清单
    filtered_out: list = field(default_factory=list)          # Sheet3: 过滤摘要（🟢⚪ + 未命中）
    summary: dict = field(default_factory=dict)               # 统计摘要


# ═══════════════════════════════════════════════════════════════
# ① 输入解析层（容错风格照搬 read_excel）
# ═══════════════════════════════════════════════════════════════

def _safe_str(val) -> str:
    """安全转字符串（与 analyzer.py 一致）"""
    if val is None:
        return ""
    return str(val).strip()


def _find_col(col_idx: dict, name: str) -> Optional[int]:
    """查找列索引。先精确匹配，再归一化（去空格+全角半角），再子串包含。"""
    if name in col_idx:
        return col_idx[name]

    def normalize(s):
        s = s.replace("（", "(").replace("）", ")")
        s = s.replace(" ", "").replace("\u3000", "")
        return s

    norm_name = normalize(name)
    for actual, idx in col_idx.items():
        if normalize(actual) == norm_name:
            return idx
    for actual, idx in col_idx.items():
        na = normalize(actual)
        if norm_name in na or na in norm_name:
            return idx
    return None


def _get_val(row: tuple, idx: Optional[int]) -> str:
    """安全获取行值（与 analyzer.py 一致）"""
    if idx is None or idx >= len(row):
        return ""
    return _safe_str(row[idx])


def _parse_bool(val: str) -> bool:
    """解析布尔值：平切 Y/是/true/平切 → True"""
    return val.upper() in ("Y", "YES", "TRUE", "1", "是", "平切")


def read_changes(path: str) -> tuple:
    """读取变更清单 Excel（三 Sheet）。

    返回 (table_changes: list[TableChange], field_changes: list[ChangeItem], type_dict: dict)

    容错策略（照搬 read_excel 风格）：
    - 列名模糊匹配（精确→归一化→子串）
    - 空值统一空串
    - 单条解析失败 continue 不中断
    - 缺 Sheet 诊断到 stderr，不抛异常
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERROR] 缺少依赖: openpyxl", file=sys.stderr)
        return [], [], {}

    wb_path = Path(path)
    if not wb_path.exists():
        print(f"[ERROR] 变更清单不存在: {path}", file=sys.stderr)
        return [], [], {}

    wb = load_workbook(wb_path, read_only=True, data_only=True)

    # ── Sheet1: 表级 mapping ──
    table_changes = []
    sheet1 = _find_sheet(wb, ["源系统切换前后表级mapping", "表级mapping", "表级"])
    if sheet1:
        table_changes = _parse_sheet1(sheet1)
    else:
        print("[WARN] 未找到表级 mapping Sheet", file=sys.stderr)

    # ── Sheet2: 字段级 mapping ──
    field_changes = []
    sheet2 = _find_sheet(wb, ["源系统切换前后字段mapping", "字段级mapping", "字段级"])
    if sheet2:
        field_changes = _parse_sheet2(sheet2)
    else:
        print("[WARN] 未找到字段级 mapping Sheet", file=sys.stderr)

    # ── Sheet3: 变动类型说明（可选，MVP 仅读为字典）──
    type_dict = {}
    sheet3 = _find_sheet(wb, ["源端变动类型", "变动类型", "类型"])
    if sheet3:
        type_dict = _parse_sheet3(sheet3)

    wb.close()
    return table_changes, field_changes, type_dict


def _find_sheet(wb, name_candidates: list):
    """按名称候选找 Sheet（模糊匹配）"""
    # 先精确
    for name in name_candidates:
        if name in wb.sheetnames:
            return wb[name]
    # 再包含
    for ws_name in wb.sheetnames:
        for candidate in name_candidates:
            if candidate in ws_name:
                return wb[ws_name]
    return None


def _parse_sheet1(ws) -> list:
    """解析表级 mapping Sheet"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 建列索引
    header = rows[0]
    col_idx = {}
    for i, h in enumerate(header):
        if h is not None:
            col_idx[_safe_str(h)] = i

    idx_before = _find_col(col_idx, "切换前表名")
    idx_after = _find_col(col_idx, "切换后表名")
    idx_ping = _find_col(col_idx, "是否平切")
    idx_note = _find_col(col_idx, "切换说明")
    idx_change_type = _find_col(col_idx, "表级变化类型") or _find_col(col_idx, "变化类型")

    table_changes = []
    for row in rows[1:]:
        before_t = _get_val(row, idx_before)
        if not before_t:
            continue  # 空行跳过
        table_changes.append(TableChange(
            before_table=before_t,
            after_table=_get_val(row, idx_after),
            is_ping_cut=_parse_bool(_get_val(row, idx_ping)),
            note=_get_val(row, idx_note),
            change_type=_get_val(row, idx_change_type),
        ))
    return table_changes


def _parse_sheet2(ws) -> list:
    """解析字段级 mapping Sheet"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    col_idx = {}
    for i, h in enumerate(header):
        if h is not None:
            col_idx[_safe_str(h)] = i

    idx_b_db = _find_col(col_idx, "切换前数据库")
    idx_b_schema = _find_col(col_idx, "切换前表schema")
    idx_b_table = _find_col(col_idx, "切换前表名")
    idx_b_field = _find_col(col_idx, "切换前表字段名")
    idx_b_field_cn = _find_col(col_idx, "切换前表字段中文名")
    idx_b_type = _find_col(col_idx, "切换前表字段类型")
    idx_a_db = _find_col(col_idx, "切换后数据库")
    idx_a_schema = _find_col(col_idx, "切换后表schema")
    idx_a_table = _find_col(col_idx, "切换后表名")
    idx_a_field = _find_col(col_idx, "切换后表字段名")
    idx_a_field_cn = _find_col(col_idx, "切换后表字段中文名")
    idx_a_type = _find_col(col_idx, "切换后表字段类型")
    idx_change = _find_col(col_idx, "字段变化类型")
    idx_recover = _find_col(col_idx, "是否可还原")
    idx_recovery = _find_col(col_idx, "还原方案详细说明")
    idx_it = _find_col(col_idx, "源端IT责任人")
    idx_biz = _find_col(col_idx, "源端业务责任人")

    field_changes = []
    for row in rows[1:]:
        before_t = _get_val(row, idx_b_table)
        before_f = _get_val(row, idx_b_field)
        # 前后都空 = 空行跳过
        if not before_t and not _get_val(row, idx_a_table):
            continue
        if not before_t and not before_f:
            continue
        try:
            field_changes.append(ChangeItem(
                before_db=_get_val(row, idx_b_db),
                before_schema=_get_val(row, idx_b_schema),
                before_table=before_t,
                before_field=before_f,
                before_field_cn=_get_val(row, idx_b_field_cn),
                before_type=_get_val(row, idx_b_type),
                after_db=_get_val(row, idx_a_db),
                after_schema=_get_val(row, idx_a_schema),
                after_table=_get_val(row, idx_a_table),
                after_field=_get_val(row, idx_a_field),
                after_field_cn=_get_val(row, idx_a_field_cn),
                after_type=_get_val(row, idx_a_type),
                change_type=_get_val(row, idx_change),
                recoverable=_get_val(row, idx_recover),
                recovery_plan=_get_val(row, idx_recovery),
                source_it_owner=_get_val(row, idx_it),
                source_biz_owner=_get_val(row, idx_biz),
            ))
        except Exception as e:
            print(f"[WARN] 字段行解析失败，已跳过: {e}", file=sys.stderr)
            continue
    return field_changes


def _parse_sheet3(ws) -> dict:
    """解析变动类型说明 Sheet → {类型: 说明}"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    idx_type = None
    idx_desc = None
    header = rows[0]
    for i, h in enumerate(header):
        if h is None:
            continue
        h_str = _safe_str(h)
        if "类型" in h_str:
            idx_type = i
        elif "说明" in h_str:
            idx_desc = i
    if idx_type is None:
        return {}
    type_dict = {}
    for row in rows[1:]:
        t = _get_val(row, idx_type)
        if not t:
            continue
        type_dict[t] = _get_val(row, idx_desc) if idx_desc is not None else ""
    return type_dict


# ═══════════════════════════════════════════════════════════════
# ② 传播层（纯函数，确定性，与判定分离）
# ═══════════════════════════════════════════════════════════════

def _norm_table(t: str) -> str:
    """表名归一化：小写 + 去 schema 前缀，用于匹配"""
    if not t:
        return ""
    t = t.strip().lower()
    # 去 schema 前缀（取最后的表名部分）
    if "." in t:
        t = t.split(".")[-1]
    return t


def _build_table_index(knowledge: dict) -> dict:
    """构建表名 → [step_info] 索引，用于表级过滤。

    返回: {norm_table: [{"step_id", "rule_code", "exec_sequence", "source_tables_from_sql"}, ...]}
    """
    index = {}
    topology = knowledge.get("topology", {})
    for s in topology.get("steps", []):
        for src_table in s.get("source_tables_from_sql", []):
            norm = _norm_table(src_table)
            if norm:
                index.setdefault(norm, []).append({
                    "step_id": s.get("step_id", ""),
                    "rule_code": s.get("rule_code", ""),
                    "exec_sequence": s.get("exec_sequence", 0),
                    "source_table_raw": src_table,
                    "is_view_step": s.get("is_view_step", False),
                    "is_exchange": s.get("is_exchange", False),
                    "exchange_temp_table": s.get("exchange_temp_table", ""),
                })
    return index


def _build_alias_resolver(knowledge: dict) -> dict:
    """构建 SQL别名 → 真实表名 映射。

    field_mappings.lineage 里的 source_table 是 SQL 别名（如 a, t1），
    需要映射回真实表名才能与变更清单匹配。

    从 data_flow 或 source 里取步骤的表别名信息。
    """
    resolver = {}
    # data_flow 里的 source_tables 可能带 alias 信息
    data_flow = knowledge.get("data_flow", {})
    for block in data_flow.get("blocks", []):
        for src in block.get("source_tables", []):
            alias = src.get("alias", "")
            table = src.get("table", "") or src.get("name", "")
            if alias and table:
                resolver[alias.lower()] = table
                resolver[alias] = table
    return resolver


def filter_and_propagate(
    table_changes: list,
    field_changes: list,
    knowledge: dict,
    type_dict: dict,
) -> AnalysisResult:
    """三层过滤 + 逐跳传播。

    流程:
      1. 表级过滤: 变更表是否在资产 source_tables 中
         - 按表级变化类型映射表判定（平切/下线/初始化/归档/权限等）
         - 表级处理过的表，字段层跳过
      2. 字段级过滤: 变更字段是否被步骤引用
      3. 传播: 沿 field_mappings.lineage 逐跳追踪到目标字段
    """
    result = AnalysisResult()
    table_index = _build_table_index(knowledge)
    alias_resolver = _build_alias_resolver(knowledge)

    # ── 表级处理 ──
    asset_tables = set(table_index.keys())
    processed_field_tables = set()  # 已在表级处理的表，字段层跳过

    for tc in table_changes:
        norm_before = _norm_table(tc.before_table)
        if not norm_before:
            continue
        if norm_before not in asset_tables:
            # 表不在资产中 → 进过滤摘要，不丢
            result.filtered_out.append({
                "status": "⚪未命中",
                "source_table": tc.before_table,
                "source_field": "",
                "change_type": tc.change_type or tc.derived_table_type,
                "reason": f"表 {tc.before_table} 不在本资产源表中",
            })
            continue

        # 确定表级变化类型：优先人工标注，再从平切/下线标志推导
        tct = tc.change_type
        if not tct:
            if tc.is_ping_cut:
                tct = "平切"
            elif tc.is_table_dropped:
                tct = "表/视图下线"

        # 查表级映射表
        mapped = _TABLE_SEVERITY_MAP.get(tct)
        if not mapped:
            # 未知类型：检查是否平切/下线（change_type 没填但标志位有）
            if tc.is_ping_cut:
                mapped = _TABLE_SEVERITY_MAP["平切"]
            elif tc.is_table_dropped:
                mapped = _TABLE_SEVERITY_MAP["表/视图下线"]
            else:
                # 有 change_type 但映射表没有 → 待确认
                mapped = None

        if mapped:
            status, severity, default_note, scope = mapped
            # 数据初始化类需联动 load_strategy
            if tct in _INIT_WITH_TS_TYPES or tct in _INIT_WITHOUT_TS_TYPES:
                status, severity, default_note = _assess_table_init(tct, knowledge)
            elif not status:
                # 映射表里 status 为空（理论不该走到，初始化已处理），兜底
                status, severity = "🟡待确认", "unknown"
        else:
            status, severity, default_note, scope = (
                "🟡待确认", "unknown",
                f"表级变更类型「{tct or '未标注'}」需人工确认",
                "all_fields" if tct else "ref_only",
            )

        steps_info = table_index.get(norm_before, [])
        touched_fields = []
        if scope == "all_fields":
            touched_fields = _find_fields_by_source_table(
                knowledge, tc.before_table, norm_before
            )

        result.table_level_impacts.append({
            "status": status,
            "type": tct or "未标注",
            "source_table": tc.before_table,
            "new_table": tc.after_table,
            "note": default_note,
            "original_note": tc.note,
            "steps": [s["step_id"] for s in steps_info],
            "rule_codes": list({s["rule_code"] for s in steps_info}),
            "touched_fields": touched_fields,
        })
        # 表级处理过的表，字段层跳过（除非是"名称变化"这类 ref_only，
        # 字段语义可能仍需检查——但 MVP 简化：表级已覆盖，字段层跳过）
        processed_field_tables.add(norm_before)

    # ── 字段级处理 ──
    for fc in field_changes:
        norm_table = _norm_table(fc.before_table)
        if not norm_table:
            continue
        if norm_table in processed_field_tables:
            # 已在表级处理，字段层跳过
            # 找到该表在 table_changes 里的 change_type 用于说明
            tc_type = ""
            for _tc in table_changes:
                if _norm_table(_tc.before_table) == norm_table:
                    tc_type = _tc.change_type or _tc.derived_table_type
                    break
            result.filtered_out.append({
                "status": "⚪跳过",
                "source_table": fc.before_table,
                "source_field": fc.before_field,
                "change_type": fc.derived_change_type,
                "reason": f"表已按「{tc_type or '表级变更'}」处理，字段级跳过",
            })
            continue
        if norm_table not in asset_tables:
            # 表不在资产中 → 未命中
            result.filtered_out.append({
                "status": "⚪未命中",
                "source_table": fc.before_table,
                "source_field": fc.before_field,
                "change_type": fc.derived_change_type,
                "reason": f"表 {fc.before_table} 不在本资产源表中",
            })
            continue

        # 字段级传播
        impact_paths = _propagate_field_change(
            fc, knowledge, table_index.get(norm_table, []), alias_resolver
        )

        if not impact_paths:
            # 字段没被任何步骤引用
            result.filtered_out.append({
                "status": "⚪未命中",
                "source_table": fc.before_table,
                "source_field": fc.before_field,
                "change_type": fc.derived_change_type,
                "reason": f"字段 {fc.before_field} 未被本资产引用",
            })
            continue

        for path in impact_paths:
            assess_severity(path, fc, knowledge, type_dict)
            # 分流：🟢无影响 + ⚪ → filtered_out；🔴🟡 → field_level_impacts
            if path.status == "🟢无影响":
                result.filtered_out.append({
                    "status": path.status,
                    "source_table": fc.before_table,
                    "source_field": fc.before_field,
                    "change_type": fc.derived_change_type,
                    "reason": path.reason,
                    "target_field": path.target_field,
                })
            elif path.status == "⚪未命中":
                result.filtered_out.append({
                    "status": path.status,
                    "source_table": fc.before_table,
                    "source_field": fc.before_field,
                    "change_type": fc.derived_change_type,
                    "reason": path.reason,
                })
            else:
                # 🔴有影响 / 🟡待确认
                result.field_level_impacts.append(_path_to_row(path))

    # ── 统计摘要 ──
    result.summary = {
        "total_field_changes": len(field_changes),
        "total_table_changes": len(table_changes),
        "impacted": len([r for r in result.field_level_impacts if r["status"] == "🔴有影响"]),
        "uncertain": len([r for r in result.field_level_impacts if r["status"] == "🟡待确认"]),
        "no_impact": len([r for r in result.filtered_out if r["status"] == "🟢无影响"]),
        "not_hit": len([r for r in result.filtered_out if "未命中" in r["status"] or "跳过" in r["status"]]),
        "table_level": len(result.table_level_impacts),
    }
    return result


def _assess_table_init(tct: str, knowledge: dict) -> tuple:
    """表级数据初始化的严重度判定（联动 load_strategy）。

    逻辑同 _assess_init：刷/不刷时间戳 × 全量/增量
    """
    strategy = knowledge.get("meta", {}).get("load_strategy", "")
    is_with_ts = tct in _INIT_WITH_TS_TYPES
    is_without_ts = tct in _INIT_WITHOUT_TS_TYPES

    if "full" in strategy.lower() or "全量" in strategy:
        return ("🟢无影响", "none",
                f"表级数据初始化({'刷时间戳' if is_with_ts else '不刷时间戳'})，"
                f"本资产为全量加载，无影响")
    if is_without_ts:
        return ("🔴有影响", "high",
                "表级数据初始化不刷时间戳 + 增量加载 = 下游增量可能漏掉此变更！")
    if is_with_ts:
        return ("🟡待确认", "unknown",
                "表级数据初始化刷时间戳 + 增量加载，会触发增量重拉，需确认")
    return ("🟡待确认", "unknown", "表级数据初始化，加载策略不明确，需人工确认")


def _find_fields_by_source_table(knowledge: dict, table: str, norm_table: str) -> list:
    """找出某个源表流向本资产的所有目标字段"""
    touched = []
    fm = knowledge.get("field_mappings", {})
    for f in fm.get("fields", []):
        for hop in f.get("lineage", []):
            # lineage.source_table 是别名，需额外匹配真实表
            # 这里用 step_id 反查 source_tables_from_sql
            step_id = hop.get("step", "")
            topology = knowledge.get("topology", {})
            step_info = None
            for s in topology.get("steps", []):
                if s.get("step_id") == step_id:
                    step_info = s
                    break
            if step_info:
                src_tables = step_info.get("source_tables_from_sql", [])
                for st in src_tables:
                    if _norm_table(st) == norm_table:
                        touched.append({
                            "target_field": f.get("target_field", ""),
                            "step": step_id,
                            "rule_code": f.get("rule_code", ""),
                        })
                        break
    return touched


def _propagate_field_change(
    fc: ChangeItem,
    knowledge: dict,
    step_infos: list,
    alias_resolver: dict,
) -> list:
    """对单个字段变更做逐跳传播。

    沿 field_mappings.fields[].lineage[] 追踪：
    - 找到引用了 (source_table, source_field) 的 lineage 跳
    - 从该跳出发，沿数据流追踪到最终目标字段
    - SELECT * 检测：如果步骤用了通配，标断链
    """
    impact_paths = []
    fm = knowledge.get("field_mappings", {})
    target_table = knowledge.get("meta", {}).get("target_table", "")
    norm_source_table = _norm_table(fc.before_table)

    for f in fm.get("fields", []):
        lineage = f.get("lineage", [])
        if not lineage:
            continue

        # 检查该字段的 lineage 是否引用了变更的源表+字段
        hit_hops = []
        for hop in lineage:
            hop_field = hop.get("source_field", "")
            hop_step = hop.get("step", "")

            # 解析 hop 的真实表名（lineage.source_table 是别名）
            hop_real_table = _resolve_hop_table(hop, hop_step, knowledge, norm_source_table)

            if hop_real_table != norm_source_table:
                continue

            # 字段匹配（大小写不敏感）
            if hop_field.lower() != fc.before_field.lower():
                continue

            hit_hops.append(hop)

        if not hit_hops:
            continue

        # 构建完整传播路径（从命中跳到目标字段）
        hops = []
        has_select_star = False
        for hop in lineage:
            expr = hop.get("raw_sql", "")
            # SELECT * 检测
            if _is_select_star_step(hop.get("step", ""), knowledge):
                has_select_star = True

            hops.append(Hop(
                step=hop.get("step", ""),
                source_table=hop.get("source_table", ""),
                source_field=hop.get("source_field", ""),
                expression=expr,
                output=f.get("target_field", ""),
                rule_code=f.get("rule_code", ""),
            ))

        reason_uncertain = ""
        if has_select_star:
            star_step = next(
                (h.step for h in hops if _is_select_star_step(h.step, knowledge)), ""
            )
            reason_uncertain = (
                f"步骤 {star_step} 使用了 SELECT *，无法枚举字段，"
                f"无法自动传播字段 {fc.before_field} 的变更"
            )

        impact_paths.append(ImpactPath(
            target_table=target_table,
            target_field=f.get("target_field", ""),
            change=fc,
            hops=hops,
            uncertain_reason=reason_uncertain,
        ))

    return impact_paths


def _resolve_hop_table(hop: dict, hop_step: str, knowledge: dict, target_norm: str) -> str:
    """解析 lineage 跳的真实表名。

    lineage.source_table 是 SQL 别名，反查 step 的 source_tables_from_sql。
    匹配策略：如果 step 的任一源表归一化后 == target_norm，则命中。
    （别名→真实表的精确映射依赖 data_flow，MVP 用 step 级宽匹配）
    """
    topology = knowledge.get("topology", {})
    for s in topology.get("steps", []):
        if s.get("step_id") == hop_step:
            for st in s.get("source_tables_from_sql", []):
                if _norm_table(st) == target_norm:
                    return target_norm
            break
    return ""


def _is_select_star_step(step_id: str, knowledge: dict) -> bool:
    """检查某步骤是否使用了 SELECT *（通配）"""
    source = knowledge.get("source", {})
    raw_sqls = source.get("raw_sql", [])
    sql = ""
    if isinstance(raw_sqls, list):
        for item in raw_sqls:
            if isinstance(item, dict) and item.get("step_id") == step_id:
                sql = item.get("sql", "")
                break
    elif isinstance(raw_sqls, dict):
        sql = raw_sqls.get(step_id, "")
    if not sql:
        return False
    sql_upper = sql.upper().replace("\n", " ")
    # 粗检测：SELECT * FROM
    import re
    if re.search(r"SELECT\s+\*", sql_upper):
        return True
    return False


def _path_to_row(path: ImpactPath) -> dict:
    """ImpactPath → Excel 行"""
    hops_desc = " → ".join(
        f"{h.rule_code}/{h.step}: {h.expression or h.source_field}"
        for h in path.hops
    )
    steps = ",".join(sorted({h.step for h in path.hops}))
    rule_codes = ",".join(sorted({h.rule_code for h in path.hops}))
    return {
        "status": path.status,
        "severity": path.severity,
        "target_table": path.target_table,
        "target_field": path.target_field,
        "source_table": path.change.before_table if path.change else "",
        "source_field": path.change.before_field if path.change else "",
        "change_type": path.change.derived_change_type if path.change else "",
        "before_type": path.change.before_type if path.change else "",
        "after_type": path.change.after_type if path.change else "",
        "reason": path.reason or path.uncertain_reason,
        "hops": hops_desc,
        "steps": steps,
        "rule_codes": rule_codes,
        "recovery_plan": path.change.recovery_plan if path.change else "",
    }


# ═══════════════════════════════════════════════════════════════
# ③ 判定层（映射表驱动，与传播分离）
# ═══════════════════════════════════════════════════════════════

# ── 字段级变化类型 → 默认状态/严重度 映射表 ──
# 这是数据驱动的：类型变了改表不改代码。
# 每项 = (状态, 严重度, 原因说明)
# 空状态("")表示需特殊处理（如类型变化需结合 DDL，见 _assess_type_change）
_FIELD_SEVERITY_MAP = {
    # ── 基数变化 ──
    "0:1新增字段": ("⚪未命中", "none", "新增字段，资产未引用"),
    "1:0废弃字段": ("🔴有影响", "high", "字段被废弃，资产取不到该数据"),
    "1:0解耦到其他表": ("🔴有影响", "high", "字段解耦到其他表，资产需改来源表"),
    # ── 1:1 变化 ──
    "1:1完全一致": ("🟢无影响", "none", "前后完全一致，仅名称变化，无实质影响"),
    "字段类型及长度变化": ("", "", ""),  # 需结合 DDL，见 _assess_type_change
    "1:1数据类型/长度变化": ("", "", ""),  # 旧写法兼容
    "字段值语义变化": ("🟡待确认", "unknown", "同名同类型但语义变化，需人工确认"),
    "字段名称变化": ("🟡待确认", "unknown", "字段重命名，需改引用名"),
    # ── 数据初始化（区分刷/不刷时间戳，联动 load_strategy）──
    # 刷时间戳：增量会重新拉取，全量无影响
    # 不刷时间戳：增量可能漏掉！全量无影响
    "字段数据初始化（刷时间戳）": ("", "", ""),  # 需联动 load_strategy，见 _assess_init
    "字段数据初始化（不刷时间戳）": ("", "", ""),  # 需联动 load_strategy，见 _assess_init
}

# ── 表级变化类型 → 默认状态/严重度/影响范围 映射表 ──
# impact_scope: "all_fields" = 列出该表所有受波及字段
#               "ref_only"   = 只列触点步骤（改表名/权限等，不波及字段语义）
#               "none"       = 不影响数据（权限取消等，看场景）
_TABLE_SEVERITY_MAP = {
    # ── 结构性变化 ──
    "表/视图下线": ("🔴有影响", "high", "来源消失", "all_fields"),
    "表/视图替换": ("🔴有影响", "high", "表被替换，需改来源表引用", "all_fields"),
    "表/视图主键变化": ("🔴有影响", "high", "主键变化影响 JOIN/去重逻辑", "all_fields"),
    "表/视图名称或者schema变化": ("🟡待确认", "unknown", "需改表名/schema引用+术+规则+调度", "ref_only"),
    "表/视图名称或者schema变化".lower(): ("🟡待确认", "unknown", "需改表名/schema引用+术+规则+调度", "ref_only"),
    # ── 平切（字段完全一致仅名称变化）──
    "平切": ("🟡待确认", "unknown", "平切：需替换表名引用及术/规则/调度依赖", "ref_only"),
    # ── 数据操作（区分刷/不刷时间戳，联动 load_strategy）──
    "表/视图数据初始化（刷时间戳）": ("", "", "", "all_fields"),  # 见 _assess_init
    "表/视图初始化（不刷时间戳）": ("", "", "", "all_fields"),     # 见 _assess_init
    "表/视图数据初始化（不刷时间戳）": ("", "", "", "all_fields"),  # 两种写法兼容
    # ── 其他 ──
    "表/视图数据归档": ("🟡待确认", "unknown", "数据被归档，需确认是否影响历史数据查询", "all_fields"),
    "表/视图取消权限": ("🔴有影响", "high", "权限取消，资产将无法访问该表", "all_fields"),
    "表/视图数据硬删除": ("🔴有影响", "high", "数据被硬删除，来源数据消失", "all_fields"),
}

# 刷/不刷时间戳的类型识别（字段级 + 表级）
_INIT_WITH_TS_TYPES = {
    "字段数据初始化（刷时间戳）",
    "表/视图数据初始化（刷时间戳）",
}
_INIT_WITHOUT_TS_TYPES = {
    "字段数据初始化（不刷时间戳）",
    "表/视图初始化（不刷时间戳）",
    "表/视图数据初始化（不刷时间戳）",
}


def assess_severity(path: ImpactPath, fc: ChangeItem, knowledge: dict, type_dict: dict):
    """影响判定：在传播结果上判定严重度。

    判定依据:
      1. SELECT * 断链 → 待确认（优先）
      2. 数据初始化类型 → 联动 load_strategy（刷/不刷时间戳）
      3. 类型/长度变化 → 结合目标 DDL（cast 吸收 vs 截断）
      4. 其他 → 查字段级映射表
    """
    # 1. SELECT * 断链优先
    if path.uncertain_reason:
        path.status = "🟡待确认"
        path.severity = "unknown"
        path.reason = path.uncertain_reason
        return

    ct = fc.derived_change_type

    # 2. 数据初始化 → 联动 load_strategy
    if ct in _INIT_WITH_TS_TYPES or ct in _INIT_WITHOUT_TS_TYPES:
        _assess_init(path, fc, knowledge)
        return

    # 3. 类型/长度变化 → 结合 DDL
    if ct in ("字段类型及长度变化", "1:1数据类型/长度变化"):
        _assess_type_change(path, fc, knowledge)
        return

    # 4. 查字段级映射表
    mapped = _FIELD_SEVERITY_MAP.get(ct)
    if mapped:
        path.status, path.severity, path.reason = mapped
        return

    # 未知类型 → 待确认
    path.status = "🟡待确认"
    path.severity = "unknown"
    path.reason = f"未识别的变化类型: {ct}，需人工确认"


def _assess_type_change(path: ImpactPath, fc: ChangeItem, knowledge: dict):
    """类型/长度变化的严重度判定。

    判定逻辑（类型兼容性链）：
      源新类型(after_type) →[沿途cast]→ 中间类型 → 目标DDL类型

    最终看"源新类型能否安全流入目标DDL"：
      1. 类型大类变了（int↔varchar）→ 🔴必须适配
      2. 同大类，长度/精度兼容（目标≥源）→ 🟢无影响
      3. 同大类，长度/精度不兼容（目标<源）→ 🔴截断/溢出风险
      4. cast 的目标类型参与判定（作为链上一个关口）

    比较双方：源新类型(after_type) vs 目标DDL(field_type)，
    沿途 cast 的目标类型作为"关口"也参与。
    """
    after_type = fc.after_type
    target_type = _get_target_field_type(path.target_field, knowledge)

    # 收集沿途 cast 的目标类型（作为兼容性链上的关口）
    cast_types = _extract_cast_types(path.hops)

    # 没有目标 DDL，也没有 cast → 无法判定
    if not target_type and not cast_types:
        path.status = "🟡待确认"
        path.severity = "unknown"
        path.reason = (f"源类型 {fc.before_type}→{after_type}，"
                       f"目标字段无 DDL 且无 cast，需人工确认")
        return

    # 解析各类型
    after_info = _parse_type_info(after_type)
    target_info = _parse_type_info(target_type) if target_type else None
    cast_infos = [_parse_type_info(ct) for ct in cast_types]

    # ── 第一步：类型大类判定 ──
    # 源新类型 vs 目标DDL vs 每个cast关口，只要有一个大类不匹配 → 有影响
    compare_targets = []
    if target_info:
        compare_targets.append(("目标DDL", target_info))
    for i, ci in enumerate(cast_infos):
        compare_targets.append((f"cast关口{i+1}", ci))

    for label, t_info in compare_targets:
        if not _same_type_family(after_info["family"], t_info["family"]):
            path.status = "🔴有影响"
            path.severity = "high"
            path.reason = (f"源新类型 {after_type}({after_info['family']}) 与{label} "
                           f"{t_info['raw']}({t_info['family']}) 类型大类不一致，必须适配")
            return

    # ── 第二步：同大类，长度/精度兼容性判定 ──
    # 取链上最窄的关口（目标DDL + 所有cast），看源新类型能否通过
    for label, t_info in compare_targets:
        compat = _check_length_compat(after_info, t_info)
        if compat == "incompatible":
            path.status = "🔴有影响"
            path.severity = "high"
            path.reason = (f"源新类型 {after_type} 长度/精度超出{label} {t_info['raw']}，"
                           f"有截断/溢出风险，需适配")
            return

    # 所有关口都兼容
    gate_desc = "、".join(f"{label}={t_info['raw']}" for label, t_info in compare_targets)
    path.status = "🟢无影响"
    path.severity = "none"
    path.reason = (f"源新类型 {after_type} 可被{gate_desc}兼容，无影响")


# ── 类型解析工具 ──

# 类型大类归一化映射：各种写法 → 标准家族名
_TYPE_FAMILY_MAP = {
    # 整数家族
    "int": "integer", "integer": "integer", "bigint": "integer",
    "smallint": "integer", "tinyint": "integer", "int2": "integer",
    "int4": "integer", "int8": "integer", "serial": "integer",
    # 字符家族
    "varchar": "varchar", "character": "varchar", "char": "varchar",
    "text": "varchar", "string": "varchar", "nvarchar": "varchar",
    "nvarchar2": "varchar", "varchar2": "varchar",
    # 数值家族（带小数）
    "numeric": "numeric", "decimal": "numeric", "number": "numeric",
    "float": "numeric", "double": "numeric", "real": "numeric",
    "float4": "numeric", "float8": "numeric", "double": "numeric",
    "precision": "numeric",
    # 日期时间家族
    "date": "datetime", "timestamp": "datetime", "time": "datetime",
    "datetime": "datetime",
    # 布尔
    "boolean": "boolean", "bool": "boolean",
}


def _parse_type_info(type_str: str) -> dict:
    """解析类型字符串为结构化信息。

    返回: {family, raw, length, precision, scale}
    - family: 归一化大类（integer/varchar/numeric/datetime/boolean/unknown）
    - raw: 原始类型字符串
    - length: 长度（varchar 的 n，或 numeric 的 precision）
    - scale: 小数位数（numeric 的 scale）
    """
    import re

    if not type_str:
        return {"family": "unknown", "raw": "", "length": None, "scale": None}

    raw = type_str.strip()
    lower = raw.lower()

    # 去掉 "character varying" 的空格 → "charactervarying"，取第一个词
    # 处理两段式类型名
    lower_compact = lower.replace("varying", "").strip()
    # 提取类型名（括号前的部分，可能含空格如 "character varying"）
    type_name_match = re.match(r'^([a-zA-Z][a-zA-Z\s]*?)(?:\s*\(|\s*$)', lower)
    if type_name_match:
        base_name = type_name_match.group(1).strip()
        # 两段式：character varying → varchar 家族
        if "character" in base_name and "varying" in lower:
            base_name = "varchar"
        elif base_name == "character":
            base_name = "char"
    else:
        base_name = lower.split("(")[0].split()[0] if lower.split() else "unknown"

    family = _TYPE_FAMILY_MAP.get(base_name, "unknown")

    # 提取括号内参数
    length = None
    scale = None
    param_match = re.search(r'\(([^)]*)\)', lower)
    if param_match:
        params = [p.strip() for p in param_match.group(1).split(",")]
        if params:
            try:
                length = int(params[0])
            except (ValueError, TypeError):
                pass
        if len(params) > 1:
            try:
                scale = int(params[1])
            except (ValueError, TypeError):
                pass

    return {"family": family, "raw": raw, "length": length, "scale": scale,
            "base_name": base_name}


def _same_type_family(f1: str, f2: str) -> bool:
    """两个类型大类是否兼容（同家族或在安全转换范围内）。

    安全跨类转换：integer → numeric（整数可安全转数值，不丢精度）
    """
    if f1 == f2:
        return True
    # integer → numeric 是安全的（整数可以精确表示为数值）
    if {f1, f2} == {"integer", "numeric"}:
        return True
    return False


def _check_length_compat(source: dict, target: dict) -> str:
    """检查长度/精度兼容性。

    返回: "compatible" | "incompatible" | "unknown"
    - 目标无长度信息 → unknown（不阻断，由其他关口判）
    - 目标长度 ≥ 源长度 → compatible
    - 目标长度 < 源长度 → incompatible
    """
    # 目标没长度信息（如 text 类型）→ 不限制，兼容
    if target["length"] is None:
        return "compatible"
    # 源没长度信息（如源端没填）→ 无法判，放行（标 unknown 但不阻断）
    if source["length"] is None:
        return "unknown"

    # 字符家族：比长度
    if source["family"] == "varchar" and target["family"] == "varchar":
        if target["length"] >= source["length"]:
            return "compatible"
        return "incompatible"

    # 数值家族：比精度+标度
    if source["family"] == "numeric" and target["family"] == "numeric":
        # 精度比较
        if target["length"] is not None and source["length"] is not None:
            if target["length"] < source["length"]:
                return "incompatible"
        # 标度比较（目标标度不能小于源标度，否则小数位丢失）
        if target["scale"] is not None and source["scale"] is not None:
            if target["scale"] < source["scale"]:
                return "incompatible"
        return "compatible"

    # 整数家族：一般不比长度（int/bigint 已在大类映射里区分了）
    # integer → numeric 的跨类安全转换，长度兼容性单独处理
    if source["family"] == "integer" and target["family"] == "numeric":
        # 整数转数值，目标精度要能容纳整数位数
        if target["length"] is not None and source["length"] is not None:
            if target["length"] < source["length"]:
                return "incompatible"
        return "compatible"

    # 整数家族同类：不比长度（bigint/int 差异由大类决定）
    if source["family"] == "integer" and target["family"] == "integer":
        return "compatible"

    return "compatible"


def _extract_cast_types(hops: list) -> list:
    """从传播路径的 hop 表达式里提取 cast 的目标类型。

    匹配 cast(xxx as TYPE) 或 convert(xxx, TYPE) 里的 TYPE。
    """
    import re

    cast_types = []
    for hop in hops:
        expr = hop.expression or ""
        # cast(field as bigint) / cast(field as varchar(50))
        for m in re.finditer(r'cast\s*\([^)]*?\s+as\s+([a-zA-Z][a-zA-Z0-9\s]*(?:\([^)]*\))?)', expr, re.IGNORECASE):
            cast_types.append(m.group(1).strip())
        # convert(type, field) — GaussDB/SQL Server 语法
        for m in re.finditer(r'convert\s*\(\s*([a-zA-Z][a-zA-Z0-9\s]*(?:\([^)]*\))?)\s*,', expr, re.IGNORECASE):
            cast_types.append(m.group(1).strip())
    return cast_types


def _assess_init(path: ImpactPath, fc: ChangeItem, knowledge: dict):
    """数据初始化变化的严重度判定（联动 load_strategy）。

    核心逻辑：刷/不刷时间戳 × 全量/增量加载 = 不同风险

    |              | 全量(TRUNCATE+INSERT) | 增量(依赖时间戳)         |
    |-------------|----------------------|-------------------------|
    | 刷时间戳      | 🟢 无影响             | 🟡 会触发重拉，需确认     |
    | 不刷时间戳    | 🟢 无影响             | 🔴 增量会漏掉！高风险     |

    这是 detect_load_strategy 能力的自然延伸。
    """
    ct = fc.derived_change_type
    strategy = knowledge.get("meta", {}).get("load_strategy", "")

    is_with_ts = ct in _INIT_WITH_TS_TYPES
    is_without_ts = ct in _INIT_WITHOUT_TS_TYPES

    # 全量加载：不管刷不刷时间戳都无影响（每次全量拉）
    if "full" in strategy.lower() or "全量" in strategy:
        path.status = "🟢无影响"
        path.severity = "none"
        ts_label = "刷时间戳" if is_with_ts else "不刷时间戳"
        path.reason = f"数据初始化({ts_label})，本资产为全量加载，每次全量拉取，无影响"
        return

    # 增量加载
    if is_without_ts:
        # 不刷时间戳 + 增量 = 高风险！
        path.status = "🔴有影响"
        path.severity = "high"
        path.reason = (
            "数据初始化不刷时间戳 + 增量加载 = 下游增量可能漏掉此变更！"
            "需确认增量逻辑是否依赖时间戳"
        )
    elif is_with_ts:
        # 刷时间戳 + 增量 = 会触发重拉，行为变化
        path.status = "🟡待确认"
        path.severity = "unknown"
        path.reason = "数据初始化刷时间戳 + 增量加载，会触发增量重拉，需确认数据正确性"

    else:
        path.status = "🟡待确认"
        path.severity = "unknown"
        path.reason = "数据初始化，加载策略不明确，需人工确认"


def _get_target_field_type(field_name: str, knowledge: dict) -> str:
    """从 field_mappings 获取目标字段类型（DDL 下注）"""
    fm = knowledge.get("field_mappings", {})
    for f in fm.get("fields", []):
        if f.get("target_field", "").lower() == field_name.lower():
            return f.get("field_type", "") or ""
    return ""


# ═══════════════════════════════════════════════════════════════
# ④ 渲染层（Excel 三 Sheet）
# ═══════════════════════════════════════════════════════════════

def render_excel(result: AnalysisResult, output_path: str, asset_name: str = ""):
    """渲染影响分析 Excel（三 Sheet）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # 颜色填充
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    gray_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    header_font = Font(bold=True)

    def _fill_for(status):
        if "🔴" in status:
            return red_fill
        if "🟡" in status:
            return yellow_fill
        if "🟢" in status:
            return green_fill
        return gray_fill

    # ── Sheet1: 影响清单（主表，🔴🟡）──
    ws1 = wb.active
    ws1.title = "影响清单"
    headers1 = ["状态", "严重度", "目标表", "目标字段", "源表", "源字段",
                "变化类型", "变化前类型", "变化后类型", "说明", "传播路径",
                "涉及步骤", "规则编码", "还原方案"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
    for row in result.field_level_impacts:
        ws1.append([row.get(h, "") for h in
                    ["status", "severity", "target_table", "target_field",
                     "source_table", "source_field", "change_type",
                     "before_type", "after_type", "reason", "hops",
                     "steps", "rule_codes", "recovery_plan"]])
        # 行着色
        fill = _fill_for(row.get("status", ""))
        for cell in ws1[ws1.max_row]:
            cell.fill = fill
    _auto_width(ws1)

    # ── Sheet2: 表级影响（含平切/下线）──
    ws2 = wb.create_sheet("表级影响")
    headers2 = ["状态", "类型", "源表", "切换后表", "影响说明", "涉及步骤", "规则编码", "受影响字段数"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
    for row in result.table_level_impacts:
        touched_count = len(row.get("touched_fields", []))
        ws2.append([
            row.get("status", ""), row.get("type", ""),
            row.get("source_table", ""), row.get("new_table", ""),
            row.get("note", ""),
            ",".join(row.get("steps", [])),
            ",".join(row.get("rule_codes", [])),
            touched_count if touched_count else "",
        ])
        fill = _fill_for(row.get("status", ""))
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
    _auto_width(ws2)

    # ── Sheet3: 过滤摘要（🟢无影响 + ⚪未命中）──
    ws3 = wb.create_sheet("过滤摘要")
    headers3 = ["状态", "源表", "源字段", "变化类型", "原因", "目标字段"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
    for row in result.filtered_out:
        ws3.append([row.get(h, "") for h in
                    ["status", "source_table", "source_field",
                     "change_type", "reason", "target_field"]])
        fill = _fill_for(row.get("status", ""))
        for cell in ws3[ws3.max_row]:
            cell.fill = fill
    _auto_width(ws3)

    # ── Sheet0: 统计摘要（放最后创建不影响顺序，放第一）──
    ws0 = wb.create_sheet("统计摘要", 0)
    ws0.append(["关联影响分析统计", asset_name or ""])
    ws0.append([])
    ws0.append(["指标", "数量"])
    for cell in ws0[3]:
        cell.font = header_font
    s = result.summary
    ws0.append(["字段变更总数", s.get("total_field_changes", 0)])
    ws0.append(["表级变更总数", s.get("total_table_changes", 0)])
    ws0.append([])
    ws0.append(["🔴 有影响", s.get("impacted", 0)])
    ws0.append(["🟡 待确认", s.get("uncertain", 0)])
    ws0.append(["表级影响", s.get("table_level", 0)])
    ws0.append([])
    ws0.append(["🟢 无影响（已过滤）", s.get("no_impact", 0)])
    ws0.append(["⚪ 未命中/跳过", s.get("not_hit", 0)])
    _auto_width(ws0)

    wb.save(output_path)


def _auto_width(ws):
    """自适应列宽：按列头语义 + 内容长度，设合理宽度，长文本列自动换行。

    策略：
      - 短内容列（状态/严重度/类型等）：取内容最大长度 + padding
      - 长文本列（说明/原因/路径/还原方案）：限宽 40-45，开自动换行
      - 中等内容列（表名/字段名/变化类型）：取内容长度，上限 25
    """
    # 长文本列名（内容可能很长，限宽+换行）
    LONG_TEXT_COLS = {"说明", "影响说明", "原因", "传播路径", "还原方案", "还原方案详细说明", "说明"}
    # 中等内容列名（表名/字段名等，上限 25）
    MEDIUM_COLS = {"目标表", "源表", "切换后表", "变化类型", "目标字段", "源字段",
                   "切换前表名", "切换后表名", "涉及步骤", "规则编码", "去哪看"}
    # 短内容列名（固定窄列）
    SHORT_COLS = {"状态", "严重度", "类型", "是否平切", "受影响字段数", "序号",
                  "是否可还原(Y/N)", "命中源数"}

    from openpyxl.utils import get_column_letter

    for col_cells in ws.columns:
        if not col_cells:
            continue
        col_letter = col_cells[0].column_letter
        header = str(col_cells[0].value or "")

        # 计算内容最大长度（考虑换行取最长行）
        max_len = len(header)
        for cell in col_cells[1:]:
            try:
                val = str(cell.value or "")
                longest_line = max((len(line) for line in val.split("\n")), default=0)
                # 中文按2算宽度
                cn_count = sum(1 for c in val if '\u4e00' <= c <= '\u9fff')
                adjusted = longest_line + cn_count
                max_len = max(max_len, adjusted)
            except Exception:
                pass

        if header in LONG_TEXT_COLS:
            width = min(max(max_len + 2, 20), 45)
            # 长文本列开自动换行
            for cell in col_cells[1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        elif header in MEDIUM_COLS:
            width = min(max(max_len + 2, 12), 25)
        elif header in SHORT_COLS:
            width = max(min(max_len + 4, 18), 8)
        else:
            # 兜底：内容长度 + padding，上限 30
            width = min(max(max_len + 2, 10), 30)

        ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════════
# CLI 入口
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="关联影响分析（单资产 MVP）"
    )
    parser.add_argument("--changes", required=True,
                        help="变更清单 Excel（三 Sheet 模板）")
    parser.add_argument("--knowledge", required=True,
                        help="资产 knowledge JSON（analyzer 产出）")
    parser.add_argument("--output", default="impact.xlsx",
                        help="输出 Excel 路径（默认 impact.xlsx）")
    parser.add_argument("--asset", default="",
                        help="资产名/规则组编码（用于报告标题）")
    args = parser.parse_args()

    # 读取 knowledge
    knowledge_path = Path(args.knowledge)
    if not knowledge_path.exists():
        print(f"[ERROR] knowledge 文件不存在: {args.knowledge}", file=sys.stderr)
        sys.exit(1)
    with open(knowledge_path, "r", encoding="utf-8") as f:
        knowledge = json.load(f)

    # 读取变更清单
    table_changes, field_changes, type_dict = read_changes(args.changes)
    print(f"[INFO] 读取变更清单: {len(table_changes)} 表级, {len(field_changes)} 字段级, {len(type_dict)} 类型")

    # 分析
    result = filter_and_propagate(table_changes, field_changes, knowledge, type_dict)

    # 渲染
    render_excel(result, args.output, args.asset)
    print(f"[INFO] 影响分析完成 → {args.output}")
    s = result.summary
    print(f"  🔴 有影响: {s.get('impacted', 0)}")
    print(f"  🟡 待确认: {s.get('uncertain', 0)}")
    print(f"  🟢 无影响: {s.get('no_impact', 0)}")
    print(f"  ⚪ 未命中: {s.get('not_hit', 0)}")
    print(f"  表级影响: {s.get('table_level', 0)}")


if __name__ == "__main__":
    main()
