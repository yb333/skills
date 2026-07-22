#!/usr/bin/env python3
"""
dws-pipeline-analyzer view_generator — 视图生成器
从 knowledge_draft.json 生成多种输出视图（AI 增强结果通过 --ai-input 注入）。

Usage:
    python run.py view_generator \
        --input knowledge_draft.json \
        --output docs/output/{target_table}/ \
        [--ai-input knowledge_ai.md] \
        [--views mapping,asset,techspec]

支持的视图:
    mapping   → mapping.xlsx        (实体级+属性级字段映射)
    asset     → asset_report.html   (资产说明书，交互式 HTML)
    techspec  → tech_design.md      (技术设计文档)

默认: all (生成全部视图)
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
import re

# ── 工具函数 ──────────────────────────────────────────────

def _clean(s):
    """清洗字符串"""
    if s is None:
        return ""
    return str(s).strip()


def _norm(name: str) -> str:
    """表名归一化（统一小写）。所有表名比较都必须走这个函数。"""
    if not name:
        return ""
    return name.strip().lower()


# 加工类型优先级（从轻到重）。HTML 视图和 Excel mapping 共用此口径，避免不一致。
TRANSFORM_PRIORITY = {
    "unknown": -1, "direct": 0, "value": 1, "fallback": 2,
    "case_when": 3, "expression": 4, "aggregate": 5, "pivot": 6, "window": 7,
}


def _describe_transform(transform_type: str, expression: str = "", field: str = "") -> str:
    """把加工类型 + 表达式翻译成人能看懂的自然语言描述。

    用于 mapping 的映射规则列（替代生硬的 SQL 表达式）。
    """
    tt = transform_type or "unknown"
    expr = (expression or "").strip()
    expr_upper = expr.upper()

    if tt == "direct":
        return "直接映射"
    if tt == "value":
        return "赋值：固定值" if "'" not in expr else f"赋值：固定值 {expr}"
    if tt == "aggregate":
        if "SUM" in expr_upper:
            return f"聚合：对 {field or '字段'} 求和"
        if "COUNT" in expr_upper:
            return "聚合：计数"
        if "AVG" in expr_upper:
            return f"聚合：对 {field or '字段'} 求平均"
        if "MAX" in expr_upper:
            return f"聚合：取 {field or '字段'} 最大值"
        if "MIN" in expr_upper:
            return f"聚合：取 {field or '字段'} 最小值"
        return "聚合"
    if tt == "case_when":
        return "条件取值"
    if tt == "pivot":
        return "行转列"
    if tt == "window":
        if "ROW_NUMBER" in expr_upper:
            return "窗口：排序编号"
        if "RANK" in expr_upper or "DENSE_RANK" in expr_upper:
            return "窗口：排名"
        if "LAG" in expr_upper:
            return "窗口：取上一行"
        if "LEAD" in expr_upper:
            return "窗口：取下一行"
        return "窗口函数"
    if tt == "fallback":
        if "COALESCE" in expr_upper or "NVL" in expr_upper:
            return f"兜底：{field or '字段'} 为空取默认值"
        return "兜底"
    if tt == "expression":
        if "||" in expr or "CONCAT" in expr_upper:
            return "拼接字段"
        if "SUBSTR" in expr_upper or "SUBSTRING" in expr_upper:
            return f"截取：{field or '字段'}"
        if "CAST" in expr_upper:
            return f"类型转换"
        if "+" in expr or "-" in expr or "*" in expr or "/" in expr:
            return "算术加工"
        return "表达式加工"
    return "加工"


# 临时表/中间表判断（与 analyzer._is_intermediate_table 同款逻辑）
_INTERMEDIATE_TBL_RE = __import__("re").compile(
    r"(?:^tmp\d*$|_tmp\d*$|^temp\d*$|_temp\d*$|^tmp_|_tmp_|^temp_|_temp_)",
    __import__("re").IGNORECASE,
)


def _is_intermediate_tbl(table_name: str) -> bool:
    """判断表名是否为临时表/中间表。"""
    if not table_name:
        return False
    short = _norm(table_name).split(".")[-1]
    return bool(_INTERMEDIATE_TBL_RE.search(short))


# 写入方式标签
_DELETE_MODE_LABEL = {
    "0": "追加写入", "1": "覆盖写入", "2": "清空后写入",
    "3": "按条件删除后写入", "4": "增量写入",
}


def _merge_block_purposes(blocks, block_purposes, depth=0):
    """把 AI 生成的块目的合并到 data_blocks。

    block_purposes 格式: {"块1": "目的描述", "块1.1": "子块目的", ...}
    块编号规则: 块N（顶层），块N.M（子块），块N.M.K（更深层）
    """
    if not block_purposes:
        return blocks

    def _merge_recursive(blocks_list, parent_prefix, depth):
        for idx, blk in enumerate(blocks_list):
            block_id = f"块{parent_prefix}{idx+1}" if parent_prefix else f"块{idx+1}"
            # 尝试匹配（块1 / 块1 (主表 xxx) 等格式）
            matched_purpose = None
            for key, val in block_purposes.items():
                if key.startswith(block_id):
                    matched_purpose = val
                    break
            if matched_purpose:
                blk["purpose"] = matched_purpose
            # 递归子块
            if blk.get("children"):
                child_prefix = f"{parent_prefix}{idx+1}." if parent_prefix else f"{idx+1}."
                _merge_recursive(blk["children"], child_prefix, depth + 1)

    _merge_recursive(blocks, "", 0)
    return blocks


def _build_step_summary_inline(topo_step, df_step, fields_list):
    """生成结构化步骤概述（自然语言，精简归类）。

    格式: 从<主表>，过滤<条件>。关联<从表>带出<字段>。<写入方式><目标表>
    主表自带字段省略，只强调从表带出和加工字段。
    """
    joins = df_step.get("joins", [])
    where_clause = (df_step.get("where_clause", "") or "").replace("WHERE ", "").strip()
    target_table = topo_step.get("target_table_full", "") or topo_step.get("target_table", "")
    dm = (topo_step.get("delete_mode", "") or "").strip()
    delete_label = _DELETE_MODE_LABEL.get(dm, "写入")
    rule_type = topo_step.get("rule_type", 1)
    step_id = topo_step.get("step_id", "")

    # 分离主表和从表
    main_table = ""
    main_alias = ""
    secondary_tables = []
    for j in joins:
        if j.get("source_table", "").startswith("(subquery:"):
            continue
        jt = (j.get("join_type", "") or "").upper()
        if jt in ("FROM", "FROM_SUBQUERY_MAIN"):
            main_table = j.get("source_table", "")
            main_alias = j.get("alias", "")
        else:
            secondary_tables.append({
                "table": j.get("source_table", ""),
                "alias": j.get("alias", ""),
            })

    parts = []
    main_short = main_table.split(".")[-1] if main_table else ""

    if rule_type == 9:
        parts.append("交换分区数据")
    elif main_short:
        if where_clause:
            parts.append(f"从 {main_table}，过滤 {where_clause}。")
        else:
            parts.append(f"从 {main_table}。")

    # 别名→表名映射
    alias_map = {}
    for j in joins:
        if j.get("alias") and j.get("source_table"):
            alias_map[j["alias"].upper()] = j["source_table"]

    # 从表关联带出字段
    for sec in secondary_tables:
        sec_alias = sec["alias"].upper()
        sec_short = sec["table"].split(".")[-1] if sec["table"] else ""
        brought = []
        for f in fields_list:
            if f.get("producing_step") != step_id:
                continue
            for l in f.get("lineage", []):
                if (l.get("source_table", "") or "").upper() == sec_alias:
                    brought.append(f["target_field"])
                    break
        field_str = "、".join(dict.fromkeys(brought)) if brought else "字段"
        parts.append(f"关联 {sec_short} 带出 {field_str}；")

    # 加工归类
    processing_types = set()
    for f in fields_list:
        if f.get("producing_step") != step_id:
            continue
        tt = f.get("transform_type", "direct")
        if tt not in ("direct", "unknown"):
            processing_types.add(tt)
    if processing_types:
        type_descs = [_describe_transform(tt, "", "") for tt in sorted(processing_types)]
        parts.append("包含" + "、".join(type_descs) + "。")

    if rule_type == 9:
        parts.append(f"交换至 {target_table}")
    else:
        parts.append(f"{delete_label} {target_table}")

    return " ".join(parts)


def _resolve_on_condition_aliases(on_condition, alias_map, join_key_lineage):
    """把 ON 条件里的中间表别名替换成物理源表，并标注传递路径。

    例: "t.bid = d.bid"，t 是中间表(tmp1)，bid 追溯到 tbl_b
        → "tbl_b.bid = d.bid（经tmp1传递）"

    Args:
        on_condition: ON 条件字符串（如 "t.bid = d.bid"）
        alias_map: {别名(UPPER): 物理表名}，本步骤的 FROM/JOIN 别名映射
        join_key_lineage: {field_lower: [追溯链]}，关联键追溯数据
    Returns: (改写后的ON条件, 传递路径标注)
    """
    import re as _re
    if not on_condition:
        return on_condition, ""

    rewritten = on_condition
    transfers = []  # 传递路径标注

    # 提取所有 alias.field 模式
    for m in _re.finditer(r'(\w+)\.(\w+)', on_condition):
        alias = m.group(1)
        field = m.group(2)
        table = alias_map.get(alias.upper(), "")
        if not table or not _is_intermediate_tbl(table):
            continue
        # 这个别名是中间表，追溯它的 field 到物理源表
        trace_chains = join_key_lineage.get(field.lower(), [])
        if not trace_chains:
            continue
        chain = trace_chains[0]
        # 找物理源表（叶节点）
        leaves = []
        def _find_leaves(n):
            if not n.get("children"):
                leaves.append(n)
            for c in n.get("children", []):
                _find_leaves(c)
        _find_leaves(chain)
        if not leaves:
            continue
        # 取第一个叶节点的物理表短名
        phys_table = leaves[0].get("table", "").split(".")[-1]
        phys_field = leaves[0].get("field", field)
        # 替换别名：t.bid → tbl_b.bid
        old_ref = f"{alias}.{field}"
        new_ref = f"{phys_table}.{phys_field}"
        rewritten = rewritten.replace(old_ref, new_ref)
        transfers.append(f"{old_ref}经{table.split('.')[-1]}来自{new_ref}")

    transfer_note = "（" + "；".join(transfers) + "）" if transfers else ""
    return rewritten, transfer_note


def _merge_ai_markdown(knowledge: dict, ai_text: str) -> None:
    """解析 AI 输出的自然语言 markdown，合并到 knowledge 的 business_logic。

    AI 输出格式:
        # 整体描述
        （描述文字）

        ## step_1
        （这步的描述）

        ## step_2
        ...

        ## 关键字段
        - 字段名: 含义
    """
    bl = knowledge.setdefault("business_logic", {})
    if "step_descriptions" not in bl:
        bl["step_descriptions"] = []
    if "key_transforms" not in bl:
        bl["key_transforms"] = []

    # 去掉 markdown 代码块标记
    text = ai_text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:markdown)?\n?", "", text)
        text = re.sub(r"\n?```$", "", text)

    # 按 ## 分段
    sections = re.split(r"^## ", text, flags=re.MULTILINE)

    for section in sections:
        section = section.strip()
        if not section:
            continue

        # 第一行是标题
        lines = section.split("\n", 1)
        title = lines[0].strip().lower()
        content = lines[1].strip() if len(lines) > 1 else ""

        # 去掉标题行的 # 前缀
        if title.startswith("# "):
            title = title[2:]

        if "整体描述" in title or title.startswith("# 整体"):
            bl["summary"] = content

        elif title.startswith("step_") or re.match(r"step_\d+", title):
            # 提取 step_id
            step_match = re.match(r"(step_\d+)", title)
            if step_match:
                step_id = step_match.group(1)

                # 分离块目的和步骤描述
                block_purposes = {}  # {块名: 目的}
                step_content_lines = []
                in_block_section = False
                for cl in content.split("\n"):
                    cl_stripped = cl.strip()
                    if cl_stripped.startswith("### 块目的"):
                        in_block_section = True
                        continue
                    if cl_stripped.startswith("### ") and in_block_section:
                        in_block_section = False
                    if in_block_section and (cl_stripped.startswith("- ") or cl_stripped.startswith("* ")):
                        # 解析 "- 块1 (xxx): 目的描述"
                        bp_match = re.match(r"[-*]\s*(块\d+[^:]*?):\s*(.*)", cl_stripped)
                        if bp_match:
                            block_purposes[bp_match.group(1).strip()] = bp_match.group(2).strip()
                            continue
                    if not in_block_section:
                        step_content_lines.append(cl)

                step_content = "\n".join(step_content_lines).strip()

                # 找已有的 step_description
                desc = next((d for d in bl["step_descriptions"] if d.get("step_id") == step_id), None)
                if desc:
                    desc["purpose"] = step_content.split("\n")[0] if step_content else desc.get("purpose", "")
                    desc["logic"] = step_content
                    desc["is_auto_generated"] = False
                else:
                    bl["step_descriptions"].append({
                        "step_id": step_id,
                        "purpose": step_content.split("\n")[0] if step_content else "",
                        "logic": step_content,
                        "is_auto_generated": False,
                    })

                # 存块目的到 step_descriptions（供 data_blocks 合并用）
                if block_purposes:
                    desc = next((d for d in bl["step_descriptions"] if d.get("step_id") == step_id), None)
                    if desc:
                        desc["block_purposes"] = block_purposes

        elif "关键字段" in title:
            # 解析 - 字段名: 含义
            for line in content.split("\n"):
                line = line.strip()
                if line.startswith("- ") or line.startswith("* "):
                    parts = line[2:].split(":", 1)
                    if len(parts) == 2:
                        fname = parts[0].strip()
                        meaning = parts[1].strip()
                        # 更新或添加 key_transforms
                        existing = next((kt for kt in bl["key_transforms"] if kt.get("field") == fname), None)
                        if existing:
                            existing["meaning"] = meaning
                        else:
                            bl["key_transforms"].append({"field": fname, "meaning": meaning})


def _schema_table(schema, table):
    """拼接 schema.table"""
    s = _clean(schema)
    t = _clean(table)
    if s and t:
        return f"{s}.{t}"
    return t or s


def _split_schema_table(full):
    """拆分 schema.table"""
    if "." in full:
        parts = full.split(".", 1)
        return parts[0].strip(), parts[1].strip()
    return "", full.strip()


def _layer_from_schema(schema, table=""):
    """从 schema 和 table 推断数仓层级"""
    s = (schema or "").upper()
    t = (table or "").upper()
    combined = s + "." + t if s else t

    if "ODS" in combined:
        return "ODS"
    if "DIM" in combined:
        return "DIM"
    if "DWB" in combined or "DWD" in combined or "DWL" in combined:
        return "DWB"
    if "DWS" in combined:
        return "DWS"
    if "ADS" in combined or "RPT" in combined or "SLPRD" in combined:
        return "ADS"
    if "TMP" in combined or "TEMP" in combined:
        return "TMP"
    # 无 schema 的短名（CTE、子查询）标记为 CTE
    if not s and t:
        return "CTE"
    return ""  # 不显示 UNKNOWN


# ── 数据转换 ──────────────────────────────────────────────

def build_report_data(knowledge):
    """将 knowledge 结构转换为 HTML 模板所需的 REPORT_DATA"""

    meta = knowledge.get("meta", {})
    topo = knowledge.get("topology", {})
    df = knowledge.get("data_flow", {})
    fm = knowledge.get("field_mappings", {})
    bl = knowledge.get("business_logic", {})
    quality = knowledge.get("quality", {})

    steps_list = topo.get("steps", [])
    data_flow_steps = df.get("steps", [])
    fields_list = fm.get("fields", [])

    target_types = meta.get("target_field_types", {})
    patterns = meta.get("patterns", [])

    # ── summary ──（取最大 exec_sequence 步骤的目标表作为最终目标表）
    target_table = ""
    if steps_list:
        _max_seq = max(s.get("exec_sequence", 0) for s in steps_list)
        _max_step = next((s for s in steps_list if s.get("exec_sequence", 0) == _max_seq), steps_list[0])
        ts = _max_step.get("target_schema", "")
        tt = _max_step.get("target_table", "")
        target_table = _schema_table(ts, tt)

    cm = quality.get("complexity_metrics", {})
    scenarios = topo.get("scenarios", [])
    # 资产信息（I 视图标注，excel 模式无此信息则不标注）
    asset_info = meta.get("asset_info", {})
    summary = {
        "target_table": target_table,
        # 表中文名：优先用 DDL 表注释（标准稳定），没有再用 AI 业务描述
        "table_cn_name": meta.get("table_comment", "") or (bl.get("summary", "").split("，")[0] if bl.get("summary") else ""),
        "description": bl.get("summary", ""),
        "rule_count": len(steps_list),
        # 目标字段数：只算最终目标表（最大 exec_sequence 步骤）的字段，不算中间步骤
        # 中间步骤的字段不是资产的字段，算进去会多计
        "field_count": len([f for f in fields_list
                           if f.get("producing_step") == _max_step.get("step_id", "")]),
        "source_count": len([t for t in df.get("tables", []) if t.get("role") == "source"]),
        "scenario_count": len([s for s in scenarios if not s.get("is_common", False)]),
        "is_multi_scenario": len(scenarios) > 1,
        "generated_at": meta.get("analysis_time", datetime.now().strftime("%Y-%m-%d %H:%M")),
        "patterns": patterns,
        "scenarios": [{"id": s["id"], "name": s["name"], "rule_count": s["rule_count"],
                       "is_common": s.get("is_common", False)} for s in scenarios],
        "complexity": {
            "max_join_count": cm.get("max_join_count", 0),
            "max_cte_count": cm.get("max_cte_count", 0),
            "total_source_tables": cm.get("total_source_tables", 0),
            "total_case_when_branches": cm.get("total_case_when_branches", 0),
        },
        # I 视图资产信息（is_view=True 时前端渲染视图标注，否则不渲染）
        "asset_info": asset_info,
        # 加工方式（增量/全量/分区/追加）
        "load_strategy": meta.get("load_strategy", {}),
        # 多规则组链路信息（单规则组时为空）
        "is_multi_group": meta.get("is_multi_group", False),
        "chain_groups": meta.get("chain_groups", []),
    }

    # ── lineage (分层布局) ──
    lineage = _build_lineage_layout(topo, df, bl)

    # ── target_schema (目标表结构) ──
    schema_fields = []
    # 去重：同名字段只保留一个（跨步骤合并）
    seen_fields = {}
    for f in fields_list:
        fname = f.get("target_field", "")
        if fname and fname not in seen_fields:
            ai_transform = next(
                (kt for kt in bl.get("key_transforms", []) if kt.get("field") == fname),
                {}
            )
            seen_fields[fname] = {
                "name": fname,
                # 优先从字段自身的 field_type 取（P2 注入，含过程表），
                # 兜底从 meta 的 target_types 取（只有目标表）
                "type": f.get("field_type", "") or target_types.get(fname.lower(), ""),
                # 业务含义：优先 DDL 注释（field_comment），兜底 AI 增强
                "meaning": f.get("field_comment", "") or ai_transform.get("meaning", ""),
                "transform_type": f.get("transform_type", ""),
                "producing_step": f.get("producing_step", ""),
            }
    schema_fields = list(seen_fields.values())

    # ── steps (步骤详情 + SQL) ──
    steps_out = []
    # 多规则组：建 target_table → rule_group_name 映射（用于步骤卡片标注归属）
    chain_group_map = {}
    for cg in meta.get("chain_groups", []):
        chain_group_map[cg.get("target_table", "").lower()] = cg.get("name", "")
    for s in steps_list:
        step_id = s["step_id"]
        df_step = next((d for d in data_flow_steps if d.get("step_id") == step_id), {})
        # 优先用 AI 的 description，没有则用脚本兜底的（来自 business_logic.step_descriptions）
        all_step_descs = bl.get("step_descriptions", [])
        ai_step = next(
            (d for d in all_step_descs if d.get("step_id") == step_id),
            {}
        )

        steps_out.append({
            "step_id": step_id,
            "rule_code": s.get("rule_code", ""),
            "rule_name": s.get("rule_name", ""),
            "exec_sequence": s.get("exec_sequence", 0),
            "scenario_id": s.get("scenario_id", ""),
            "scenario_name": s.get("scenario_name", ""),
            "is_common_step": s.get("is_common_step", False),
            "structured_summary": _build_step_summary_inline(s, df_step, fields_list),
            "data_blocks": _merge_block_purposes(df_step.get("data_blocks", []),
                                                  ai_step.get("block_purposes", {})),
            "delete_mode_label": s.get("delete_mode_label", ""),
            "delete_condition": s.get("delete_condition", ""),
            "source_tables": s.get("source_tables_from_sql", []),
            "target_table": s.get("target_table", ""),
            "purpose": ai_step.get("purpose", ""),
            "logic": ai_step.get("logic", ""),
            "raw_sql": df_step.get("raw_sql", ""),
            # I 视图步骤标注：从 topology steps 的 is_view_step 取（统一真相源）
            "is_view_step": s.get("is_view_step", False),
            # 多规则组：该步骤属于哪个规则组（从 target_table 反查 chain_groups）
            "rule_group_name": chain_group_map.get(s.get("target_table", "").lower(), ""),
            "join_usage": df_step.get("join_usage", []),
            "where_usage": df_step.get("where_usage", []),
            "groupby_usage": df_step.get("groupby_usage", []),
            "join_paths": df_step.get("join_paths", {}),
            "union_branches": df_step.get("union_branches", []),
            "ctes": [
                {
                    "name": c.get("name", ""),
                    "source_tables": [
                        {"name": st.get("name", ""), "alias": st.get("alias", ""), "join_type": st.get("join_type", "FROM")}
                        for st in c.get("source_tables", [])
                    ],
                    "field_count": len(c.get("fields", [])),
                }
                for c in df_step.get("ctes", [])
            ],
        })

    # data_dependencies（供字段链路树连线过滤用）
    data_deps = topo.get("data_dependencies", [])

    # ── 构建 alias→物理表名 映射（用于字段来源翻译）──
    alias_table_map = {}  # {step_id: {alias(UPPER): physical_table}}
    for s in data_flow_steps:
        sid = s.get("step_id", "")
        amap = {}
        for j in s.get("joins", []):
            alias = (j.get("alias") or "").upper()
            tbl = j.get("source_table", "")
            if alias and tbl:
                amap[alias] = tbl
        alias_table_map[sid] = amap

    # ── 构建 CTE 索引（用于字段来源 CTE 穿透）──
    cte_index = _build_cte_index(data_flow_steps)
    cte_names_upper = set(cte_index.keys())

    # ── 构建 union_branches 索引（step_id → {字段(LOWER): [物理来源]}）──
    # UNION 步骤的字段来源用分支的物理穿透来源（替代假名层）
    union_branch_sources = {}  # {step_id: {field_lower: [{"table","field","branch"}]}}
    for ds in data_flow_steps:
        sid = ds.get("step_id", "")
        branches = ds.get("union_branches", [])
        if not branches:
            continue
        fmap = {}
        for b in branches:
            bidx = b.get("branch_index", 0)
            for col in b.get("columns", []):
                col_name = (col.get("alias", "") or "").lower()
                if not col_name:
                    continue
                for ps in col.get("physical_sources", []):
                    fmap.setdefault(col_name, []).append({
                        "table": ps.get("table", ""),
                        "field": ps.get("field", ""),
                        "alias": ps.get("alias", ""),
                        "branch": bidx,
                    })
        union_branch_sources[sid] = fmap

    # ── fields (按场景分组，同场景内按目标表+字段名去重) ──
    def _resolve_sources(field_data, step_id):
        """构建字段来源列表，把别名翻译成物理表名，CTE 穿透到物理源表。

        UNION 步骤优先用 union_branches 的物理穿透来源（含分支归属）。
        """
        # 优先：UNION 分支物理来源（已穿透子查询到物理表，含 branch 归属）
        ub = union_branch_sources.get(step_id)
        if ub:
            fname = (field_data.get("target_field", "") or "").lower()
            if fname in ub:
                return [{"table": s["table"], "alias": s.get("alias", ""), "field": s["field"],
                         "branch": s["branch"]} for s in ub[fname]]

        amap = alias_table_map.get(step_id, {})
        sources = []
        for l in field_data.get("lineage", []):
            src_alias = l.get("source_table", "") or l.get("alias", "")
            if not src_alias or src_alias.upper() in ("NULL", "NONE"):
                continue

            # 检查是否来自 CTE
            cte_name = l.get("cte_name", "")
            if cte_name and cte_name.upper() in cte_index:
                # CTE 穿透：从 lineage 的 cte_source_fields 取物理表字段
                cte_info = cte_index[cte_name.upper()]
                cte_field_info = cte_info["fields_map"].get(
                    (l.get("source_field", "") or "").upper(), {}
                )
                cte_alias_to_table = cte_info["alias_to_table"]
                cte_source_fields = cte_field_info.get("source_fields", l.get("cte_source_fields", []))
                for csf in cte_source_fields:
                    csf_alias = (csf.get("alias", "") or "").upper()
                    csf_field = csf.get("field", "")
                    physical_table = cte_alias_to_table.get(csf_alias, "")
                    if physical_table:
                        sources.append({
                            "table": physical_table,
                            "alias": csf.get("alias", ""),
                            "field": csf_field,
                        })
            else:
                # 普通字段：别名 → 物理表名
                physical_table = amap.get(src_alias.upper(), src_alias)
                sources.append({
                    "table": physical_table,
                    "alias": src_alias,
                    "field": l.get("source_field", ""),
                })
        return sources

    step_info_map = {}
    for s in steps_list:
        sid = s["step_id"]
        df_step = next((d for d in data_flow_steps if d.get("step_id") == sid), {})
        # join_paths 合并：第一分支 + 所有 UNION 分支的 join_paths
        # （不同分支的别名通常不同如 d1/d2，合并后渲染时按 source.alias 查找即可）
        merged_jp = dict(df_step.get("join_paths", {}))
        for b in df_step.get("union_branches", []):
            for alias, info in b.get("join_paths", {}).items():
                if alias not in merged_jp:
                    merged_jp[alias] = info
        step_info_map[sid] = {
            "target_table": s.get("target_table", ""),
            "scenario_name": s.get("scenario_name", "默认"),
            "rule_name": s.get("rule_name", ""),
            "rule_code": s.get("rule_code", ""),
            "exec_sequence": s.get("exec_sequence", 0),
            "is_view_step": s.get("is_view_step", False),  # 从 topology 传播（统一真相源）
            "join_paths": merged_jp,
            "join_key_lineage": df_step.get("join_key_lineage", {}),
        }

    # ── field_chain_map (字段 → 完整链路树，供详情面板用) ──
    # 提前构建（fields_out 需要用它计算最重加工类型）
    field_chain_map = {}
    for f in fields_list:
        fname = f.get("target_field", "")
        fname_lower = fname.lower()
        si = step_info_map.get(f.get("producing_step", ""), {})
        step_id = f.get("producing_step", "")
        sources = _resolve_sources(f, f.get("producing_step", ""))
        raw_sql = ""
        field_lineage = f.get("lineage", [])
        if field_lineage:
            raw_sql = field_lineage[0].get("raw_sql", "")

        if fname_lower not in field_chain_map:
            field_chain_map[fname_lower] = {"target_field": fname, "chains": []}
        field_chain_map[fname_lower]["chains"].append({
            "step_id": step_id,
            "rule_code": si.get("rule_code", ""),
            "rule_name": si.get("rule_name", step_id),
            "scenario": si.get("scenario_name", ""),
            "exec_sequence": si.get("exec_sequence", 0),
            "target_table": si.get("target_table", ""),
            "transform_type": f.get("transform_type", "expression"),
            "sources": sources,
            "raw_sql": raw_sql,
            "join_paths": si.get("join_paths", {}),
        })

    # ── fields (全局去重，不按场景分组；取链路中最重加工类型) ──
    fields_out = []
    seen_fields_lower = {}  # {field_lower: idx}

    # ── 字段排序 + 去重 + I 视图穿透合并 ──
    # 统一原则：以最终目标为终点展示字段。
    # I 视图场景：以 I 视图字段为基准（对外终点），F 表字段穿透合并进来，
    #   F 有 I 没有的字段追加并标注"未暴露"。不做场景区分，线性穿透。
    # 无 I 视图：以最终步骤字段为准（现有行为）。
    _base_table_norm = _norm(asset_info.get("base_table", "")) if asset_info else ""

    # 收集 I 视图步骤和 F 表步骤的字段（用于穿透合并）
    # 用 step_info_map 的 is_view_step 判断（统一真相源，不依赖 asset_info.view_step）
    view_fields = []   # I 视图步骤的字段
    base_fields = {}   # F 表步骤的字段（按字段名小写索引，用于穿透合并）
    other_fields = []  # 其他步骤的字段（中间过程）
    has_view_step = any(si.get("is_view_step") for si in step_info_map.values())
    for f in fields_list:
        ps = f.get("producing_step", "")
        si_f = step_info_map.get(ps, {})
        if has_view_step and si_f.get("is_view_step"):
            view_fields.append(f)
        elif _base_table_norm and _norm(si_f.get("target_table", "")) == _base_table_norm:
            base_fields[(f.get("target_field", "") or "").lower()] = f
        else:
            other_fields.append(f)

    # 合并后的字段列表：以 I 视图为基准，穿透 F 表链路
    if has_view_step and view_fields:
        # I 视图字段为基准，穿透合并 F 表的同名字段信息
        merged_fields = []
        view_field_names = set()
        for vf in view_fields:
            fname_lower = (vf.get("target_field", "") or "").lower()
            view_field_names.add(fname_lower)
            bf = base_fields.get(fname_lower)  # F 表同名字段
            if bf:
                # 穿透合并：用 F 表的 transform_type/field_type/field_comment/lineage
                # （F 表是加工终点，有完整血缘和类型），但 producing_step 标记为 I 视图
                merged = dict(vf)  # 以 I 视图字段为基准
                merged["transform_type"] = bf.get("transform_type", vf.get("transform_type", "expression"))
                merged["field_type"] = bf.get("field_type", "") or vf.get("field_type", "")
                merged["field_comment"] = bf.get("field_comment", "") or vf.get("field_comment", "")
                merged["lineage"] = bf.get("lineage", vf.get("lineage", []))  # F 表的链路
                merged["physical_source"] = bf.get("physical_source", vf.get("physical_source", []))
                merged["is_view_inherited"] = True  # 标注：穿透自 F 表
                merged_fields.append(merged)
            else:
                # I 有 F 没有：I 视图有额外逻辑
                vf["is_view_extra"] = True  # 标注：I 视图独有
                merged_fields.append(vf)

        # F 有 I 没有：追加并标注"F 表有但 I 视图未暴露"
        for bfname, bf in base_fields.items():
            if bfname not in view_field_names:
                bf["is_base_only"] = True  # 标注：F 有 I 没有
                merged_fields.append(bf)

        sorted_fields = merged_fields + other_fields
    else:
        # 无 I 视图：现有行为（按 exec_sequence 倒序去重）
        sorted_fields = sorted(fields_list,
            key=lambda f: -step_info_map.get(f.get("producing_step", ""), {}).get("exec_sequence", 0))

    for f in sorted_fields:
        fname = f.get("target_field", "")
        fname_lower = fname.lower()
        si = step_info_map.get(f.get("producing_step", ""), {})
        scenario = si.get("scenario_name", "默认")
        target_table = si.get("target_table", "")
        sources = _resolve_sources(f, f.get("producing_step", ""))

        # 取链路中最重的加工类型
        chain_for_field = field_chain_map.get(fname_lower, {}).get("chains", [])
        chain_priority = TRANSFORM_PRIORITY
        best_tt = f.get("transform_type", "expression")
        for cc in chain_for_field:
            cc_tt = cc.get("transform_type", "expression")
            if chain_priority.get(cc_tt, 0) > chain_priority.get(best_tt, 0):
                best_tt = cc_tt

        # 取最初来源（seq 最小步骤的 sources，穿透到物理源表）
        origin_sources = []
        if chain_for_field:
            min_seq = min(c.get("exec_sequence", 0) for c in chain_for_field)
            origin_chains = [c for c in chain_for_field if c.get("exec_sequence", 0) == min_seq]
            seen_origin = set()
            for oc in origin_chains:
                for s in oc.get("sources", []):
                    key = f"{s.get('table','')}.{s.get('field','')}"
                    if key not in seen_origin and s.get("table"):
                        seen_origin.add(key)
                        origin_sources.append(s)

        # 判断是否在最终目标表中（精确匹配，非子串包含）
        _max_seq_step = next((s for s in steps_list if s.get("exec_sequence", 0) == max(s.get("exec_sequence", 0) for s in steps_list)), None) if steps_list else None
        _final_target = _norm(_max_seq_step.get("target_table", "")) if _max_seq_step else ""
        _producing_target = _norm(si.get("target_table", ""))
        is_final_field = bool(_final_target) and _final_target == _producing_target
        # I 视图场景：F 表也是"写入目标表"（底表是资产的一部分，不是中间过程）。
        # asset_info 里的 base_table 就是 F 表，写入 F 表的字段也是 final。
        if not is_final_field and asset_info:
            _base_table = _norm(asset_info.get("base_table", ""))
            if _base_table and _base_table == _producing_target:
                is_final_field = True

        if fname_lower in seen_fields_lower:
            continue
        seen_fields_lower[fname_lower] = len(fields_out)
        fields_out.append({
            "target_field": fname,
            "producing_step": f.get("producing_step", ""),
            "target_table": target_table,
            "scenario": scenario,
            "transform_type": best_tt,
            # 字段类型+业务含义：从 P2 注入的 field_type/field_comment 取（写入表的类型）
            "field_type": f.get("field_type", ""),
            "field_comment": f.get("field_comment", ""),
            "origin_sources": origin_sources,
            "is_final_field": is_final_field,
            "in_target_fields": f.get("in_target_fields", False),
            "excel_source_field": f.get("excel_source_field", ""),
            "sources": sources,
            # I 视图穿透标注（无 asset_info 时这些都不存在，兼容 excel 模式）
            "is_view_inherited": f.get("is_view_inherited", False),  # 穿透自F表
            "is_view_extra": f.get("is_view_extra", False),          # I视图独有
            "is_base_only": f.get("is_base_only", False),            # F有I没有
        })

    # ── field_chain_map (字段 → 完整链路树，供详情面板用) ──
    # 每个 field 收集它在所有步骤+所有场景中的来源
    field_chain_map = {}
    for f in fields_list:
        fname = f.get("target_field", "")
        fname_lower = fname.lower()
        si = step_info_map.get(f.get("producing_step", ""), {})
        step_id = f.get("producing_step", "")
        sources = _resolve_sources(f, f.get("producing_step", ""))

        if fname_lower not in field_chain_map:
            field_chain_map[fname_lower] = {
                "target_field": fname,
                "chains": [],  # 所有步骤的来源
            }
        # 取 raw_sql（加工表达式）
        raw_sql = ""
        field_lineage = f.get("lineage", [])
        if field_lineage:
            raw_sql = field_lineage[0].get("raw_sql", "")

        field_chain_map[fname_lower]["chains"].append({
            "rule_code": si.get("rule_code", ""),
            "step_id": step_id,
            "rule_name": si.get("rule_name", step_id),
            "scenario": si.get("scenario_name", ""),
            "exec_sequence": si.get("exec_sequence", 0),
            "target_table": si.get("target_table", ""),
            "transform_type": f.get("transform_type", "expression"),
            "sources": sources,
            "raw_sql": raw_sql,
            "join_paths": si.get("join_paths", {}),
            "join_key_lineage": si.get("join_key_lineage", {}),
        })

    # ── field_details (CTE 穿透血缘链) ──
    field_details = {}
    for f in fields_list:
        fname = f.get("target_field", "")
        new_type = f.get("transform_type", "expression")
        if fname in field_details:
            existing_type = field_details[fname].get("transform_type", "expression")
            if existing_type != "direct" and new_type == "direct":
                continue
        ai_transform = next(
            (kt for kt in bl.get("key_transforms", []) if kt.get("field") == fname),
            {}
        )

        # CTE 穿透血缘链
        penetration_chain = _build_penetration_chain(f.get("lineage", []), cte_index, cte_names_upper)

        field_details[fname] = {
            "target_field": fname,
            "producing_step": f.get("producing_step", ""),
            "rule_code": f.get("rule_code", ""),
            "transform_type": new_type,
            "in_target_fields": f.get("in_target_fields", False),
            "lineage": f.get("lineage", []),
            "penetration_chain": penetration_chain,
            "validation": f.get("validation", None),
            "meaning": ai_transform.get("meaning", ""),
        }

    # ── quality ──
    quality_out = {
        "complexity": quality.get("complexity_metrics", {}),
        "issues": quality.get("issues", []),
        "ai_insights": quality.get("ai_insights", []),
    }

    # ── 字段使用汇总（字段名 → 跨步骤的关联/过滤/分组信息）──
    field_usage_map = {}  # {field_lower: {join: [...], where: [...], groupby: [...]}}
    for s in steps_out:
        sid = s.get("step_id", "")
        sname = s.get("rule_name", "") or sid
        scenario = s.get("scenario_name", "")

        for ju in s.get("join_usage", []):
            fname = (ju.get("field") or "").lower()
            if not fname:
                continue
            if fname not in field_usage_map:
                field_usage_map[fname] = {"join": [], "where": [], "groupby": []}
            field_usage_map[fname]["join"].append({
                "step_id": sid, "step_name": sname, "scenario": scenario,
                "join_type": ju.get("join_type", ""),
                "on_condition": ju.get("on_condition", ""),
                "tables": ju.get("tables", []),
            })

        for wu in s.get("where_usage", []):
            fname = (wu.get("field") or "").lower()
            if not fname:
                continue
            if fname not in field_usage_map:
                field_usage_map[fname] = {"join": [], "where": [], "groupby": []}
            field_usage_map[fname]["where"].append({
                "step_id": sid, "step_name": sname, "scenario": scenario,
                "condition": wu.get("condition", ""),
            })

        for gu in s.get("groupby_usage", []):
            fname = (gu.get("field") or "").lower()
            if not fname:
                continue
            if fname not in field_usage_map:
                field_usage_map[fname] = {"join": [], "where": [], "groupby": []}
            field_usage_map[fname]["groupby"].append({
                "step_id": sid, "step_name": sname, "scenario": scenario,
            })

    # ── 辅助字段（出现在 usage 里但不在 fields_out 里的字段）──
    write_field_names = set(f["target_field"].lower() for f in fields_out)
    auxiliary_fields = []
    for fname, usage in field_usage_map.items():
        if fname not in write_field_names:
            roles = []
            if usage["join"]:
                roles.append("关联键")
            if usage["where"]:
                roles.append("过滤")
            if usage["groupby"]:
                roles.append("分组")
            auxiliary_fields.append({
                "field": fname,
                "roles": roles,
                "usage": usage,
            })

    return {
        "summary": summary,
        "lineage": lineage,
        "schema_fields": schema_fields,
        "steps": steps_out,
        "fields": fields_out,
        "field_chain_map": field_chain_map,
        "data_deps": data_deps,
        "field_details": field_details,
        "field_usage_map": field_usage_map,
        "auxiliary_fields": auxiliary_fields,
        "quality": quality_out,
    }


def _format_join(join_type: str, source_table: str) -> str:
    """格式化 JOIN 描述（避免 'LEFT JOIN JOIN ON' 重复）"""
    if join_type == "FROM":
        return f"FROM {source_table}"
    return f"{join_type} {source_table}"


def _build_cte_index(data_flow_steps):
    """构建 CTE 索引用于穿透。"""
    cte_index = {}
    for s in data_flow_steps:
        for cte in s.get("ctes", []):
            cte_key = cte.get("name", "").upper()
            alias_to_table = {}
            for st in cte.get("source_tables", []):
                talias = st.get("alias", "").upper()
                tname = st.get("name", "")
                if talias:
                    alias_to_table[talias] = tname
            fields_map = {}
            for cf in cte.get("fields", []):
                if isinstance(cf, dict) and cf.get("name"):
                    fields_map[cf["name"].upper()] = cf
            cte_index[cte_key] = {
                "alias_to_table": alias_to_table,
                "fields_map": fields_map,
                "source_tables": cte.get("source_tables", []),
            }
    return cte_index


def _build_penetration_chain(lineages, cte_index, cte_names_upper, visited=None, depth=0):
    """构建字段穿透血缘链（用于详情面板展示）。

    返回: [{level, node_type, name, field, transform_type, expression}, ...]
    """
    if visited is None:
        visited = set()
    if depth > 10 or not lineages:
        return []

    chain = []
    for l in lineages:
        src_table = l.get("source_table", "")
        src_field = l.get("source_field", "")
        cte_name = l.get("cte_name", "")
        raw_sql = l.get("raw_sql", "")
        cte_transform = l.get("cte_transform_type", "")

        if cte_name and cte_name.upper() in cte_index:
            if cte_name.upper() in visited:
                continue
            visited.add(cte_name.upper())

            # CTE 节点
            chain.append({
                "level": depth,
                "node_type": "cte",
                "name": cte_name,
                "field": src_field,
                "transform_type": cte_transform or "unknown",
                "expression": l.get("cte_expression", raw_sql),
            })

            # 递归 CTE 内部
            cte_info = cte_index[cte_name.upper()]
            cte_field_info = cte_info["fields_map"].get(src_field.upper(), {})
            cte_source_fields = cte_field_info.get("source_fields", l.get("cte_source_fields", []))
            cte_alias_to_table = cte_info["alias_to_table"]

            for csf in cte_source_fields:
                csf_alias = csf.get("alias", "").upper()
                csf_field = csf.get("field", "")
                physical_table = cte_alias_to_table.get(csf_alias, "")

                if physical_table:
                    sch, tbl = _split_schema_table(physical_table)
                    chain.append({
                        "level": depth + 1,
                        "node_type": "physical",
                        "name": physical_table,
                        "field": csf_field,
                        "alias": csf.get("alias", ""),
                        "transform_type": "direct",
                        "expression": "",
                    })
                elif csf_alias in cte_names_upper:
                    # 嵌套 CTE
                    nested_chain = _build_penetration_chain(
                        [{"source_table": csf_alias, "source_field": csf_field, "cte_name": csf_alias}],
                        cte_index, cte_names_upper, visited, depth + 2
                    )
                    chain.extend(nested_chain)

            visited.discard(cte_name.upper())
        elif src_table and src_table.upper() not in cte_names_upper:
            # 物理源表
            chain.append({
                "level": depth,
                "node_type": "physical",
                "name": src_table,
                "field": src_field,
                "alias": l.get("alias", ""),
                "transform_type": l.get("transform", "direct"),
                "expression": raw_sql,
            })

    return chain


def _build_lineage_layout(topo, df, bl=None):
    """构建数据流向图布局。

    布局策略：
    - 步骤按 exec_sequence 分列（同 seq 同列）
    - 中间表放在产出步骤和消费步骤之间
    - 目标表放在最后一列
    - 来源表标记 hidden=true（默认不显示，可切换）
    - 来源表标注引用次数（被几个步骤使用）
    - 交互高亮：点击步骤高亮直接前后关系
    """
    tables = df.get("tables", [])
    steps_list = topo.get("steps", [])
    data_deps = topo.get("data_dependencies", [])
    data_flow_steps = df.get("steps", [])
    bl = bl or {}

    # ── 1. 分类节点（统一用 _norm 做大小写归一化）──
    all_target_tables = {}  # {norm_table: step_id}
    for s in steps_list:
        tf = _schema_table(s.get("target_schema", ""), s.get("target_table", ""))
        all_target_tables[_norm(tf)] = s["step_id"]

    final_targets = set()
    if steps_list:
        max_seq = max(s.get("exec_sequence", 0) for s in steps_list)
        for s in steps_list:
            if s.get("exec_sequence", 0) == max_seq:
                tf = _schema_table(s.get("target_schema", ""), s.get("target_table", ""))
                final_targets.add(_norm(tf))

    # CTE 名
    # 先全量收集所有 CTE 名（统一大写），再构建 source_map（避免 CTE_A 引用
    # CTE_B 时 B 还没进 cte_names 而被误当物理表——顺序依赖 bug）
    cte_names_upper = set()
    for s in data_flow_steps:
        for cte in s.get("ctes", []):
            cn = cte.get("name", "")
            if cn:
                cte_names_upper.add(cn.upper())
    cte_source_map = {}
    for s in data_flow_steps:
        for cte in s.get("ctes", []):
            cn = cte.get("name", "")
            if cn:
                phys = []
                for st in cte.get("source_tables", []):
                    tname = st.get("name", "")
                    if tname and tname.upper() not in cte_names_upper:
                        phys.append(tname)
                cte_source_map[cn] = phys
    cte_names = cte_names_upper  # 兼容下游（下游比较都用 upper）

    # 来源表引用计数 + 主表识别
    source_ref_count = {}  # {table_name(UPPER): count}
    step_primary_tables = {}  # {step_id: set(table_name(UPPER))}  主表集合
    for s in steps_list:
        sid = s["step_id"]
        primary = set()
        for src in s.get("source_tables_from_sql", []):
            su = _norm(src)
            source_ref_count[su] = source_ref_count.get(su, 0) + 1
        # 从 data_flow 获取 JOIN 类型，识别主表
        df_step = next((d for d in data_flow_steps if d.get("step_id") == sid), {})
        for j in df_step.get("joins", []):
            jt = (j.get("join_type") or "").upper()
            tbl = _norm(j.get("source_table") or "")
            if jt == "FROM":
                # FROM 表始终是主表
                primary.add(tbl)
            elif jt == "FROM_SUBQUERY_MAIN":
                # FROM 子查询内部的 FROM 主表（透传主表属性）
                primary.add(tbl)
            elif "INNER" in jt or "CROSS" in jt:
                # INNER JOIN / CROSS JOIN 也是主表
                primary.add(tbl)
        step_primary_tables[sid] = primary
    # CTE 内部表也算
    for cte_name, phys_list in cte_source_map.items():
        for p in phys_list:
            pu = _norm(p)
            source_ref_count[pu] = source_ref_count.get(pu, 0) + 1

    # ── 2. 构建节点 ──
    nodes = {}
    edges_list = []

    # 步骤节点
    step_seq_map = {}  # {step_id: seq}
    for s in steps_list:
        sid = s["step_id"]
        seq = s.get("exec_sequence", 0)
        step_seq_map[sid] = seq
        rule_name = s.get("rule_name", "")
        scenario_name = s.get("scenario_name", "")
        label = rule_name if rule_name else sid
        if scenario_name:
            label = f"[{scenario_name}] {label}"
        nodes[sid] = {
            "type": "step",
            "label": label,
            "hidden": False,
            "step_data": {
                "rule_code": s.get("rule_code", ""),
                "exec_sequence": seq,
                "scenario_name": scenario_name,
                "rule_type": s.get("rule_type", 1),
            },
        }

    # 表节点（去重按 _norm 归一化，避免大小写不同的同名表重复出现）
    seen_tables = set()
    for t in tables:
        tname = _schema_table(t.get("schema", ""), t.get("name", ""))
        if not tname or _norm(tname) in seen_tables:
            continue
        seen_tables.add(_norm(tname))
        if tname in cte_names:
            continue

        if _norm(tname) in final_targets:
            node_type = "target"
            cn = bl.get("summary", "").split("，")[0] if bl.get("summary") else ""
            label = tname if not cn else f"{tname} ({cn})"
            hidden = False
        elif _norm(tname) in all_target_tables:
            node_type = "intermediate"
            label = tname
            hidden = False
        else:
            node_type = "source"
            ref_count = source_ref_count.get(_norm(tname), 1)
            label = f"{tname} (×{ref_count})" if ref_count > 1 else tname
            hidden = True  # 来源表默认隐藏

        # 判断是否是某个步骤的主表
        is_primary = False
        for sid, primary_set in step_primary_tables.items():
            if _norm(tname) in primary_set:
                is_primary = True
                break

        nodes[tname] = {"type": node_type, "label": label, "hidden": hidden, "step_data": None,
                        "is_primary": is_primary}

    # CTE 内部物理表
    for cte_name, phys_list in cte_source_map.items():
        for p in phys_list:
            if _norm(p) not in seen_tables and p not in cte_names:
                seen_tables.add(_norm(p))
                ref_count = source_ref_count.get(_norm(p), 1)
                label = f"{p} (×{ref_count})" if ref_count > 1 else p
                nodes[p] = {"type": "source", "label": label, "hidden": True, "step_data": None,
                            "is_primary": False}

    # ── 3. 构建边 ──
    # 步骤 → 目标表/中间表
    step_to_table = {}  # {step_id: table_name}
    # 构建 norm → node_name 映射（大小写不敏感查找）
    norm_nodes = {}
    for nid in nodes:
        norm_nodes[_norm(nid)] = nid

    for s in steps_list:
        sid = s["step_id"]
        tf = _schema_table(s.get("target_schema", ""), s.get("target_table", ""))
        tf_norm = _norm(tf)
        if tf_norm in norm_nodes:
            tf_actual = norm_nodes[tf_norm]
            edges_list.append({"from": sid, "to": tf_actual, "label": "", "type": "step_to_table"})
            step_to_table[sid] = tf_actual

    # 来源表 → 步骤
    table_to_steps = {}  # {table: [step_ids]}
    for s in steps_list:
        sid = s["step_id"]
        for src in s.get("source_tables_from_sql", []):
            src_norm = _norm(src)
            if src_norm in norm_nodes and src != sid:
                src_actual = norm_nodes[src_norm]
                edges_list.append({"from": src_actual, "to": sid, "label": "", "type": "source_to_step"})
                table_to_steps.setdefault(src_actual, []).append(sid)
        # CTE 内部源表 → 步骤
        df_step = next((d for d in data_flow_steps if d.get("step_id") == sid), {})
        for cte in df_step.get("ctes", []):
            for st in cte.get("source_tables", []):
                tname = st.get("name", "")
                tname_norm = _norm(tname)
                if tname_norm in norm_nodes and tname != sid:
                    tname_actual = norm_nodes[tname_norm]
                    edges_list.append({"from": tname_actual, "to": sid, "label": f"CTE:{cte.get('name','')}", "type": "source_to_step"})
                    table_to_steps.setdefault(tname_actual, []).append(sid)

    # 数据依赖（中间表 → 后续步骤）
    # 排除 delete_before_write 类型（它只是执行顺序依赖，不是数据传递，
    # 画成"目标表→步骤"会误导，让人以为步骤从目标表取数据）
    for dep in data_deps:
        if dep.get("type") == "delete_before_write":
            continue
        from_step = dep.get("from", "")
        to_step = dep.get("to", "")
        intermediate_table = step_to_table.get(from_step, "")
        if intermediate_table and _norm(intermediate_table) in norm_nodes and to_step in nodes:
            edges_list.append({"from": intermediate_table, "to": to_step, "label": "", "type": "dep"})

    # ── 4. 按 exec_sequence 分列计算坐标 ──
    LAYER_WIDTH = 280
    NODE_HEIGHT = 36
    NODE_GAP = 10
    MARGIN_TOP = 30
    MARGIN_LEFT = 20

    # 步骤的列 = exec_sequence
    # 中间表的列 = 产出步骤的列 + 0.5（放在步骤右边、下一步骤左边）
    # 目标表的列 = max_seq + 1
    # 来源表的列 = 使用它的步骤的列 - 0.5（放在步骤左边），但默认隐藏不渲染

    max_seq = max(step_seq_map.values()) if step_seq_map else 0

    # 计算每个节点的列号（浮点数，整数列放步骤，小数列放表）
    # 步骤按 exec_sequence 排列，目标表在最后，写同一目标的步骤保持调度序
    node_col = {}
    for nid, ninfo in nodes.items():
        if ninfo["type"] == "step":
            node_col[nid] = step_seq_map.get(nid, 0)
        elif ninfo["type"] == "target":
            node_col[nid] = max_seq + 1
        elif ninfo["type"] == "intermediate":
            # 找产出它的步骤
            producer_step = all_target_tables.get(_norm(nid))
            if producer_step:
                node_col[nid] = step_seq_map.get(producer_step, 0) + 0.5
            else:
                node_col[nid] = max_seq + 0.5
        else:  # source
            # 放在使用它的最早步骤（最小 exec_sequence）的列 - 0.5
            # 这样所有引用此表的步骤都在表的右边，避免视觉回连
            using_steps = table_to_steps.get(nid, [])
            if using_steps:
                min_seq = min(step_seq_map.get(sid, 999) for sid in using_steps)
                node_col[nid] = min_seq - 0.5
            else:
                node_col[nid] = -0.5

    # 计算实际渲染列（包含所有节点，隐藏节点也占位以便开关打开时有坐标）
    visible_cols = sorted(set(node_col.values()))
    col_to_x = {}
    for i, col in enumerate(visible_cols):
        col_to_x[col] = MARGIN_LEFT + i * LAYER_WIDTH

    # 计算每个节点的 Y 坐标（同列垂直排列，隐藏节点也分配坐标）
    col_nodes = {}
    for nid in nodes:
        col = node_col[nid]
        col_nodes.setdefault(col, []).append(nid)

    max_nodes_in_col = max(len(v) for v in col_nodes.values()) if col_nodes else 0
    total_height = MARGIN_TOP * 2 + max_nodes_in_col * (NODE_HEIGHT + NODE_GAP)

    positions = {}
    node_meta = {}
    node_id_counter = 0

    # 同列节点按自然顺序垂直排列即可避免重叠。
    # 不再用 stagger_steps 固定偏移（多场景同 exec_sequence 时会导致位置撞车）。
    for col in sorted(col_nodes.keys()):
        x = col_to_x.get(col, MARGIN_LEFT)
        # 排序：步骤在前（按 exec_sequence + step_id），表在后
        col_nids = sorted(col_nodes[col], key=lambda n: (
            nodes[n]["type"] != "step",  # step 排前面
            step_seq_map.get(n, 0) if nodes[n]["type"] == "step" else 0,
            n,  # step_id 兜底（同 exec_sequence 时按 step_id 排序，保证唯一顺序）
        ))
        col_count = len(col_nids)
        total_h = col_count * (NODE_HEIGHT + NODE_GAP) - NODE_GAP
        start_y = max(MARGIN_TOP, (total_height - total_h) / 2)

        for ni, name in enumerate(col_nids):
            y = start_y + ni * (NODE_HEIGHT + NODE_GAP)
            node_id = f"n_{node_id_counter}"
            node_id_counter += 1
            positions[name] = node_id

            ninfo = nodes.get(name, {})
            node_meta[node_id] = {
                "id": node_id,
                "name": name,
                "label": ninfo.get("label", name),
                "x": x,
                "y": y,
                "width": 220,
                "height": NODE_HEIGHT,
                "type": ninfo.get("type", "source"),
                "layer": visible_cols.index(col) if col in visible_cols else 0,
                "hidden": ninfo.get("hidden", False),
                "step_data": ninfo.get("step_data"),
                "source_ref_count": source_ref_count.get(_norm(name), 1) if ninfo.get("type") == "source" else 0,
                "is_primary": ninfo.get("is_primary", False),
            }

    # 隐藏节点也需要 id（用于交互高亮时显示）
    for nid in nodes:
        if nid not in positions:
            node_id = f"n_{node_id_counter}"
            node_id_counter += 1
            positions[nid] = node_id
            ninfo = nodes[nid]
            node_meta[node_id] = {
                "id": node_id,
                "name": nid,
                "label": ninfo.get("label", nid),
                "x": 0,
                "y": 0,
                "width": 220,
                "height": NODE_HEIGHT,
                "type": ninfo.get("type", "source"),
                "layer": -1,
                "hidden": True,
                "step_data": None,
                "source_ref_count": source_ref_count.get(_norm(nid), 1),
                "is_primary": ninfo.get("is_primary", False),
            }

    # ── 5. 构建边输出 ──
    edges_out = []
    for e in edges_list:
        f, t = e["from"], e["to"]
        if f in positions and t in positions:
            edges_out.append({
                "from": positions[f],
                "to": positions[t],
                "label": e.get("label", ""),
                "type": e.get("type", "data_flow"),
                "from_hidden": nodes.get(f, {}).get("hidden", False),
                "to_hidden": nodes.get(t, {}).get("hidden", False),
            })

    self_ref_ids = []
    for sr in topo.get("self_references", []):
        target = sr.get("table", "")
        if target in positions:
            self_ref_ids.append(positions[target])

    total_width = MARGIN_LEFT * 2 + len(visible_cols) * LAYER_WIDTH
    return {
        "nodes": list(node_meta.values()),
        "edges": edges_out,
        "self_references": self_ref_ids,
        "layout": {
            "width": total_width,
            "height": total_height,
            "layer_count": len(visible_cols),
            "compact": max_nodes_in_col > 8,
            "layer_width": LAYER_WIDTH,
            "has_hidden_sources": any(n.get("hidden") for n in node_meta.values()),
        },
        "schedule_groups": [
            {"sequence": g.get("sequence", 0), "steps": g.get("parallel_steps", [])}
            for g in topo.get("schedule_plan", [])
        ],
    }


# ── 视图生成: asset_report.html ─────────────────────────

def generate_asset_report(knowledge, output_dir):
    """生成资产说明书 HTML"""
    report_data = build_report_data(knowledge)

    # 读取模板
    template_path = Path(__file__).parent / "templates" / "asset_report.html"
    if not template_path.exists():
        print(f"  错误: 模板文件不存在: {template_path}", file=sys.stderr)
        return False

    template = template_path.read_text(encoding="utf-8")

    # 替换占位符
    json_str = json.dumps(report_data, ensure_ascii=False, indent=2)
    # 转义 </script> 避免浏览器误解析
    json_str_safe = json_str.replace("</script>", "<\\/script>")
    html = template.replace("{{REPORT_DATA}}", json_str_safe)

    # 替换 title 占位符
    target_table = report_data["summary"]["target_table"]
    html = html.replace("{{TARGET_TABLE}}", target_table)

    # 写入
    output_path = Path(output_dir) / "asset_report.html"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8", newline="\n")
    print(f"  [OK] 资产说明书: {output_path}")
    return True


# ── 视图生成: mapping.xlsx ──────────────────────────────

def generate_mapping(knowledge, output_dir):
    """生成 Mapping Excel

    规则：
    - 实体级 mapping：只展示物理源表 → 目标表（CTE/中间表不显示）
    - 属性级 mapping：穿透 CTE 到物理源表字段
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  错误: 缺少 openpyxl，请 pip install openpyxl", file=sys.stderr)
        return False

    topo = knowledge.get("topology", {})
    df = knowledge.get("data_flow", {})
    fm = knowledge.get("field_mappings", {})
    bl = knowledge.get("business_logic", {})

    steps_list = topo.get("steps", [])
    data_flow_steps = df.get("steps", [])
    fields_list = fm.get("fields", [])
    tables_info = df.get("tables", [])

    # ── 构建 CTE 字典：{CTE名(UPPER): {source_tables, fields_map}} ──
    cte_index = {}
    cte_names_upper = set()
    for s in data_flow_steps:
        for cte in s.get("ctes", []):
            cte_key = cte.get("name", "").upper()
            cte_names_upper.add(cte_key)
            src_tables = cte.get("source_tables", [])
            # 构建 alias → 物理表名 的映射
            alias_to_table = {}
            for st in src_tables:
                talias = st.get("alias", "").upper()
                tname = st.get("name", "")
                if talias:
                    alias_to_table[talias] = tname
            # 构建 field_name → source_fields 的映射
            fields_map = {}
            for cf in cte.get("fields", []):
                if isinstance(cf, dict) and cf.get("name"):
                    fields_map[cf["name"].upper()] = cf
            cte_index[cte_key] = {
                "alias_to_table": alias_to_table,
                "fields_map": fields_map,
                "source_tables": src_tables,
            }

    # ── 构建 target table 列表（用于过滤，归一化避免大小写漏匹配）──
    target_tables_set = set()
    for s in steps_list:
        tf = _schema_table(s.get("target_schema", ""), s.get("target_table", ""))
        target_tables_set.add(_norm(tf))

    wb = Workbook()

    # DDL 元数据（字段类型+中文名）
    meta = knowledge.get("meta", {})
    ddl_types = meta.get("target_field_types", {})
    ddl_comments = meta.get("target_field_comments", {})

    # ════════════════════════════════════════════════════════════
    # Sheet 1: 实体级 mapping — 物理源表 → 目标表
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "实体级mapping"
    # 源表别名放在源表表名后面
    headers1 = [
        "源表schema", "源表物理表名", "源表别名", "源表中文名",
        "目标表schema", "目标表中文名", "目标表物理表名",
        "关联&限定条件", "备注", "调度任务名称", "执行路径", "依赖参数"
    ]
    ws1.append(headers1)

    # 目标表 schema/table 都从 exec_sequence 最大的步骤取（统一口径）
    target_schema = ""
    target_table = ""
    if steps_list:
        _max_seq = max(s.get("exec_sequence", 0) for s in steps_list)
        _max_steps = [s for s in steps_list if s.get("exec_sequence", 0) == _max_seq]
        if _max_steps:
            target_schema = _max_steps[0].get("target_schema", "")
            target_table = _max_steps[0].get("target_table", "")
    target_cn = bl.get("summary", "").split("，")[0] if bl.get("summary") else target_table

    # 收集所有物理源表（含 CTE/子查询内部物理表），排除假名、CTE 名和目标表自身
    def _join_type_label(jt, join_cond):
        """把 join_type 翻译成友好关联描述"""
        if jt == "FROM":
            return "主表"
        if jt == "FROM_SUBQUERY_MAIN":
            return "子查询内部主表"
        if jt == "FROM_SUBQUERY":
            return "子查询内部从表"
        if jt == "JOIN_SUBQUERY_INNER":
            return "JOIN子查询内部从表"
        if jt == "JOIN_SUBQUERY":
            return "JOIN子查询"
        return f"{jt} ON {join_cond}" if join_cond else jt

    # 源表去重：seen_global（string key）跨步骤+跨段全局去重，
    # seen_branches（tuple key）UNION 分支内去重。两者分离避免 key 类型混用。
    seen_global = set()
    seen_branches = set()
    entity_rows = []
    for s in steps_list:
        sid = s["step_id"]
        df_step = next((d for d in data_flow_steps if d.get("step_id") == sid), {})
        joins = df_step.get("joins", [])
        ctes = df_step.get("ctes", [])
        ub_list = df_step.get("union_branches", [])
        has_union = len(ub_list) >= 1

        if has_union:
            # UNION 步骤：按分支收集源表（带分支归属）
            for b in ub_list:
                bidx = b.get("branch_index", 0)
                for j in b.get("source_tables", []):
                    src_full = j.get("source_table", "")
                    if not src_full or src_full.startswith("(subquery:"):
                        continue
                    if _norm(src_full) in target_tables_set:
                        continue
                    if src_full.upper() in cte_names_upper:
                        continue
                    # 同表在不同分支都显示（分支+表 做去重 key）
                    dedup_key = (bidx, _norm(src_full))
                    if dedup_key in seen_branches:
                        continue
                    seen_branches.add(dedup_key)
                    seen_global.add(_norm(src_full))
                    sch, tbl = _split_schema_table(src_full)
                    relation = _join_type_label(j.get("join_type", ""), j.get("join_condition", ""))
                    entity_rows.append([
                        sch, tbl, j.get("alias", ""), "",
                        target_schema, target_cn, target_table,
                        relation, "", "", f"分支{bidx}", "",
                    ])
        else:
            # 非 UNION：主查询 JOIN（物理表），过滤子查询假名和中间表
            where_clause = (df_step.get("where_clause", "") or "").replace("WHERE ", "")
            # 本步骤的别名映射 + 关联键追溯（用于 ON 条件中间表别名替换）
            step_alias_map = {}
            for j in joins:
                if j.get("alias") and j.get("source_table"):
                    step_alias_map[j["alias"].upper()] = j["source_table"]
            step_jkl = df_step.get("join_key_lineage", {})
            for j in joins:
                src_full = j.get("source_table", "")
                if not src_full or src_full.startswith("(subquery:"):
                    continue
                # 过滤中间表（实体级只显示物理源表，中间表不出现）
                if _is_intermediate_tbl(src_full):
                    continue
                if _norm(src_full) in seen_global or _norm(src_full) in target_tables_set:
                    continue
                if src_full.upper() in cte_names_upper:
                    continue
                seen_global.add(_norm(src_full))
                sch, tbl = _split_schema_table(src_full)
                # ON 条件里的中间表别名替换为物理源表
                raw_cond = j.get("join_condition", "")
                rewritten_cond, transfer_note = _resolve_on_condition_aliases(
                    raw_cond, step_alias_map, step_jkl)
                relation = _join_type_label(j.get("join_type", ""), rewritten_cond)
                if transfer_note:
                    relation += transfer_note
                # WHERE 条件放在备注列（只有第一个源表行放，避免重复）
                remark = where_clause if not entity_rows or not entity_rows[-1][8] else ""
                entity_rows.append([
                    sch, tbl, j.get("alias", ""), "",
                    target_schema, target_cn, target_table,
                    relation, remark, "", "", "",
                ])

        # CTE 内部物理表（UNION 和非 UNION 都可能有 CTE）
        for cte in ctes:
            for st in cte.get("source_tables", []):
                tname = st.get("name", "")
                if not tname or _norm(tname) in seen_global or _norm(tname) in target_tables_set:
                    continue
                if tname.upper() in cte_names_upper:
                    continue
                seen_global.add(_norm(tname))
                sch, tbl = _split_schema_table(tname)
                talias = st.get("alias", "")
                jt = st.get("join_type", "FROM")
                relation = f"CTE {cte['name']} 内部 {jt}" if jt != "FROM" else f"CTE {cte['name']} 主表"
                entity_rows.append([
                    sch, tbl, talias, "",
                    target_schema, target_cn, target_table,
                    relation, "", "", "", "",
                ])

    for row in entity_rows:
        ws1.append(row)

    # ════════════════════════════════════════════════════════════
    # Sheet 2: 属性级 mapping — 穿透 CTE 到物理源表字段
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("属性级mapping")
    # 源表别名放在源表物理表名后面
    headers2 = [
        "场景", "源表schema", "源表物理表名", "源表别名", "源字段名", "源字段类型",
        "映射描述", "映射表达式",
        "目标字段名", "目标字段中文名", "目标字段类型"
    ]
    ws2.append(headers2)

    # 过滤视图步骤
    def _is_view_step(step_id: str) -> bool:
        df_step = next((d for d in data_flow_steps if d.get("step_id") == step_id), {})
        raw_sql = (df_step.get("raw_sql", "") or "").upper()
        if "CREATE VIEW" in raw_sql or "CREATE OR REPLACE VIEW" in raw_sql:
            return True
        return False

    # 构建 step_id → scenario_name 映射
    step_scenario = {}
    for s in steps_list:
        sid = s.get("step_id", "")
        step_scenario[sid] = s.get("scenario_name", "")

    # 字段按场景分组，同场景内字段去重（优先保留加工版本）
    # 不同场景可以有相同字段名（各自独立映射）
    scenario_fields: dict[str, list] = {}  # {scenario: [fields]}
    seen_in_scenario: dict[str, set] = {}   # {scenario: {field_names}}

    for f in fields_list:
        fname = f.get("target_field", "")
        step_id = f.get("producing_step", "")
        if _is_view_step(step_id) or not fname:
            continue

        # 跳过中间表步骤的字段（只展示最终目标表，中间表靠 physical_source 穿透体现）
        step_info = next((s for s in steps_list if s.get("step_id") == step_id), {})
        step_target = step_info.get("target_table_full", "") or step_info.get("target_table", "")
        if _is_intermediate_tbl(step_target):
            continue

        scenario = step_scenario.get(step_id, "默认场景")
        if scenario not in scenario_fields:
            scenario_fields[scenario] = []
            seen_in_scenario[scenario] = set()

        # 同场景内去重
        if fname in seen_in_scenario[scenario]:
            # 找到已有的，优先保留加工版本
            for i, existing in enumerate(scenario_fields[scenario]):
                if existing.get("target_field") == fname:
                    existing_tt = existing.get("transform_type", "expression")
                    new_tt = f.get("transform_type", "expression")
                    priority = TRANSFORM_PRIORITY
                    if priority.get(new_tt, 0) > priority.get(existing_tt, 0):
                        scenario_fields[scenario][i] = f
                    break
        else:
            seen_in_scenario[scenario].add(fname)
            scenario_fields[scenario].append(f)

    # 按场景顺序输出
    for scenario, sc_fields in scenario_fields.items():
        for f in sc_fields:
            tt = f.get("transform_type", "expression")
            rule_map = {"direct": "直取", "value": "赋值"}
            rule = rule_map.get(tt, "加工")

            target_field_name = f.get("target_field", "")
            # 优先从字段自身的 field_type/field_comment 取（P2 注入，含过程表），
            # 兜底从 meta 的 ddl_types/ddl_comments 取（只有目标表）
            field_cn = f.get("field_comment", "") or ddl_comments.get(target_field_name.lower(), "")
            field_type = f.get("field_type", "") or ddl_types.get(target_field_name.lower(), "")

            # 优先用跨步骤穿透的 physical_source（追到物理源表），没有则回退 CTE 穿透
            phys_sources = f.get("physical_source", [])
            if not phys_sources:
                lineages = f.get("lineage", [])
                phys_sources = _resolve_physical_sources(lineages, cte_index, cte_names_upper, set())

            if not phys_sources:
                ws2.append([
                    scenario, "", "", "", "", "",
                    _describe_transform(tt), "",
                    target_field_name, field_cn, field_type,
                ])
            else:
                for ps in phys_sources:
                    ps_tt = ps.get("transform", tt)
                    ps_field = ps.get("field", target_field_name)
                    ps_raw = ps.get("raw_sql", "")
                    describe = _describe_transform(ps_tt, ps_raw, ps_field)
                    # schema/table 拆分
                    ptable = ps.get("table", "")
                    parts = ptable.split(".")
                    p_schema = parts[0] if len(parts) > 1 else ""
                    p_table = parts[-1] if len(parts) > 1 else parts[0]
                    ws2.append([
                        scenario, p_schema, p_table, ps.get("alias", ""), ps_field, "",
                        describe, ps_raw,
                        target_field_name, field_cn, field_type,
                    ])

    # 格式化：冻结首行 + 筛选器 + 列宽
    for ws in [ws1, ws2]:
        ws.freeze_panes = "A2"  # 冻结首行
        # autofilter 覆盖所有有数据的列
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row > 0 and max_col > 0:
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=max_col).column_letter}{max_row}"

    # 实体级列宽
    col_widths_1 = [12, 14, 18, 10, 10, 14, 18, 35, 30, 10, 10, 10]
    for i, w in enumerate(col_widths_1, 1):
        if i <= ws1.max_column:
            ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w
    # 属性级列宽
    col_widths_2 = [12, 12, 18, 10, 16, 12, 18, 30, 18, 14, 12]
    for i, w in enumerate(col_widths_2, 1):
        if i <= ws2.max_column:
            ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # 写入
    output_path = Path(output_dir) / "mapping.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  [OK] Mapping Excel: {output_path}")
    return True


def _resolve_physical_sources(lineages, cte_index, cte_names_upper, visited, depth=0):
    """穿透 CTE 到物理源表字段。

    lineages: field.lineage 列表
    cte_index: {CTE名(UPPER): {alias_to_table, fields_map, source_tables}}
    cte_names_upper: 所有 CTE 名的大写集合
    visited: 已访问的 CTE 名（防循环）
    depth: 递归深度（最大 10）

    返回: [{schema, table, field, alias, raw_sql}, ...]
    """
    if depth > 10 or not lineages:
        return []

    results = []
    for l in lineages:
        src_table = l.get("source_table", "")
        src_field = l.get("source_field", "")
        cte_name = l.get("cte_name", "")
        raw_sql = l.get("raw_sql", "")

        if cte_name and cte_name.upper() in cte_index:
            # 来源是 CTE — 递归穿透
            cte_info = cte_index[cte_name.upper()]
            if cte_name.upper() in visited:
                continue
            visited.add(cte_name.upper())

            cte_field_info = cte_info["fields_map"].get(src_field.upper(), {})
            cte_source_fields = cte_field_info.get("source_fields", l.get("cte_source_fields", []))
            cte_alias_to_table = cte_info["alias_to_table"]

            for csf in cte_source_fields:
                csf_alias = csf.get("alias", "").upper()
                csf_field = csf.get("field", "")

                # alias → 物理表
                physical_table = cte_alias_to_table.get(csf_alias, "")

                if physical_table:
                    # 找到物理表
                    sch, tbl = _split_schema_table(physical_table)
                    results.append({
                        "schema": sch,
                        "table": tbl,
                        "field": csf_field,
                        "alias": csf.get("alias", ""),
                        "raw_sql": raw_sql,
                    })
                elif csf_alias in cte_names_upper:
                    # 嵌套 CTE — 递归穿透
                    nested_cte = cte_index.get(csf_alias, {})
                    nested_fields_map = nested_cte.get("fields_map", {})
                    nested_field = nested_fields_map.get(csf_field.upper(), {})
                    nested_sources = nested_field.get("source_fields", [])
                    nested_alias_to_table = nested_cte.get("alias_to_table", {})
                    for nsf in nested_sources:
                        nsf_alias = nsf.get("alias", "").upper()
                        nsf_table = nested_alias_to_table.get(nsf_alias, "")
                        if nsf_table:
                            sch, tbl = _split_schema_table(nsf_table)
                            results.append({
                                "schema": sch,
                                "table": tbl,
                                "field": nsf.get("field", ""),
                                "alias": nsf.get("alias", ""),
                                "raw_sql": raw_sql,
                            })

            visited.discard(cte_name.upper())
        elif src_table and src_table.upper() not in cte_names_upper:
            # 直接物理源表
            sch, tbl = _split_schema_table(src_table)
            results.append({
                "schema": sch,
                "table": tbl,
                "field": src_field,
                "alias": l.get("alias", ""),
                "raw_sql": raw_sql,
            })

    return results


# ── 视图生成: tech_design.md ────────────────────────────

def generate_tech_design(knowledge, output_dir):
    """生成技术设计文档 Markdown"""
    topo = knowledge.get("topology", {})
    df = knowledge.get("data_flow", {})
    fm = knowledge.get("field_mappings", {})
    bl = knowledge.get("business_logic", {})
    quality = knowledge.get("quality", {})
    source = knowledge.get("source", {})

    steps_list = topo.get("steps", [])
    data_flow_steps = df.get("steps", [])
    fields_list = fm.get("fields", [])
    sched_plan = topo.get("schedule_plan", [])
    self_refs = topo.get("self_references", [])

    # 目标表 schema/table 都从 exec_sequence 最大的步骤取（统一口径）
    target_schema = ""
    target_table = ""
    if steps_list:
        _max_seq = max(s.get("exec_sequence", 0) for s in steps_list)
        _max_steps = [s for s in steps_list if s.get("exec_sequence", 0) == _max_seq]
        if _max_steps:
            target_schema = _max_steps[0].get("target_schema", "")
            target_table = _max_steps[0].get("target_table", "")
    target_full = _schema_table(target_schema, target_table)

    lines = []
    lines.append(f"# {target_table} 技术设计文档")
    lines.append("")
    lines.append(f"> 由 dws-pipeline-analyzer 从制品包反向生成")
    lines.append(f"> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")

    # ── 1. 概述 ──
    lines.append("## 1. 概述")
    lines.append("")
    lines.append("| 项目 | 值 |")
    lines.append("|------|-----|")
    lines.append(f"| 目标表 | {target_full} |")
    lines.append(f"| 中文名 | {bl.get('summary', '').split('，')[0] if bl.get('summary') else '-'} |")
    lines.append(f"| 业务定位 | {bl.get('summary', '-')} |")
    lines.append(f"| 步骤数 | {len(steps_list)} |")
    lines.append(f"| 源表数 | {len(df.get('tables', []))} |")
    lines.append(f"| 字段数 | {len(fields_list)} |")
    lines.append(f"| 方言 | {knowledge.get('meta', {}).get('dialect', 'dws')} |")
    lines.append("")

    # ── 2. 复杂度分析 ──
    cm = quality.get("complexity_metrics", {})
    lines.append("## 2. 复杂度分析")
    lines.append("")
    lines.append("| 维度 | 值 |")
    lines.append("|------|-----|")
    lines.append(f"| 最大 JOIN 数 | {cm.get('max_join_count', '-')} |")
    lines.append(f"| CTE 数 | {cm.get('max_cte_count', '-')} |")
    lines.append(f"| 源表总数 | {cm.get('total_source_tables', '-')} |")
    lines.append(f"| CASE WHEN 分支 | {cm.get('total_case_when_branches', '-')} |")
    td = cm.get("transform_distribution", {})
    if td:
        dist_str = ", ".join(f"{k}={v}" for k, v in td.items())
        lines.append(f"| 转换类型分布 | {dist_str} |")
    lines.append("")

    # ── 3. 分段策略 ──
    lines.append("## 3. 分段策略")
    lines.append("")
    lines.append("| 步骤 | 规则编码 | 执行序列 | 源表 | 写入模式 |")
    lines.append("|------|---------|---------|------|---------|")
    for s in steps_list:
        wm = "TRUNCATE+INSERT" if s.get("delete_mode") == "1" else "APPEND"
        srcs = ", ".join(s.get("source_tables_from_sql", []))
        lines.append(f"| {s['step_id']} | {s.get('rule_code', '')} | {s.get('exec_sequence', 0)} | {srcs} | {wm} |")
    lines.append("")

    # 并行/串行说明
    if sched_plan:
        lines.append("**并行/串行关系:**")
        for g in sched_plan:
            seq = g.get("sequence", 0)
            psteps = ", ".join(g.get("parallel_steps", []))
            if len(g.get("parallel_steps", [])) > 1:
                lines.append(f"- 序列 {seq}: {psteps} （并行）")
            else:
                lines.append(f"- 序列 {seq}: {psteps}")
        lines.append("")

    # ── 4. 表级血缘 ──
    lines.append("## 4. 表级血缘")
    lines.append("")
    lines.append("```mermaid")
    lines.append("flowchart LR")

    # 构建 CTE 索引
    cte_index = _build_cte_index(data_flow_steps)
    cte_names_upper = set(cte_index.keys())
    target_tables_set = set()
    for s in steps_list:
        tf = _schema_table(s.get("target_schema", ""), s.get("target_table", ""))
        target_tables_set.add(_norm(tf))

    # 收集所有物理源表（含 CTE 内部表），seen_src 存归一化名用于去重
    all_phys_sources = []
    seen_src = set()
    for s in data_flow_steps:
        # 主查询 JOIN
        for j in s.get("joins", []):
            src_tbl = j.get("source_table", "")
            if src_tbl and src_tbl.upper() not in cte_names_upper and _norm(src_tbl) not in target_tables_set and _norm(src_tbl) not in seen_src:
                seen_src.add(_norm(src_tbl))
                all_phys_sources.append(src_tbl)
        # CTE 内部物理表
        for cte in s.get("ctes", []):
            for st in cte.get("source_tables", []):
                tname = st.get("name", "")
                if tname and tname.upper() not in cte_names_upper and _norm(tname) not in target_tables_set and _norm(tname) not in seen_src:
                    seen_src.add(_norm(tname))
                    all_phys_sources.append(tname)

    # 画物理源表 → 目标表
    target_safe = target_table.replace(".", "_")
    for src in all_phys_sources:
        src_safe = src.replace(".", "_")
        lines.append(f'    {src_safe}["{src}"] --> {target_safe}["{target_full}"]')

    # 目标表 → 下游视图
    for s in steps_list[1:]:
        tf = _schema_table(s.get("target_schema", ""), s.get("target_table", ""))
        if tf != target_full:
            tf_safe = tf.replace(".", "_")
            lines.append(f'    {target_safe}["{target_full}"] --> {tf_safe}["{tf}"]')

    # 自引用
    if self_refs:
        for sr in self_refs:
            lines.append(f'    {target_safe} -.->|自引用| {target_safe}')
    lines.append("```")
    lines.append("")

    # ── 5. 字段映射对照表 ──
    lines.append("## 5. 字段映射对照表")
    lines.append("")
    for s in steps_list:
        sid = s["step_id"]
        rc = s.get("rule_code", "")
        step_fields = [f for f in fields_list if f.get("producing_step") == sid]
        if not step_fields:
            continue

        lines.append(f"### {sid} ({rc})")
        lines.append("")
        lines.append("| # | 目标字段 | 物理源表 | 物理源字段 | 转换类型 | 映射表达式 |")
        lines.append("|---|---------|---------|-----------|---------|-----------|")
        for i, f in enumerate(step_fields, 1):
            # CTE 穿透获取物理源
            phys_sources = _resolve_physical_sources(
                f.get("lineage", []), cte_index, cte_names_upper, set()
            )
            tt = f.get('transform_type', '')
            raw_sql = ""
            if f.get("lineage"):
                raw_sql = f["lineage"][0].get("raw_sql", "")
            if len(raw_sql) > 100:
                raw_sql = raw_sql[:97] + "..."

            if phys_sources:
                ps0 = phys_sources[0]
                lines.append(f"| {i} | {f.get('target_field', '')} | {ps0.get('table', '-')} | {ps0.get('field', '-')} | {tt} | {raw_sql} |")
            else:
                lines.append(f"| {i} | {f.get('target_field', '')} | - | - | {tt} | {raw_sql} |")
        lines.append("")

    # ── 6. 数据处理逻辑 ──
    lines.append("## 6. 数据处理逻辑")
    lines.append("")
    for s in data_flow_steps:
        sid = s.get("step_id", "")
        ai_step = next(
            (d for d in bl.get("step_descriptions", []) if d.get("step_id") == sid), {}
        )
        lines.append(f"### {sid}: {ai_step.get('purpose', '')}")
        if ai_step.get("logic"):
            lines.append(f"- **加工逻辑**: {ai_step['logic']}")
        where = s.get("where_clause", "")
        if where:
            lines.append(f"- **过滤条件**: {where}")
        group_by = s.get("group_by", [])
        if group_by:
            lines.append(f"- **分组**: {', '.join(group_by)}")
        lines.append("")
        lines.append("```sql")
        lines.append(s.get("raw_sql", ""))
        lines.append("```")
        lines.append("")

    # ── 7. 质量评估 ──
    lines.append("## 7. 质量评估")
    lines.append("")
    issues = quality.get("issues", [])
    ai_insights = quality.get("ai_insights", [])
    if issues:
        lines.append("### 检测到的问题")
        lines.append("")
        lines.append("| 级别 | 类别 | 描述 |")
        lines.append("|------|------|------|")
        for iss in issues:
            lines.append(f"| {iss.get('severity', '')} | {iss.get('category', '')} | {iss.get('title', '')} |")
        lines.append("")

    if ai_insights:
        lines.append("### AI 建议")
        lines.append("")
        for ins in ai_insights:
            lines.append(f"- **[{ins.get('severity', '')}]** {ins.get('title', '')}")
            if ins.get("detail"):
                lines.append(f"  - {ins['detail']}")
            if ins.get("suggestion"):
                lines.append(f"  - 建议: {ins['suggestion']}")
        lines.append("")

    # ── 8. 上游任务依赖 ──
    lines.append("## 8. 上游任务依赖")
    lines.append("")
    lines.append("| 源表 | 别名 | 关联方式 | CTE归属 | 调度任务 |")
    lines.append("|------|------|---------|---------|---------|")
    seen_dep = set()
    for s in data_flow_steps:
        # 主查询 JOIN
        for j in s.get("joins", []):
            src = j.get("source_table", "")
            if _norm(src) in seen_dep or src.upper() in cte_names_upper:
                continue
            seen_dep.add(_norm(src))
            jt = j.get("join_type", "")
            lines.append(f"| {src} | {j.get('alias', '-')} | {_format_join(jt, src)} | - | 待配置 |")
        # CTE 内部物理表
        for cte in s.get("ctes", []):
            cte_name = cte.get("name", "")
            for st in cte.get("source_tables", []):
                tname = st.get("name", "")
                if _norm(tname) in seen_dep or tname.upper() in cte_names_upper:
                    continue
                seen_dep.add(_norm(tname))
                jt = st.get("join_type", "FROM")
                lines.append(f"| {tname} | {st.get('alias', '-')} | {_format_join(jt, tname)} | {cte_name} | 待配置 |")
    lines.append("")

    # ── 9. 执行平台配置 ──
    lines.append("## 9. 执行平台配置")
    lines.append("")
    raw_rules = source.get("rule_sheet_raw", [])
    if raw_rules:
        lines.append("| 配置项 | 值 |")
        lines.append("|--------|-----|")
        r0 = raw_rules[0] if raw_rules else {}
        config_map = {
            "项目编码": r0.get("project_code", ""),
            "数据源": r0.get("data_source", ""),
        }
        for k, v in config_map.items():
            lines.append(f"| {k} | {v or '待配置'} |")
    else:
        lines.append("*无执行平台配置信息*")
    lines.append("")

    # 写入
    output_path = Path(output_dir) / "tech_design.md"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8", newline="\n")
    print(f"  [OK] 技术设计文档: {output_path}")
    return True


# ── CLI ──────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="dws-pipeline-analyzer 视图生成器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python run.py view_generator --input knowledge_draft.json --output docs/output/table/
  python run.py view_generator --input knowledge_draft.json --output docs/output/table/ --views mapping,asset
        """,
    )
    parser.add_argument("--input", required=True, help="knowledge_draft.json 路径")
    parser.add_argument("--ai-input", default=None, help="knowledge_ai.md 路径（AI 增强结果，可选）")
    parser.add_argument("--output", required=True, help="输出目录")
    parser.add_argument(
        "--views",
        default="all",
        help="要生成的视图，逗号分隔: mapping,asset,techspec (默认: all)",
    )

    args = parser.parse_args()

    # 读取 knowledge
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 输入文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        knowledge = json.load(f)

    # 合并 AI 增强结果（可选）
    if args.ai_input:
        ai_path = Path(args.ai_input)
        if ai_path.exists():
            ai_text = ai_path.read_text(encoding="utf-8")
            _merge_ai_markdown(knowledge, ai_text)
            print(f"  [OK] 已合并 AI 增强: {ai_path}")
        else:
            print(f"  [WARN] AI 输入文件不存在: {ai_path}（跳过）")

    # 输出目录: 直接用用户指定的 output 目录
    views_dir = Path(args.output)
    views_dir.mkdir(parents=True, exist_ok=True)
    views_str = args.views.strip().lower()
    if views_str == "all":
        views = ["mapping", "asset", "techspec"]
    else:
        views = [v.strip() for v in views_str.split(",") if v.strip()]

    print(f"=== dws-pipeline-analyzer 视图生成器 ===")
    print(f"输入: {input_path}")
    print(f"输出: {views_dir}")
    print(f"视图: {', '.join(views)}")
    print()

    results = {}
    for view in views:
        if view == "mapping":
            results["mapping"] = generate_mapping(knowledge, str(views_dir))
        elif view == "asset":
            results["asset"] = generate_asset_report(knowledge, str(views_dir))
        elif view == "techspec":
            results["techspec"] = generate_tech_design(knowledge, str(views_dir))
        else:
            print(f"  警告: 未知视图类型 '{view}'，跳过", file=sys.stderr)

    print()
    success = sum(1 for v in results.values() if v)
    total = len(results)
    print(f"=== 完成: {success}/{total} 视图生成成功 ===")

    if success < total:
        sys.exit(1)


if __name__ == "__main__":
    main()
