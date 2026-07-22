#!/usr/bin/env python3
"""dws-pipeline-analyzer — 制品包分析器（数据层 + CLI）。

三层架构（详见 architecture.md）：
    ① 数据层（analyzer.py）— 本模块：read_excel / CLI
    ② 理解引擎（engine.py）— SQL 理解与血缘解析
    ③ 任务层 — 文档化 / 字段检索 / 关联影响分析 / ...

本模块职责：
    - read_excel()  读取 execution_tasks.xlsx → rules / target_fields / 配置
    - main()        CLI 入口，调 engine.analyze_pipeline + 写文件
    - _generate_ai_summary()  AI 兜底摘要生成

Usage:
    python analyzer.py --input execution_tasks.xlsx --output docs/output/
    python analyzer.py --input execution_tasks.xlsx --output docs/ --ddl-dir ddl/

Author: 院博
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

# Windows UTF-8
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl。pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import yaml  # noqa: F401（read_yml 内部 import，这里只做启动检查）
except ImportError:
    print("提示: PyYAML 未安装，代码仓 yml 输入将不可用。pip install pyyaml", file=sys.stderr)

# 引擎层（单向依赖：analyzer → engine）
import re
from engine import (
    analyze_pipeline,
    detect_dialect, parse_single_sql,
    ParsedSQL, RawRule,
    SELECT_RULE_TYPES, RULE_TYPE_MAP,
    _strip_dws_clauses, _replace_placeholders,
    _normalize_table_name, _norm_table, _is_intermediate_table,
    _table_match, _infer_layer, _clean_name,
)

# ═══════════════════════════════════════════════════════════════
# Excel 列映射（read_excel 专用）
# ═══════════════════════════════════════════════════════════════

RULE_COLUMNS_MAP = {
    "rule_code": "规则编码",
    "exec_sequence": "执行序列",
    "target_schema": "目标Schema",
    "target_table": "目标表",
    "delete_mode": "删除模式",
    "delete_condition": "删除条件",
    "query_sql": "(生成的）查询语句1",
    "project_code": "项目编码",
    "data_source": "数据源",
    "business_owner": "业务责任人",
    "rule_group_code": "规则组编码",
    "rule_group_en": "规则组英文名称",
    "rule_type": "规则类型",
    "rule_name": "规则中文名称",
    "exchange_source_table": "交换分区来源表",
}

# 规则类型语义映射
RULE_TYPE_MAP = {
    1: "取数规则",
    2: "删数规则",
    3: "备份规则",
    4: "查询规则",
    5: "逻辑视图",
    6: "物理视图",
    7: "度量规则",
    8: "物理表规则",
    9: "分区交换",
    10: "SP规则",
    11: "API规则",
    12: "参数变量",
    13: "维护类",
    14: "Spark取数",
    15: "判断类",
}

# SELECT 类规则（需要完整解析 SQL + 字段映射 + 血缘）
SELECT_RULE_TYPES = {1, 14}

# 记录类规则（不解析 SQL，但记录操作信息）
RECORD_RULE_TYPES = {2, 9}

# 参数变量（记录到 variables，不算 step）
VARIABLE_RULE_TYPES = {12}

# 删除模式语义映射
DELETE_MODE_MAP = {
    "1": "TRUNCATE TABLE",
    "2": "NO DELETE (追加)",
    "3": "TRUNCATE SUBPARTITION",
    "4": "DELETE",
    "5": "TRUNCATE PARTITION",
    "6": "MERGE INTO",
    "7": "RPT_ITEM",
}

# 分区级删除模式（有分区场景标识）
PARTITION_DELETE_MODES = {"3", "5"}

# TargetFields sheet 列名
TF_COLUMNS_MAP = {
    "rule_code": "规则编码",
    "target_field": "目标字段名称",
    "source_field": "来源字段名称",
    "encryption": "加密方式",
    "alias": "别名",
    "field_type": "字段类型",
    "remark": "备注",
}

# GroupVariables sheet 列名
GV_COLUMNS_MAP = {
    "rule_code": "规则编码",
    "var_name": "动态参数/变量名",
    "default_value": "变量默认值",
}

# ═══════════════════════════════════════════════════════════════
# read_excel 专用辅助函数（其余工具函数已在 engine.py）
# ═══════════════════════════════════════════════════════════════

def _safe_str(val) -> str:
    """安全转字符串"""
    if val is None:
        return ""
    return str(val).strip()


def _find_col(col_idx: dict, name: str) -> int | None:
    """查找列索引。先精确匹配，再模糊匹配（去空格+全角半角归一化）。"""
    # 精确匹配
    if name in col_idx:
        return col_idx[name]
    # 模糊匹配：去空格、全角括号转半角
    def normalize(s):
        # 全角括号 → 半角
        s = s.replace("（", "(").replace("）", ")")
        # 去所有空格
        s = s.replace(" ", "").replace("\u3000", "")
        return s
    norm_name = normalize(name)
    for actual, idx in col_idx.items():
        if normalize(actual) == norm_name:
            return idx
    # 包含匹配（期望列名是实际列名的子串，或反过来）
    for actual, idx in col_idx.items():
        na = normalize(actual)
        if norm_name in na or na in norm_name:
            return idx
    return None


def _get_val(row: tuple, idx: int | None) -> str:
    """安全获取行值"""
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return _safe_str(val)


def _read_query_sql(row: tuple, col_idx: dict) -> str:
    """读取查询语句（支持多列拼接）。

    超长 SQL 会分散在「查询语句1」「查询语句2」... 多列。
    按列序号拼接非空内容，最后一个非空列去掉末尾 \\r\\n。
    不 strip 中间列（保留列内的空格和换行，避免拼接时 SQL 断裂）。
    """
    import re as _re
    # 找所有「查询语句N」列，按 N 排序
    sql_cols = []
    for col_name, idx in col_idx.items():
        m = _re.match(r'.*查询语句\s*(\d+)', str(col_name))
        if m:
            sql_cols.append((int(m.group(1)), idx))
    sql_cols.sort()

    if not sql_cols:
        return ""

    # 直接读取原始值（不走 _get_val 的 strip），保留列内空格换行
    parts = []
    for _, idx in sql_cols:
        if idx is None or idx >= len(row):
            continue
        val = row[idx]
        if val is None:
            continue
        val_str = str(val)
        if val_str.strip():  # 整列只有空白/换行则跳过，但保留非空内容
            parts.append(val_str)

    if not parts:
        return ""

    # 完整拼接，保留所有原始字符（包括末尾 \r\n，它是完整 SQL 的一部分）
    return "".join(parts)


# ═══════════════════════════════════════════════════════════════
# Step 1: read_excel()
# ═══════════════════════════════════════════════════════════════

def read_excel(excel_path: str) -> dict:
    """读取制品包 Excel，返回结构化数据。

    Returns: {
        "rules": [RawRule, ...],
        "target_fields": {"规则编码": [RawTargetField, ...]},
        "group_variables": {"规则编码": [RawGroupVariable, ...]},
        "variables": ["P_CYCLE_ID", ...],
        "rule_group_code": "GR123456",
    }
    """
    # read_only=False 更可靠（部分 Excel 文件在 read_only=True 时列读取不完整）
    wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
    result = {
        "rules": [],
        "target_fields": {},
        "group_variables": {},
        "variables": [],
        "rule_group_code": "",
        "rule_group_en": "",
    }

    # ── RULE sheet ──
    if "RULE" not in wb.sheetnames:
        print("错误: Excel 中没有 RULE sheet", file=sys.stderr)
        wb.close()
        return result

    ws = wb["RULE"]
    col_idx = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False)):
        if cell.value:
            col_idx[cell.value.strip()] = cell.column - 1

    # 映射列名到索引
    ci = {k: _find_col(col_idx, v) for k, v in RULE_COLUMNS_MAP.items()}
    col_rule_type = ci.get("rule_type")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        rule_type_str = _get_val(row, col_rule_type)
        try:
            rt = int(float(rule_type_str)) if rule_type_str else 0
        except (ValueError, TypeError):
            rt = 0

        query = _read_query_sql(row, col_idx)
        exec_seq_str = _get_val(row, ci.get("exec_sequence"))
        try:
            # int(float()) 兼容数值、字符串 "1"、字符串 "1.0" 三种格式
            exec_seq = int(float(exec_seq_str)) if exec_seq_str else 0
        except (ValueError, TypeError):
            exec_seq = 0

        # 类型 12（参数变量）→ 记录到 variables
        if rt in VARIABLE_RULE_TYPES:
            var_name = _get_val(row, ci.get("rule_name")) or _get_val(row, ci.get("rule_code"))
            if var_name:
                result["variables"].append(var_name)
            continue

        # 类型 10/11/13/15（SP/API/维护/判断）→ 完全跳过
        if rt in {10, 11, 13, 15}:
            continue

        # 类型 1/14（取数/Spark取数）→ 完整解析，必须有 SQL
        # 类型 2/9（删数/分区交换）→ 记录操作，SQL 可选
        # 其他类型 → 记录但不解析

        rule = RawRule(
            rule_code=_get_val(row, ci.get("rule_code")),
            rule_name=_get_val(row, ci.get("rule_name")),
            rule_type=rt,
            exec_sequence=exec_seq,
            target_schema=_get_val(row, ci.get("target_schema")),
            target_table=_get_val(row, ci.get("target_table")),
            delete_mode=_get_val(row, ci.get("delete_mode")),
            delete_condition=_get_val(row, ci.get("delete_condition")),
            query_sql=query.strip() if query else "",
            project_code=_get_val(row, ci.get("project_code")),
            data_source=_get_val(row, ci.get("data_source")),
            business_owner=_get_val(row, ci.get("business_owner")),
            rule_group_code=_get_val(row, ci.get("rule_group_code")),
            rule_group_en=str(_get_val(row, ci.get("rule_group_en")) or "").strip(),
            exchange_source_table=_get_val(row, ci.get("exchange_source_table")),
        )

        # SELECT 类规则必须有 SQL
        if rt in SELECT_RULE_TYPES and not rule.query_sql:
            continue

        result["rules"].append(rule)

        if rule.rule_group_code and not result["rule_group_code"]:
            result["rule_group_code"] = rule.rule_group_code

        # 规则组英文名称（取第一个非空值，作为输出目录名）
        rule_group_en = _get_val(row, ci.get("rule_group_en"))
        if rule_group_en and not result["rule_group_en"]:
            result["rule_group_en"] = str(rule_group_en).strip()

    # ── TargetFields sheet ──
    if "TargetFields" in wb.sheetnames:
        ws_tf = wb["TargetFields"]
        tf_col_idx = {}
        for cell in next(ws_tf.iter_rows(min_row=1, max_row=1, values_only=False)):
            if cell.value:
                tf_col_idx[cell.value.strip()] = cell.column - 1

        tf_ci = {k: _find_col(tf_col_idx, v) for k, v in TF_COLUMNS_MAP.items()}

        for tf_row in ws_tf.iter_rows(min_row=2, values_only=True):
            if not tf_row:
                continue
            rc = _get_val(tf_row, tf_ci.get("rule_code"))
            tf = RawTargetField(
                rule_code=rc,
                target_field=_get_val(tf_row, tf_ci.get("target_field")),
                source_field=_get_val(tf_row, tf_ci.get("source_field")),
                encryption=_get_val(tf_row, tf_ci.get("encryption")),
                alias=_get_val(tf_row, tf_ci.get("alias")),
                field_type=_get_val(tf_row, tf_ci.get("field_type")),
                remark=_get_val(tf_row, tf_ci.get("remark")),
            )
            if rc:
                result["target_fields"].setdefault(rc, []).append(tf)

    # ── GroupVariables sheet ──
    if "GroupVariables" in wb.sheetnames:
        ws_gv = wb["GroupVariables"]
        gv_col_idx = {}
        for cell in next(ws_gv.iter_rows(min_row=1, max_row=1, values_only=False)):
            if cell.value:
                gv_col_idx[cell.value.strip()] = cell.column - 1

        gv_ci = {k: _find_col(gv_col_idx, v) for k, v in GV_COLUMNS_MAP.items()}
        all_vars = set()

        for gv_row in ws_gv.iter_rows(min_row=2, values_only=True):
            if not gv_row:
                continue
            rc = _get_val(gv_row, gv_ci.get("rule_code"))
            var_name = _get_val(gv_row, gv_ci.get("var_name"))
            gv = RawGroupVariable(
                rule_code=rc,
                var_name=var_name,
                default_value=_get_val(gv_row, gv_ci.get("default_value")),
            )
            if rc:
                result["group_variables"].setdefault(rc, []).append(gv)
            if var_name:
                all_vars.add(var_name)

        result["variables"] = sorted(all_vars)

    wb.close()
    return result


# ═══════════════════════════════════════════════════════════════
# Step 1b: read_yml() — 代码仓 yml 加载（和 read_excel 产出完全一致）
# ═══════════════════════════════════════════════════════════════

# yml key → RawRule 字段的映射（和 RULE_COLUMNS_MAP 对应，值是 yml 里的中文 key）
# 注意查询语句的 key 容错：yml 用半角括号，Excel 用全角括号+后缀1
_YML_RULE_KEY_ALIASES = {
    "rule_code": ["规则编码"],
    "rule_name": ["规则中文名称"],
    "rule_type": ["规则类型"],
    "exec_sequence": ["执行序列"],
    "target_schema": ["目标Schema", "目标schema", "目标SCHEMA"],
    "target_table": ["目标表"],
    "delete_mode": ["删除模式"],
    "delete_condition": ["删除条件"],
    "query_sql": ["(生成的)查询语句", "(生成的）查询语句", "(生成的)查询语句1", "(生成的）查询语句1"],
    "project_code": ["项目编码"],
    "data_source": ["数据源"],
    "business_owner": ["业务责任人"],
    "rule_group_code": ["规则组编码"],
    "rule_group_en": ["规则组英文名称"],
    "exchange_source_table": ["交换分区来源表"],
}

_YML_TF_KEY_ALIASES = {
    "rule_code": ["规则编码"],
    "target_field": ["目标字段名称"],
    "source_field": ["来源字段名称"],
    "encryption": ["加密方式"],
    "alias": ["别名"],
    "field_type": ["字段类型"],
    "remark": ["备注"],
}

_YML_GV_KEY_ALIASES = {
    "rule_code": ["规则编码"],
    "var_name": ["动态参数/变量名", "变量名"],
    "default_value": ["变量默认值", "默认值"],
}


def _yml_get(d: dict, field: str, aliases: dict) -> str:
    """从 yml dict 按 alias 列表取值（容错多种 key 写法），返回字符串。"""
    for key in aliases.get(field, []):
        if key in d:
            val = d[key]
            return "" if val is None else str(val)
    return ""


def _parse_int(val: str) -> int:
    """字符串转 int（兼容 '1'/'1.0'/1 等格式，和 read_excel 的转换一致）。"""
    try:
        return int(float(val)) if val else 0
    except (ValueError, TypeError):
        return 0


def _yml_read_query_sql(data: dict) -> str:
    """从 yml dict 读取查询语句（支持多段拼接，和 read_excel 的 _read_query_sql 对齐）。

    真实 yml 里超长 SQL 会被拆成多条（和 Excel 多列一样）：
        (生成的)查询语句1: |-   ← 第一段
          WITH agg AS (...
        (生成的)查询语句2: |-   ← 第二段
          FROM ods.t WHERE...)

    本函数找所有「查询语句N」key，按 N 排序拼接。
    key 格式容错：半角/全角括号、有无后缀数字、| literal block scalar。
    """
    import re
    sql_parts = []  # [(序号, 内容)]
    for key, val in data.items():
        key_str = str(key)
        # 匹配「查询语句N」或「查询语句」（无数字视为1）
        m = re.match(r'.*查询语句\s*(\d*)', key_str)
        if m:
            num = int(m.group(1)) if m.group(1) else 1
            val_str = "" if val is None else str(val)
            if val_str.strip():
                sql_parts.append((num, val_str))
    if not sql_parts:
        return ""
    sql_parts.sort(key=lambda x: x[0])
    return "".join(part for _, part in sql_parts)


def read_yml(yml_dir: str) -> dict:
    """读取代码仓规则组目录下的 yml 文件，返回和 read_excel 完全一致的结构。

    一个规则组目录下有多个 *.yml 文件（一个 yml = 一条规则）。
    本函数遍历目录下所有 yml，合并为一个规则组，产出结构同 read_excel：
        {rules, target_fields, group_variables, variables, rule_group_code, rule_group_en}

    yml 格式（详见 sample_rule.yml）：
        顶层 = RULE sheet 字段（中文 key）
        额外信息（其他sheet页信息）.TargetFields = TargetFields sheet
        额外信息（其他sheet页信息）.GroupVariables = GroupVariables sheet
    """
    import yaml

    yml_path = Path(yml_dir)
    result = {
        "rules": [],
        "target_fields": {},
        "group_variables": {},
        "variables": [],
        "rule_group_code": "",
        "rule_group_en": "",
    }

    # 收集目录下所有 yml 文件（一个 yml = 一条规则）
    yml_files = sorted(yml_path.glob("*.yml")) + sorted(yml_path.glob("*.yaml"))
    if not yml_files:
        print(f"错误: 目录下没有 yml 文件: {yml_dir}", file=sys.stderr)
        return result

    all_vars = set()

    for yf in yml_files:
        try:
            data = yaml.safe_load(yf.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  [yml解析错误] {yf.name}: {e}", file=sys.stderr)
            continue
        if not data or not isinstance(data, dict):
            continue

        # ── 解析 RULE 主信息 → RawRule ──
        rt = _parse_int(_yml_get(data, "rule_type", _YML_RULE_KEY_ALIASES))

        # 类型 12（参数变量）→ 记录到 variables，不作为规则
        if rt in VARIABLE_RULE_TYPES:
            var_name = _yml_get(data, "rule_name", _YML_RULE_KEY_ALIASES) or \
                       _yml_get(data, "rule_code", _YML_RULE_KEY_ALIASES)
            if var_name:
                all_vars.add(var_name)
            continue

        # 类型 10/11/13/15（SP/API/维护/判断）→ 跳过（和 read_excel 一致）
        if rt in {10, 11, 13, 15}:
            continue

        rule = RawRule(
            rule_code=_yml_get(data, "rule_code", _YML_RULE_KEY_ALIASES),
            rule_name=_yml_get(data, "rule_name", _YML_RULE_KEY_ALIASES),
            rule_type=rt,
            exec_sequence=_parse_int(_yml_get(data, "exec_sequence", _YML_RULE_KEY_ALIASES)),
            target_schema=_yml_get(data, "target_schema", _YML_RULE_KEY_ALIASES),
            target_table=_yml_get(data, "target_table", _YML_RULE_KEY_ALIASES),
            delete_mode=_yml_get(data, "delete_mode", _YML_RULE_KEY_ALIASES),
            delete_condition=_yml_get(data, "delete_condition", _YML_RULE_KEY_ALIASES),
            query_sql=_yml_read_query_sql(data).strip(),
            project_code=_yml_get(data, "project_code", _YML_RULE_KEY_ALIASES),
            data_source=_yml_get(data, "data_source", _YML_RULE_KEY_ALIASES),
            business_owner=_yml_get(data, "business_owner", _YML_RULE_KEY_ALIASES),
            rule_group_code=_yml_get(data, "rule_group_code", _YML_RULE_KEY_ALIASES),
            rule_group_en=_yml_get(data, "rule_group_en", _YML_RULE_KEY_ALIASES).strip(),
            exchange_source_table=_yml_get(data, "exchange_source_table", _YML_RULE_KEY_ALIASES),
        )

        # SELECT 类规则必须有 SQL
        if rt in SELECT_RULE_TYPES and not rule.query_sql:
            continue

        result["rules"].append(rule)

        if rule.rule_group_code and not result["rule_group_code"]:
            result["rule_group_code"] = rule.rule_group_code
        if rule.rule_group_en and not result["rule_group_en"]:
            result["rule_group_en"] = rule.rule_group_en

        # ── 解析额外信息（TargetFields / GroupVariables，有则读，无则跳过）──
        # 容错：真实 yml 里额外信息/TargetFields/GroupVariables 的冒号后可能有 |
        # （literal block scalar），导致 YAML 把它们解析成字符串而非 dict/list。
        # 遇到字符串时再 yaml.safe_load 一次，解析出真实结构。
        extra = data.get("额外信息（其他sheet页信息）") or data.get("额外信息") or {}
        if isinstance(extra, str):
            try:
                extra = yaml.safe_load(extra) or {}
            except Exception:
                extra = {}
        if not isinstance(extra, dict):
            extra = {}

        rc = rule.rule_code

        # TargetFields（值可能因 | 变成字符串，需再解析）
        tf_list = extra.get("TargetFields") or []
        if isinstance(tf_list, str):
            try:
                tf_list = yaml.safe_load(tf_list) or []
            except Exception:
                tf_list = []
        if isinstance(tf_list, list):
            for tf_item in tf_list:
                if not isinstance(tf_item, dict):
                    continue
                tf = RawTargetField(
                    rule_code=_yml_get(tf_item, "rule_code", _YML_TF_KEY_ALIASES) or rc,
                    target_field=_yml_get(tf_item, "target_field", _YML_TF_KEY_ALIASES),
                    source_field=_yml_get(tf_item, "source_field", _YML_TF_KEY_ALIASES),
                    encryption=_yml_get(tf_item, "encryption", _YML_TF_KEY_ALIASES),
                    alias=_yml_get(tf_item, "alias", _YML_TF_KEY_ALIASES),
                    field_type=_yml_get(tf_item, "field_type", _YML_TF_KEY_ALIASES),
                    remark=_yml_get(tf_item, "remark", _YML_TF_KEY_ALIASES),
                )
                tf_rc = tf.rule_code or rc
                if tf_rc:
                    result["target_fields"].setdefault(tf_rc, []).append(tf)

        # GroupVariables（值同样可能因 | 变成字符串，需再解析）
        gv_list = extra.get("GroupVariables") or []
        if isinstance(gv_list, str):
            try:
                gv_list = yaml.safe_load(gv_list) or []
            except Exception:
                gv_list = []
        if isinstance(gv_list, list):
            for gv_item in gv_list:
                if not isinstance(gv_item, dict):
                    continue
                gv = RawGroupVariable(
                    rule_code=_yml_get(gv_item, "rule_code", _YML_GV_KEY_ALIASES) or rc,
                    var_name=_yml_get(gv_item, "var_name", _YML_GV_KEY_ALIASES),
                    default_value=_yml_get(gv_item, "default_value", _YML_GV_KEY_ALIASES),
                )
                gv_rc = gv.rule_code or rc
                if gv_rc:
                    result["group_variables"].setdefault(gv_rc, []).append(gv)
                if gv.var_name:
                    all_vars.add(gv.var_name)

    result["variables"] = sorted(all_vars)
    return result


def _find_repo_root(start_dir: Path) -> Path | None:
    """从 start_dir 逐级向上，找到代码仓根（含 BFT/ + DDL/ 的目录）。

    代码仓根目录下有 BFT/DDL/DQ/LTS/ADMS/Release 等目录，且这些目录名在
    根目录下一层唯一（无同名），向上探测可靠零误判。
    """
    current = start_dir.resolve()
    for parent in [current] + list(current.parents):
        if (parent / "BFT").is_dir() and (parent / "DDL").is_dir():
            return parent
    return None


def _auto_discover_ddl_from_repo(yml_dir: Path, rules: list) -> str:
    """从代码仓结构自动发现目标表的 DDL 文件路径。

    代码仓 DDL 目录结构：DDL/{DWS_EDW|DWS_RT_EDW}/{schema}/table/{target_table}.sql
    从 yml 目录向上找仓根，再按 target_schema + target_table 定位 DDL。
    两层（DWS_EDW / DWS_RT_EDW）都试，找到第一个就返回其父目录（parse_ddl 接收目录）。

    Returns: DDL 文件所在目录路径（供 parse_ddl_for_metadata 扫描）；找不到返回 ""。
    """
    repo_root = _find_repo_root(yml_dir)
    if not repo_root:
        return ""

    # 取目标表信息，优先级：
    # 1. 交换分区的 exchange_source_table（交换分区时真正目标表）
    # 2. 最后一个非中间表、非视图步骤的 target_table
    # 3. 兜底：最后一个规则的 target_table
    target_schema = ""
    target_table = ""
    from engine import _is_intermediate_table
    # 先找交换分区的 exchange_source_table
    for rule in reversed(rules):
        if rule.rule_type == 9 and rule.exchange_source_table:
            target_schema = rule.target_schema
            target_table = rule.exchange_source_table
            break
    # 没有交换分区，找最后一个非中间表
    if not target_table:
        for rule in reversed(rules):
            if getattr(rule, "is_view_step", False):
                continue
            if not _is_intermediate_table(rule.target_table):
                target_schema = rule.target_schema
                target_table = rule.target_table
                break
    # 兜底
    if not target_table and rules:
        target_schema = rules[-1].target_schema
        target_table = rules[-1].target_table
    if not target_table:
        return ""

    table_lower = target_table.lower()
    schema_lower = target_schema.lower() if target_schema else ""

    # 两层都试：DWS_EDW（离线）/ DWS_RT_EDW（实时）
    for layer in ("DWS_EDW", "DWS_RT_EDW"):
        ddl_root = repo_root / "DDL" / layer
        if not ddl_root.is_dir():
            continue
        # schema 目录可能是 dws / DWS 等，大小写容错
        schema_dir = None
        if schema_lower:
            for sd in ddl_root.iterdir():
                if sd.is_dir() and sd.name.lower() == schema_lower:
                    schema_dir = sd
                    break
        if not schema_dir:
            continue
        # table 目录下找目标表 DDL（包含匹配 + 多扩展名，不靠严格相等）
        # 匹配规则：文件名（去扩展名）等于表名，或文件名以"表名."开头（schema前缀），
        # 或表名是文件名的一部分（容错 create_table_xxx / xxx_v2 等前缀后缀）。
        table_dir = schema_dir / "table"
        if table_dir.is_dir():
            if _find_ddl_file(table_dir, table_lower):
                # 返回 schema 目录（而非 table 目录），parse_ddl 用 rglob 递归扫描，
                # 能同时覆盖 table/（F表DDL）和 view/（I视图DDL）
                return str(schema_dir)
    return ""


def _find_ddl_file(table_dir: Path, table_lower: str) -> Path | None:
    """在 DDL 目录里找目标表的 DDL 文件（包含匹配 + 多扩展名）。

    匹配优先级（从严到宽，避免误匹配）：
    1. 文件名(去扩展名) == 表名（最精确，如 dwb_trade.sql 匹配 dwb_trade）
    2. 文件名以"表名."开头（schema 前缀，如 ods.dwb_trade.sql 匹配 dwb_trade）
    3. 文件名(去扩展名) 包含表名（容错 create_table_dwb_trade / dwb_trade_v2）

    支持的扩展名：.sql / .ddl / .txt（不区分大小写）
    """
    candidates = []
    for ext in ("*.sql", "*.ddl", "*.txt", "*.SQL", "*.DDL", "*.TXT"):
        candidates.extend(table_dir.glob(ext))
    # 去重（大小写不同的 glob 可能重复）
    seen = set()
    files = []
    for f in candidates:
        real = str(f.resolve())
        if real not in seen:
            seen.add(real)
            files.append(f)

    for f in files:  # 优先级1：精确匹配
        if f.stem.lower() == table_lower:
            return f
    for f in files:  # 优先级2：schema 前缀（文件名以"表名."开头）
        if f.name.lower().startswith(table_lower + "."):
            return f
    for f in files:  # 优先级3：包含（容错前缀后缀，但要求表名足够长避免误匹配）
        if len(table_lower) >= 4 and table_lower in f.stem.lower():
            return f
    return None


# ═══════════════════════════════════════════════════════════════
# I 视图发现（资产是 I 视图，F 表是底表，分析链路需补 F→I 这一段）
# ═══════════════════════════════════════════════════════════════

def _extract_view_sql(view_content: str) -> str:
    """从 CREATE VIEW 文件内容里提取 SELECT 语句。

    CREATE VIEW xxx AS SELECT ... → 返回 SELECT ...
    CREATE OR REPLACE VIEW xxx AS SELECT ... → 同上

    只提取 AS 后的 SELECT 语句，截断后续的 COMMENT/CREATE 等语句
    （视图 DDL 文件常含 COMMENT ON COLUMN，不能混入 query_sql）。
    """
    import re
    # 匹配 CREATE [OR REPLACE] VIEW xxx AS 后面的内容
    m = re.search(
        r'CREATE\s+(?:OR\s+REPLACE\s+)?VIEW\s+\S+\s+AS\s+(.*)',
        view_content, re.IGNORECASE | re.DOTALL
    )
    if not m:
        return ""
    sql = m.group(1).strip()
    # 先剔除注释（行注释 -- 和块注释 /* */），避免注释里的分号
    # 导致 find(";") 提前截断，丢失后续 JOIN 的表名
    sql = re.sub(r"--[^\n]*", "", sql)            # 行注释
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)  # 块注释
    # 截断到第一个独立语句结束（SELECT 的分号），去掉后续的 COMMENT/CREATE 等
    first_semicolon = sql.find(";")
    if first_semicolon >= 0:
        sql = sql[:first_semicolon]
    return sql.strip()


def _find_view_by_name(view_dir: Path, view_name_lower: str) -> Path | None:
    """在 view/ 目录里按视图名找 DDL 文件（复用 _find_ddl_file 的匹配逻辑）。"""
    return _find_ddl_file(view_dir, view_name_lower)


def _search_view_by_source(view_dirs: list, f_table_full: str) -> tuple | None:
    """全局搜索 view/ 目录，找 CREATE VIEW ... FROM ... f_table 的视图。

    遍历所有 view_dir 下的 .sql/.ddl 文件，匹配 FROM 子句引用了 f_table 的。
    用于命名不规律的 I 视图（Step 2 兜底）。

    Returns: (view_file_path, view_name) 或 None
    """
    import re
    f_table_lower = f_table_full.lower()
    # 也匹配不带 schema 的表名（FROM dwb_xxx_f / FROM dws.dwb_xxx_f）
    f_table_short = f_table_lower.split(".")[-1]

    for view_dir in view_dirs:
        if not view_dir.is_dir():
            continue
        for ext in ("*.sql", "*.ddl", "*.SQL", "*.DDL"):
            for vf in view_dir.glob(ext):
                try:
                    content = vf.read_text(encoding="utf-8", errors="ignore")
                except Exception:
                    continue
                content_lower = content.lower()
                if "create view" not in content_lower and "create or replace view" not in content_lower:
                    continue
                # 检查 FROM 子句是否引用了 F 表
                if f_table_lower in content_lower or f_table_short in content_lower:
                    # 提取视图名（CREATE VIEW 视图名 AS）
                    nm = re.search(
                        r'CREATE\s+(?:OR\s+REPLACE\s+)?VIEW\s+(\S+)',
                        content, re.IGNORECASE
                    )
                    view_name = nm.group(1) if nm else vf.stem
                    return (vf, view_name)
    return None


def discover_i_view(yml_dir: Path, f_schema: str, f_table: str) -> dict | None:
    """发现 F 表对应的 I 视图（两步走策略）。

    资产是 I 视图（dwb_xxx_i），F 表是底表。此函数找到 I 视图的 CREATE VIEW
    定义，返回视图信息供分析链路追加 F→I 步骤。

    两步走发现：
      Step 1: 按名字快速找（_f → _i，覆盖95%直封，快路径）
      Step 2: 名字没找到 → 全局搜索 view/ 里 FROM 引用 F 表的（兜底）

    Args:
        yml_dir: 规则组目录（用于向上找代码仓根）
        f_schema: F 表的 schema（如 dws）
        f_table: F 表名（如 dwb_trade_order_f）

    Returns: {
        "view_name": 视图名（如 dwb_trade_order_i）,
        "view_sql": SELECT 语句（从 CREATE VIEW 提取）,
        "view_schema": schema,
        "ddl_path": DDL 文件路径,
    } 或 None（找不到 I 视图，以 F 表为终点）
    """
    repo_root = _find_repo_root(yml_dir)
    if not repo_root:
        return None

    # 推导 I 视图名：_f → _i（_F → _I 大小写容错）
    # 只有 _f 结尾的表才找 I 视图（资产 = I 视图，F 表是底表）
    # 非 _f 表（如 dwb_xxx_d）不做 I 视图发现
    f_table_lower = f_table.lower()
    if f_table_lower.endswith("_f"):
        i_table = f_table[:-2] + "_i"
    else:
        return None  # 非 _f 表不找 I 视图

    i_table_lower = (i_table or "").lower()
    schema_lower = f_schema.lower() if f_schema else ""

    # 收集所有层的 view 目录
    view_dirs = []
    for layer in ("DWS_EDW", "DWS_RT_EDW"):
        ddl_root = repo_root / "DDL" / layer
        if not ddl_root.is_dir():
            continue
        # schema 目录大小写容错
        if schema_lower:
            for sd in ddl_root.iterdir():
                if sd.is_dir() and sd.name.lower() == schema_lower:
                    vd = sd / "view"
                    if vd.is_dir():
                        view_dirs.append(vd)
                    break

    # Step 1: 按名字快速找
    if i_table_lower:
        for vd in view_dirs:
            vf = _find_view_by_name(vd, i_table_lower)
            if vf:
                content = vf.read_text(encoding="utf-8", errors="ignore")
                view_sql = _extract_view_sql(content)
                if view_sql:
                    return {
                        "view_name": i_table,
                        "view_sql": view_sql,
                        "view_schema": f_schema,
                        "ddl_path": str(vf),
                    }

    # Step 2: 全局搜索来源表（兜底）
    f_full = f"{f_schema}.{f_table}" if f_schema else f_table
    result = _search_view_by_source(view_dirs, f_full)
    if result:
        vf, view_name = result
        content = vf.read_text(encoding="utf-8", errors="ignore")
        view_sql = _extract_view_sql(content)
        if view_sql:
            return {
                "view_name": view_name,
                "view_sql": view_sql,
                "view_schema": f_schema,
                "ddl_path": str(vf),
            }

    return None


def _generate_ai_summary(knowledge, rules, parsed_map, topology, field_mappings, quality, data_flow) -> str:
    """生成 AI 增强用的精简摘要 markdown。

    包含:
    - 目标表 + 场景结构
    - 每个步骤的关键信息（规则名/来源表/CTE/加工类型/SQL前200字符）
    - 字段加工类型分布
    - 质量问题
    """
    lines = []
    lines.append("# ETL 分析摘要（AI 增强用）")
    lines.append("")
    lines.append("> AI 请基于以下信息，补充每个步骤的业务目的和加工逻辑，")
    lines.append("> 每个步骤的逻辑块结构已经列出，请基于块结构推理：")
    lines.append("> 1. 每个块的业务目的（块目的）")
    lines.append("> 2. 整个步骤的业务目的（块目的的组合概括）")
    lines.append("> 输出格式见文末模板，保存为 knowledge_ai.md")
    lines.append("")

    # ── 基本信息 ──
    meta = knowledge.get("meta", {})
    lines.append(f"## 基本信息")
    lines.append(f"- 目标表: {meta.get('target_table', '')}")
    lines.append(f"- 规则数: {len(rules)}")
    lines.append(f"- 加工模式: {', '.join(p.get('label','') for p in meta.get('patterns', []))}")
    lines.append("")

    # ── 场景结构 ──
    scenarios = topology.get("scenarios", [])
    if scenarios:
        lines.append("## 场景结构")
        for sc in scenarios:
            label = "公共步骤" if sc.get("is_common") else sc["name"]
            lines.append(f"- {label}: {sc['rule_count']} 个规则 ({', '.join(sc['rule_codes'])})")
        lines.append("")

    # ── 步骤详情（分层压缩，控制总量让 AI 能完整读完）──
    lines.append("## 步骤详情")
    SUMMARY_MAX_LINES = 150  # 总量控制阈值
    for rule in rules:
        parsed = parsed_map.get(rule.rule_code)
        step = next((s for s in topology["steps"] if s["rule_code"] == rule.rule_code), None)
        sid = step["step_id"] if step else ""

        # 总量控制：超过阈值时只输出标题行（避免 AI 吃不下）
        if len(lines) > SUMMARY_MAX_LINES:
            lines.append(f"### {sid} ({rule.rule_code}) {rule.rule_name or ''}")
            lines.append("- (详情已省略，summary 超过总量阈值，请参考 knowledge_draft.json)")
            lines.append("")
            continue

        # 兜底描述
        auto_desc = next((d for d in knowledge.get("business_logic", {}).get("step_descriptions", [])
                         if d.get("step_id") == sid), {})

        lines.append(f"### {sid} ({rule.rule_code}) {rule.rule_name or ''}")

        # 判断步骤复杂度（决定展开程度）
        join_count = len([j for j in parsed.source_tables]) if parsed else 0
        cte_count = len(parsed.ctes) if parsed else 0
        is_simple = join_count <= 1 and cte_count == 0  # 单表直取 = 简单

        # 基本信息（压缩：简单步骤合并为一行）
        rt_label = RULE_TYPE_MAP.get(rule.rule_type, "")
        dm_label = DELETE_MODE_MAP.get((rule.delete_mode or "").strip(), "")
        dc = rule.delete_condition or ""
        if is_simple and parsed:
            # 简单步骤压缩：来源→目标+字段数，一行搞定
            src_tables = [j.source_table for j in parsed.source_tables]
            tt_dist = Counter(c.transform_type for c in parsed.select_columns)
            tt_str = ", ".join(f"{k}={v}" for k, v in tt_dist.most_common())
            line = f"- {rt_label}: {' → '.join(src_tables[:3])} → {rule.target_table}"
            line += f" | {len(parsed.select_columns)}列({tt_str})"
            line += f" | {dm_label}" + (f"[{dc}]" if dc else "")
            lines.append(line)
        else:
            # 复杂步骤展开
            lines.append(f"- 类型: {rt_label} | 写入: {dm_label}" + (f" [{dc}]" if dc else ""))
            if rule.rule_type == 9 and rule.exchange_source_table:
                lines.append(f"- 分区交换: {rule.target_table} → {rule.exchange_source_table}")
            if parsed:
                src_tables = [j.source_table for j in parsed.source_tables]
                lines.append(f"- 来源表: {', '.join(src_tables[:5])}")
                if parsed.ctes:
                    cte_names = [c.name for c in parsed.ctes]
                    lines.append(f"- CTE: {', '.join(cte_names)}")
                tt_dist = Counter(c.transform_type for c in parsed.select_columns)
                tt_str = ", ".join(f"{k}={v}" for k, v in tt_dist.most_common())
                lines.append(f"- 字段加工: {len(parsed.select_columns)} 列 ({tt_str})")

                # SQL 关键特征（替代原来的200字符前缀，信息密度更高）
                features = []
                if join_count > 1:
                    features.append(f"{join_count} JOIN")
                if cte_count:
                    features.append(f"{cte_count} CTE")
                where_count = len(parsed.where_usage) if hasattr(parsed, 'where_usage') else 0
                if where_count:
                    features.append(f"{where_count} 条件")
                if parsed.union_branches:
                    features.append(f"{len(parsed.union_branches)} UNION分支")
                if features:
                    lines.append(f"- SQL特征: {', '.join(features)}")

            # SQL 注释（如果有，帮助 AI 理解业务含义）
            if rule.query_sql and parsed and not parsed.parse_error:
                import re as _re
                comments = _re.findall(r'/\*\s*(.*?)\s*\*/', rule.query_sql)
                if comments:
                    lines.append(f"- SQL注释: {'; '.join(comments[:3])}")

            # 逻辑块（只顶层概要，不递归 children）
            df_step = next((s for s in data_flow.get("steps", []) if s.get("step_id") == sid), None)
            if df_step and df_step.get("data_blocks"):
                blocks = df_step["data_blocks"]
                # 顶层概要：每个块一行，不递归
                block_summary_parts = []
                for idx, blk in enumerate(blocks[:6]):  # 最多展示6个顶层块
                    role = blk.get("role", "")
                    table = blk.get("table", "")
                    jt = blk.get("join_type", "")
                    ops = blk.get("ops", [])
                    part = f"{table}"
                    if jt and jt != "FROM":
                        part += f"({jt})"
                    if ops:
                        part += f"[{','.join(ops[:3])}]"
                    block_summary_parts.append(part)
                lines.append(f"- 逻辑块({len(blocks)}): {' + '.join(block_summary_parts)}")

        # 兜底描述（脚本已生成）
        if auto_desc.get("purpose"):
            lines.append(f"- 脚本兜底: {auto_desc['purpose']}")
        lines.append("")

    # ── 质量问题 ──
    issues = quality.get("issues", [])
    if issues:
        lines.append("## 质量问题")
        for iss in issues[:10]:
            lines.append(f"- [{iss.get('severity','')}] {iss.get('title','')}")
        lines.append("")

    # ── AI 输出模板 ──
    lines.append("---")
    lines.append("")
    lines.append("## AI 输出模板（按此格式输出，保存为 knowledge_ai.md）")
    lines.append("")
    lines.append("```markdown")
    lines.append("# 整体描述")
    lines.append("（2-3句话描述这个ETL是干什么的）")
    lines.append("")
    for rule in rules:
        step = next((s for s in topology["steps"] if s["rule_code"] == rule.rule_code), None)
        sid = step["step_id"] if step else ""
        lines.append(f"## {sid}")
        lines.append(f"（基于以下逻辑块推理这步的业务目的，1-2句话）")
        lines.append(f"### 块目的")
        lines.append(f"（为每个块补充业务目的，格式: - 块N (角色 表): 目的）")
        # 列出该步骤的逻辑块，让 AI 为每个块补充目的
        df_step_tpl = next((s for s in data_flow.get("steps", []) if s.get("step_id") == sid), None)
        if df_step_tpl and df_step_tpl.get("data_blocks"):
            for idx, blk in enumerate(df_step_tpl["data_blocks"]):
                _append_block_template(lines, blk, idx, indent=0)
        lines.append("")
    lines.append("## 关键字段")
    lines.append("- 字段名: 业务含义")
    lines.append("```")
    lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="dws-pipeline-analyzer — 制品包深度分析器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--input", required=True, help="execution_tasks.xlsx 文件路径")
    parser.add_argument("--output", required=True, help="输出基础目录（脚本会在此目录下按规则组英文名称建子目录）")
    parser.add_argument("--dialect", default="", help="SQL 方言 (oracle/dws/auto)，默认自动检测")
    parser.add_argument("--ddl-dir", default="", help="DDL 文件目录（可选，用于补充字段类型）")
    args = parser.parse_args()

    input_path = Path(args.input)
    base_output_dir = Path(args.output)

    if not input_path.exists():
        print(f"错误: 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    print(f"=== dws-pipeline-analyzer ===")
    print(f"输入: {input_path}")
    print(f"输出基础目录: {base_output_dir}")
    print()

    # ── Step 1: 读取输入 ──
    # 输入分流：.xlsx 文件 → read_excel；目录 → read_yml（代码仓 yml 场景）
    is_yml_mode = input_path.is_dir()
    if is_yml_mode:
        print("Step 1: 读取代码仓 yml...")
        raw = read_yml(str(input_path))
    else:
        print("Step 1: 读取制品包 Excel...")
        raw = read_excel(str(input_path))
    rules = raw["rules"]

    # 确定输出目录：基础目录 / 规则组英文名称（兜底用规则组编码或 output）
    group_en = (raw.get("rule_group_en") or "").strip()
    if not group_en:
        group_en = (raw.get("rule_group_code") or "").strip() or "output"
    # 清理目录名（去掉非法字符）
    safe_group_en = re.sub(r'[<>:"/\\|?*\s]', "_", group_en)
    output_dir = base_output_dir / safe_group_en
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"输出目录: {output_dir}")
    print()

    if not rules and not is_yml_mode:
        # 详细诊断
        print("错误: 未找到有效的 RULE 行", file=sys.stderr)
        print("", file=sys.stderr)
        print("诊断信息:", file=sys.stderr)

        # 检查 RULE sheet 是否存在
        wb_diag = openpyxl.load_workbook(str(input_path), read_only=False, data_only=True)
        if "RULE" not in wb_diag.sheetnames:
            print(f"  - Excel 中没有 'RULE' sheet，实际 sheet: {wb_diag.sheetnames}", file=sys.stderr)
        else:
            ws_diag = wb_diag["RULE"]
            diag_headers = []
            for cell in next(ws_diag.iter_rows(min_row=1, max_row=1, values_only=False)):
                if cell.value:
                    diag_headers.append(str(cell.value).strip())

            data_rows = 0
            skipped_no_sql = 0
            skipped_rule_type = []
            for row in ws_diag.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                data_rows += 1

            # 找规则类型列
            rt_col_idx = None
            for h_idx, h in enumerate(diag_headers):
                if "规则类型" in h:
                    rt_col_idx = h_idx
                    break

            # 找查询语句列
            sql_col_idx = None
            for h_idx, h in enumerate(diag_headers):
                if "查询语句" in h:
                    sql_col_idx = h_idx
                    break

            print(f"  - RULE sheet 存在，数据行数: {data_rows}", file=sys.stderr)
            print(f"  - 表头列数: {len(diag_headers)}", file=sys.stderr)

            if rt_col_idx is None:
                print(f"  - [FAIL] 找不到 '规则类型' 列！表头: {diag_headers[:10]}", file=sys.stderr)
            else:
                print(f"  - 规则类型列 idx={rt_col_idx} ('{diag_headers[rt_col_idx]}')", file=sys.stderr)
                # 检查规则类型值
                rt_values = set()
                for row in ws_diag.iter_rows(min_row=2, values_only=True):
                    if row and rt_col_idx < len(row):
                        rt_val = row[rt_col_idx]
                        rt_values.add(str(rt_val) if rt_val is not None else "None")
                print(f"  - 规则类型值: {rt_values}", file=sys.stderr)

            if sql_col_idx is None:
                print(f"  - [FAIL] 找不到 '查询语句' 列！表头含'查询': {[h for h in diag_headers if '查询' in h or 'sql' in h.lower()]}", file=sys.stderr)
            else:
                print(f"  - 查询语句列 idx={sql_col_idx} ('{diag_headers[sql_col_idx]}')", file=sys.stderr)
                # 检查 SQL 是否为空
                empty_sql = 0
                for row in ws_diag.iter_rows(min_row=2, values_only=True):
                    if row and sql_col_idx < len(row):
                        sql_val = row[sql_col_idx]
                        if not sql_val or not str(sql_val).strip():
                            empty_sql += 1
                if empty_sql > 0:
                    print(f"  - [WARN] {empty_sql} 行的 SQL 为空", file=sys.stderr)

        wb_diag.close()
        sys.exit(1)

    print(f"  RULE 行: {len(rules)}")
    print(f"  TargetFields: {sum(len(v) for v in raw['target_fields'].values())} 行")
    print(f"  GroupVariables: {sum(len(v) for v in raw['group_variables'].values())} 行")
    print(f"  变量: {raw['variables']}")
    print()

    # ── Step 2: 方言检测 ──
    dialect = args.dialect
    if not dialect or dialect == "auto":
        sql_texts = [r.query_sql for r in rules if r.query_sql]
        dialect = detect_dialect(sql_texts)
    print(f"Step 2: 方言 = {dialect}")
    print()

    # ── Step 2b: I 视图发现（yml 场景，资产是 I 视图不是 F 表）──
    # F 表是加工底表，I 视图是对外资产。发现 I 视图后追加为链路最后一步（F→I），
    # 使完整链路为 ODS→F表→I视图。找不到 I 视图→以 F 表为终点（容错，不阻塞）。
    i_view_info = None  # 记录 I 视图信息，供后续写入 knowledge.meta.asset_info
    if is_yml_mode and rules:
        # 取最终目标表（max exec_sequence 的 F 表）
        # 注意：交换分区（rule_type=9）的 target_table 是临时表，
        # exchange_source_table 才是真正的目标表（F 表），不能拿临时表去找 I 视图
        from engine import _is_intermediate_table, RawRule
        f_schema = ""
        f_table = ""
        f_rule = None
        for rule in reversed(rules):
            if rule.rule_type == 9 and rule.exchange_source_table:
                # 交换分区：真正目标表是 exchange_source_table
                f_schema = rule.target_schema
                f_table = rule.exchange_source_table
                f_rule = rule
                break
            if not _is_intermediate_table(rule.target_table):
                f_schema = rule.target_schema
                f_table = rule.target_table
                f_rule = rule
                break
        if f_rule and f_table:
            i_view = discover_i_view(input_path, f_schema, f_table)
            if i_view:
                # 追加 I 视图作为最后一步
                max_seq = max((r.exec_sequence for r in rules), default=0)
                i_rule = RawRule(
                    rule_code=f"{f_rule.rule_code}_VIEW",
                    rule_name=f"{f_rule.rule_name}（I视图）",
                    rule_type=1,
                    exec_sequence=max_seq + 1,
                    target_schema=i_view["view_schema"],
                    target_table=i_view["view_name"],
                    delete_mode="",
                    query_sql=i_view["view_sql"],
                    rule_group_code=f_rule.rule_group_code,
                    rule_group_en=f_rule.rule_group_en,
                    is_view_step=True,  # 标记为视图步骤，下游统一读这个字段
                )
                raw["rules"].append(i_rule)
                rules = raw["rules"]
                # 记录 I 视图信息，供 knowledge.meta.asset_info
                # 注意：view_step 不在这里算（exec_sequence≠step_id），等 topology 生成后取
                i_view_info = {
                    "is_view": True,
                    "view_table": i_view["view_name"],
                    "view_schema": i_view["view_schema"],
                    "base_table": f_table,
                    "base_schema": f_schema,
                    "view_rule_code": f"{f_rule.rule_code}_VIEW",  # 用于从 topology 查 step_id
                }
                print(f"Step 2b: 发现 I 视图 {i_view['view_schema']}.{i_view['view_name']}")
                print(f"  链路扩展: F表 {f_table} → I视图 {i_view['view_name']}")
                print()
            else:
                print(f"Step 2b: 未找到 I 视图（以 F 表 {f_rule.target_table} 为终点）")
                print()

    # ── DDL 发现 ──
    # yml 场景：从代码仓根定位 DDL/{DWS_EDW|DWS_RT_EDW}/{schema}/table/{target_table}.sql
    # xlsx 场景：不自动发现（xlsx 是临时导出，DDL 位置不统一），需 --ddl-dir 显式指定
    ddl_dir = args.ddl_dir
    if not ddl_dir and is_yml_mode:
        # yml 场景：从规则组目录向上找代码仓根，再定位 DDL
        ddl_dir = _auto_discover_ddl_from_repo(input_path, rules)

    # ── Step 3~7: 核心解析（与批量路径共用 analyze_pipeline，避免两套逻辑漂移）──
    print("Step 3-7: 解析 + 组装 knowledge...")
    knowledge, parsed_map = analyze_pipeline(
        rules, raw["target_fields"], raw["group_variables"], dialect,
        ddl_dir=ddl_dir, source_file=input_path.name,
        rule_group_code=raw["rule_group_code"],
    )
    # 从 knowledge 取回 AI summary 需要的中间结构
    topology = knowledge["topology"]
    data_flow = knowledge["data_flow"]
    field_mappings = knowledge["field_mappings"]
    quality = knowledge["quality"]
    target_name = knowledge["meta"]["target_table"]
    # 写入资产信息（I 视图标注，供 view_generator 渲染特殊样式）
    # excel 模式无 I 视图 → i_view_info 为 None → asset_info 不写入
    if i_view_info:
        # 从 topology 取 I 视图步骤的正确 step_id（不自己算，避免 exec_sequence≠step_id 的错位）
        view_rule_code = i_view_info.pop("view_rule_code", "")
        view_step = ""
        for s in topology["steps"]:
            if s.get("rule_code") == view_rule_code:
                view_step = s["step_id"]
                break
        i_view_info["view_step"] = view_step
        knowledge["meta"]["asset_info"] = i_view_info
    stats = field_mappings["statistics"]
    print(f"  步骤数: {len(rules)}, 字段数: {stats['total_in_sql']}, "
          f"问题数: {len(quality['issues'])}")
    print()

    # 性能诊断：读各阶段耗时（在写 JSON 前读，因为 _timings 不写进 JSON）
    timings = knowledge.get("meta", {}).get("_timings", {})

    # 写入文件（_timings 是性能诊断用，不写进 JSON）
    knowledge["meta"].pop("_timings", None)
    output_file = output_dir / "knowledge_draft.json"
    output_file.write_text(
        json.dumps(knowledge, ensure_ascii=False, indent=2),
        encoding="utf-8",
        newline="\n",
    )

    # ── 生成 AI 增强用摘要 ──
    summary_file = output_dir / "knowledge_summary.md"
    summary_text = _generate_ai_summary(knowledge, rules, parsed_map, topology, field_mappings, quality, data_flow)
    summary_file.write_text(summary_text, encoding="utf-8", newline="\n")

    print(f"\n=== 完成 ===")
    print(f"输出: {output_file}")
    print(f"摘要: {summary_file}")
    print(f"目标表: {target_name}")
    print(f"步骤数: {len(rules)}")
    print(f"字段数: {stats['total_in_sql']}")
    print(f"问题数: {len(quality['issues'])}")

    # 性能诊断：输出慢阶段（>0.5s 的才显示，避免正常情况刷屏）
    slow = sorted(timings.items(), key=lambda x: -x[1])
    slow_filtered = [(n, t) for n, t in slow if t > 0.5]
    total_t = sum(timings.values()) if timings else 0
    if slow_filtered:
        print(f"\n分析耗时: {total_t:.1f}s（慢阶段:）")
        for name, t in slow_filtered:
            print(f"  {name}: {t:.1f}s")

    print(f"\n下一步: AI 读 knowledge_summary.md，输出自然语言补充，保存为 knowledge_ai.md")
    print(f"        然后: python run.py view_generator --input knowledge_draft.json --ai-input knowledge_ai.md ...")


# ═══════════════════════════════════════════════════════════════
# 多规则组链路分析（/analyze-chain）
# ═══════════════════════════════════════════════════════════════


def build_target_index(repo_root, scope_dir=None):
    """扫描代码仓，建 {target_table → [规则组目录]} 索引。

    Args:
        repo_root: 代码仓根目录（含 BFT/）
        scope_dir: 搜索范围（子项目目录），None 则扫整个 BFT/BftWideTable

    Returns:
        {norm_target_table: [{dir, schema, rule_group_en, rule_group_code}]}
    """
    import yaml as _yaml
    from engine import _norm_table

    repo = Path(repo_root)
    bft = repo / "BFT" / "BftWideTable"
    if scope_dir:
        # 在指定子项目目录下扫
        search_root = Path(scope_dir)
    else:
        search_root = bft

    if not search_root.is_dir():
        return {}

    index = {}
    for yml_file in search_root.rglob("*.yml"):
        group_dir = yml_file.parent
        try:
            data = _yaml.safe_load(yml_file.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not data or not isinstance(data, dict):
            continue

        target_table = _yml_get(data, "target_table", _YML_RULE_KEY_ALIASES)
        target_schema = _yml_get(data, "target_schema", _YML_RULE_KEY_ALIASES)
        rule_group_en = _yml_get(data, "rule_group_en", _YML_RULE_KEY_ALIASES).strip()
        rule_group_code = _yml_get(data, "rule_group_code", _YML_RULE_KEY_ALIASES)

        if not target_table:
            continue

        key = _norm_table(target_table)
        index.setdefault(key, []).append({
            "dir": str(group_dir),
            "schema": target_schema,
            "rule_group_en": rule_group_en,
            "rule_group_code": rule_group_code,
            "target_table": target_table,
        })

    return index


def trace_upstream_rule_groups(
    final_group_dir, repo_root, max_depth=6,
):
    """从最终 F 表规则组出发，递归追溯所有上游 mid/tmp 规则组。

    算法：
      1. 分析最终规则组 → 取穿透后的源表
      2. 筛出"被代码仓里某规则组写的"表（在索引里能找到的）
      3. 对每个上游规则组递归 Step 1-2
      4. 直到源表都是 ods（索引里找不到写入者）

    Args:
        final_group_dir: 最终 F 表规则组目录
        repo_root: 代码仓根目录

    Returns:
        {
            "groups": [{dir, rule_group_en, target_table, source_tables, depth}],
            "not_found": [表名列表],  # 源表找不到写入者（可能是 ods 源表）
            "cycle_detected": bool,
        }
    """
    from engine import analyze_pipeline, _norm_table
    from analyzer import read_yml, detect_dialect

    # 确定搜索范围：先子项目，后项目级
    final_path = Path(final_group_dir).resolve()
    # 子项目 = 规则组目录的父目录（如 SUB_TRADE）
    sub_project_dir = final_path.parent
    # 项目 = 子项目的父目录（如 P_TRADE）
    project_dir = sub_project_dir.parent

    # 先在子项目下建索引
    sub_index = build_target_index(repo_root, sub_project_dir)

    groups = []
    visited_dirs = set()
    not_found_tables = set()

    def _trace(group_dir, depth):
        group_dir_resolved = str(Path(group_dir).resolve())
        if group_dir_resolved in visited_dirs:
            return  # 环检测
        if depth > max_depth:
            return
        visited_dirs.add(group_dir_resolved)

        # 读规则组
        raw = read_yml(group_dir)
        if not raw.get("rules"):
            return

        rules = raw["rules"]
        rule_group_en = raw.get("rule_group_en", "")
        dialect = detect_dialect(rules)

        # 分析 → 取穿透后的源表
        target_fields = raw.get("target_fields", {})
        group_variables = raw.get("group_variables", {})
        _, parsed_map = analyze_pipeline(rules, target_fields, group_variables, dialect)

        # 收集所有步骤的穿透源表
        source_tables = set()
        for rule in rules:
            parsed = parsed_map.get(rule.rule_code)
            if not parsed:
                continue
            for j in parsed.source_tables:
                if not j.source_table or j.source_table.startswith("(subquery:"):
                    continue
                source_tables.add(j.source_table)
            # CTE 内部表也要收
            for cte in parsed.ctes:
                for t in cte.source_tables:
                    tname = t.get("name", "")
                    if tname:
                        source_tables.add(tname)

        # 目标表（这个规则组写的）
        target_table = ""
        if rules:
            # 取最大 exec_sequence 的规则的目标表（最终产出）
            max_seq_rule = max(rules, key=lambda r: r.exec_sequence)
            target_table = max_seq_rule.target_table

        groups.append({
            "dir": str(group_dir),
            "rule_group_en": rule_group_en,
            "target_table": target_table,
            "source_tables": list(source_tables),
            "depth": depth,
        })

        # 对每个源表，找写入者
        for src_table in source_tables:
            # 归一化：去 schema 前缀（dws.dwb_trade_mid_f → dwb_trade_mid_f）
            # 索引 key 是 _norm_table(target_table)，yml 里 target_table 不带 schema
            table_part = src_table.split(".")[-1] if "." in src_table else src_table
            key = _norm_table(table_part)
            writers = sub_index.get(key, [])
            if not writers:
                # 子项目找不到，试项目级
                if not hasattr(_trace, "_project_index"):
                    _trace._project_index = build_target_index(repo_root, project_dir)
                writers = _trace._project_index.get(key, [])

            if not writers:
                not_found_tables.add(src_table)
                continue

            for writer in writers:
                _trace(writer["dir"], depth + 1)

    _trace(final_group_dir, 0)

    return {
        "groups": groups,
        "not_found": sorted(not_found_tables),
        "cycle_detected": len(visited_dirs) < len(groups),
    }


def merge_rule_groups(groups_info, repo_root, ddl_dir=""):
    """合并多个规则组的 rules，重新编号 exec_sequence，返回合并后的 rules 列表。

    拓扑排序：被依赖的（depth 大的，上游的）排前面，依赖别人的（depth 小的，下游的）排后面。

    Args:
        groups_info: trace_upstream_rule_groups 的返回值

    Returns:
        merged_rules: 合并后的 RawRule 列表（exec_sequence 已按依赖拓扑排序重编号）
    """
    from engine import RawRule, _norm_table

    groups = groups_info["groups"]
    if not groups:
        return []

    # 读所有规则组的 rules（缓存，避免重复 read_yml）
    group_rules_cache = {}
    for g in groups:
        raw = read_yml(g["dir"])
        group_rules_cache[g["dir"]] = raw.get("rules", [])

    # 算每个规则组的 target_table（归一化）
    group_target = {}  # {group_dir: norm_target_table}
    for g in groups:
        rules = group_rules_cache.get(g["dir"], [])
        if rules:
            max_seq_rule = max(rules, key=lambda r: r.exec_sequence)
            group_target[g["dir"]] = _norm_table(max_seq_rule.target_table)

    # 算组间依赖关系：A 写的表被 B 的源表引用 → B 依赖 A
    # group_deps[dir] = set(被依赖的 dir)
    group_deps = {g["dir"]: set() for g in groups}
    for g in groups:
        for src_table in g.get("source_tables", []):
            table_part = src_table.split(".")[-1] if "." in src_table else src_table
            src_key = _norm_table(table_part)
            # 找哪个规则组写了这张表
            for other_dir, other_target in group_target.items():
                if other_target == src_key and other_dir != g["dir"]:
                    group_deps[g["dir"]].add(other_dir)

    # 按依赖拓扑排序（被依赖的先处理）
    # 用深度优先拓扑排序
    sorted_dirs = []
    visited = set()
    temp_mark = set()

    def _topo_sort(dir_path):
        if dir_path in visited:
            return
        if dir_path in temp_mark:
            return  # 环检测
        temp_mark.add(dir_path)
        for dep_dir in group_deps.get(dir_path, set()):
            _topo_sort(dep_dir)
        temp_mark.discard(dir_path)
        visited.add(dir_path)
        sorted_dirs.append(dir_path)

    for g in groups:
        _topo_sort(g["dir"])

    # 计算每个组的偏移量：= max(依赖组的全局最大seq) + 1
    # 互相不依赖的组 seq 重叠（都从能开始的最早seq开始）
    group_offset = {}  # {dir: offset}
    group_global_max = {}  # {dir: 该组加偏移后的全局最大seq}

    for dir_path in sorted_dirs:
        rules = group_rules_cache.get(dir_path, [])
        if not rules:
            group_offset[dir_path] = 0
            group_global_max[dir_path] = 0
            continue
        max_seq_in_group = max(r.exec_sequence for r in rules)

        # 偏移 = 依赖的所有组中最大的全局seq + 1
        deps = group_deps.get(dir_path, set())
        if deps:
            offset = max(group_global_max.get(d, 0) for d in deps) + 1
        else:
            offset = 0  # 没有依赖，从0开始（可能跟其他无依赖组重叠）

        group_offset[dir_path] = offset
        group_global_max[dir_path] = offset + max_seq_in_group

    # 合并 rules，应用偏移量
    merged_rules = []
    for g in groups:
        dir_path = g["dir"]
        offset = group_offset.get(dir_path, 0)
        for rule in group_rules_cache.get(dir_path, []):
            merged_rules.append(RawRule(
                rule_code=rule.rule_code,
                rule_name=rule.rule_name,
                rule_type=rule.rule_type,
                exec_sequence=rule.exec_sequence + offset,
                target_schema=rule.target_schema,
                target_table=rule.target_table,
                delete_mode=rule.delete_mode,
                query_sql=rule.query_sql,
                exchange_source_table=getattr(rule, "exchange_source_table", ""),
                rule_group_code=rule.rule_group_code,
                rule_group_en=rule.rule_group_en,
            ))

    return merged_rules


def main_chain():
    """多规则组链路分析 CLI 入口。"""
    parser = argparse.ArgumentParser(
        description="多规则组链路分析（/analyze-chain）",
    )
    parser.add_argument("--input", required=True,
                        help="最终 F 表规则组目录或表名")
    parser.add_argument("--output", required=True,
                        help="输出基础目录")
    parser.add_argument("--repo-root", default="",
                        help="代码仓根目录（含 BFT/），不指定则自动从 input 向上找")
    parser.add_argument("--ddl-dir", default="",
                        help="DDL 文件目录（可选）")
    args = parser.parse_args()

    input_path = Path(args.input)

    # 定位代码仓根
    if args.repo_root:
        repo_root = Path(args.repo_root)
    else:
        repo_root = _find_repo_root(input_path if input_path.is_dir() else input_path.parent)
        if not repo_root:
            print("[ERROR] 无法定位代码仓根（需含 BFT/ + DDL/），请用 --repo-root 指定", file=sys.stderr)
            sys.exit(1)

    # 定位最终 F 表规则组目录
    if input_path.is_dir():
        final_group_dir = input_path
    else:
        # 按表名定位（复用现有 locate_asset_dirs 逻辑）
        from impact_analyzer import locate_asset_dirs
        located = locate_asset_dirs([args.input], str(repo_root))
        info = located.get(args.input, {})
        if not info.get("found"):
            print(f"[ERROR] 未定位到规则组目录: {args.input}", file=sys.stderr)
            sys.exit(1)
        final_group_dir = Path(info["dir"])

    print(f"=== 多规则组链路分析 ===")
    print(f"最终规则组: {final_group_dir.name}")
    print(f"代码仓根: {repo_root}")

    # 追溯上游
    print(f"\n[Step 1] 追溯上游规则组...")
    result = trace_upstream_rule_groups(final_group_dir, repo_root)
    print(f"  找到 {len(result['groups'])} 个规则组:")
    for g in sorted(result["groups"], key=lambda x: x["depth"]):
        indent = "    " * g["depth"]
        print(f"  {indent}{g['rule_group_en']} → 写 {g['target_table']}")
    if result["not_found"]:
        print(f"  源表（不追溯）: {', '.join(result['not_found'])}")

    # 合并
    print(f"\n[Step 2] 合并规则组...")
    merged_rules = merge_rule_groups(result, repo_root)
    print(f"  合并后 {len(merged_rules)} 条规则（exec_sequence 已重编号）")

    # DDL 发现
    ddl_dir = args.ddl_dir
    if not ddl_dir and repo_root:
        try:
            ddl_dir = _auto_discover_ddl_from_repo(merged_rules, final_group_dir) or ""
        except Exception:
            ddl_dir = ""

    # 分析
    print(f"\n[Step 3] 分析完整链路...")
    dialect = detect_dialect(merged_rules)
    knowledge, parsed_map = analyze_pipeline(
        merged_rules, {}, {}, dialect,
        ddl_dir=ddl_dir, source_file="",
        rule_group_code="CHAIN",
    )

    # 标记这是多规则组链路分析（供报告区分）
    knowledge["meta"]["is_multi_group"] = True
    knowledge["meta"]["chain_groups"] = [
        {"name": g["rule_group_en"], "target_table": g["target_table"], "depth": g["depth"]}
        for g in sorted(result["groups"], key=lambda x: x["depth"])
    ]

    # 输出
    safe_name = "chain_" + final_group_dir.name
    output_dir = Path(args.output) / safe_name
    output_dir.mkdir(parents=True, exist_ok=True)

    knowledge["meta"].pop("_timings", None)
    output_file = output_dir / "knowledge_draft.json"
    output_file.write_text(
        json.dumps(knowledge, ensure_ascii=False, indent=2),
        encoding="utf-8", newline="\n",
    )

    # 生成 AI 摘要
    summary_text = _generate_ai_summary(
        knowledge, merged_rules, parsed_map,
        knowledge["topology"], knowledge["field_mappings"],
        knowledge["quality"], knowledge["data_flow"],
    )
    (output_dir / "knowledge_summary.md").write_text(summary_text, encoding="utf-8", newline="\n")

    # 生成三件套（mapping + asset_report + tech_design）
    print(f"\n[Step 4] 生成视图...")
    try:
        from view_generator import generate_mapping, generate_asset_report, generate_tech_design
        generate_mapping(knowledge, output_dir)
        generate_asset_report(knowledge, output_dir)
        generate_tech_design(knowledge, output_dir)
    except Exception as e:
        print(f"  [WARN] 视图生成失败: {e}", file=sys.stderr)

    stats = knowledge["field_mappings"]["statistics"]
    target_table = knowledge["meta"]["target_table"]
    print(f"\n=== 完成 ===")
    print(f"输出目录: {output_dir}")
    print(f"规则组数: {len(result['groups'])}")
    print(f"步骤数: {len(merged_rules)}")
    print(f"字段数: {stats['total_in_sql']}")
    print(f"目标表: {target_table}")
    print(f"\n已生成:")
    print(f"  - knowledge_draft.json")
    print(f"  - mapping.xlsx")
    print(f"  - asset_report.html")
    print(f"  - tech_design.md")


# ═══════════════════════════════════════════════════════════════
# 兼容层：re-export engine 的符号
# 引擎代码已物理搬入 engine.py。以下 re-export 保证现有代码
# `from analyzer import xxx` 继续可用（过渡期，新代码请直接 from engine import）。
# ═══════════════════════════════════════════════════════════════
from engine import (  # noqa: E402,F401
    ParsedSQL, ParsedJoin, RawRule, RawTargetField, RawGroupVariable,
    ParsedColumn, ParsedCTE, TableRef, ColumnRef, QueryUnit,
    analyze_pipeline,
    detect_dialect, parse_single_sql, classify_transform,
    build_topology, build_data_flow, build_field_mappings, analyze_quality,
    build_join_key_lineage, enrich_join_key_lineage, enrich_field_physical_sources,
    build_data_blocks, build_structured_step_summary, generate_step_description,
    detect_patterns, build_source, parse_ddl_for_metadata,
    _append_block_summary, _append_block_template,
    _is_intermediate_table, _strip_dws_clauses, _replace_placeholders,
    _normalize_table_name, _norm_table, _table_match, _infer_layer, _clean_name,
    SELECT_RULE_TYPES, RULE_TYPE_MAP, DELETE_MODE_MAP,
)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--chain":
        sys.argv.pop(1)
        main_chain()
    else:
        main()
