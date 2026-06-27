"""字段使用情况批量搜索工具。

从多个规则组的 Excel 中，按关键字搜索字段的使用情况，
输出一张 Excel（按目标表分组）。

设计原则: 复用 analyzer 的完整解析能力（含 enrich 追溯），
field_search 只负责"搜索关键字 + 组织输出"，不复制解析逻辑。

使用:
    python run.py field_search --input execution_tasks.xlsx --keyword amount --output field_usage.xlsx
    python run.py field_search --input execution_tasks.xlsx --keyword "amount,user_id" --output field_usage.xlsx
"""

import sys
from pathlib import Path
from dataclasses import dataclass


@dataclass
class FieldUsage:
    """单个字段的使用记录。"""
    target_table: str = ""       # 目标表（schema.table）
    field_name: str = ""         # 字段名
    role: str = ""               # 字段角色：写入目标表/临时过程使用/辅助字段
    situation: str = ""          # 字段情况：直取/加工/关联带出/关联键/过滤条件
    source: str = ""             # 最初来源（物理源表.字段，辅助字段填使用步骤）
    detail: str = ""             # 详情（加工表达式/关联条件/过滤条件）


def read_excel_grouped(excel_path: str) -> list:
    """读取 Excel，按规则组编码分组返回。

    Returns: [{rule_group_code, rule_group_en, rules: [RawRule], ...}, ...]
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from analyzer import read_excel

    raw = read_excel(excel_path)
    all_rules = raw["rules"]

    # 按 rule_group_code 分组；无 code 的行不混入同一组，按 rule_code 单独成组
    groups_map = {}
    unknown_idx = 0
    for rule in all_rules:
        code = rule.rule_group_code
        if not code:
            # 无规则组编码：按 rule_code 单独成组（避免无关规则混入）
            code = f"_SOLO_{rule.rule_code or f'ROW{unknown_idx}'}"
            unknown_idx += 1
        if code not in groups_map:
            groups_map[code] = {
                "rule_group_code": code,
                "rule_group_en": code,
                "rules": [],
            }
        groups_map[code]["rules"].append(rule)

    return list(groups_map.values())


def search_field_usage(excel_path: str, keywords: list) -> list:
    """主入口：搜索字段使用情况。

    轻量解析策略：只 parse_single_sql + build_field_mappings（跳过 topology/
    data_flow/enrich），对匹配字段按需追溯。避免大数据量时 enrich 全量计算过慢。

    Args:
        excel_path: Excel 文件路径
        keywords: 关键字列表（如 ["amount", "user_id"]）

    Returns: [FieldUsage, ...] 按目标表分组排序
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from analyzer import detect_dialect, parse_single_sql, build_field_mappings

    groups = read_excel_grouped(excel_path)
    all_usages = []
    keywords_lower = [k.lower() for k in keywords]

    for group in groups:
        rules = group["rules"]
        if not rules:
            continue

        # 预扫描：检查该规则组的 SQL 文本是否含任一关键字，不含则跳过
        all_sql = " ".join(r.query_sql or "" for r in rules).lower()
        if not any(k in all_sql for k in keywords_lower):
            continue

        # 轻量解析：只 parse + field_mappings（跳过 topology/data_flow/enrich）
        sqls = [r.query_sql for r in rules if r.query_sql]
        dialect = detect_dialect(sqls)
        parsed_map = {r.rule_code: parse_single_sql(r.query_sql, dialect) for r in rules}
        field_mappings = build_field_mappings(rules, parsed_map, {})

        # 检测 SELECT * 和解析失败（提示用户）
        for r in rules:
            p = parsed_map.get(r.rule_code)
            if not p:
                continue
            if p.has_star:
                print(f"  [WARN] 规则 {r.rule_code} 使用了 SELECT *（{r.target_table}），无法追踪字段血缘", file=sys.stderr)
            if p.parse_error:
                print(f"  [WARN] 规则 {r.rule_code} SQL 解析失败: {p.parse_error[:50]}", file=sys.stderr)

        # 搜索字段（按需追溯，只对匹配字段算）
        _search_group(rules, parsed_map, field_mappings, keywords_lower, all_usages)

    # 按目标表分组排序
    all_usages.sort(key=lambda u: (u.target_table, u.role, u.field_name))
    return all_usages


def _search_group(rules, parsed_map, field_mappings, keywords_lower, all_usages):
    """搜索单个规则组的字段使用情况。

    每个字段穿透到最终目标表，合并所有步骤的角色（斜杠分隔），一行输出。
    """
    fields_list = field_mappings.get("fields", [])

    # 最终目标表（非中间表的最后一个步骤的 target）
    final_target = ""
    for rule in reversed(rules):
        if not _is_intermediate(rule.target_table):
            final_target = f"{rule.target_schema}.{rule.target_table}" if rule.target_schema else rule.target_table
            break

    # 收集所有匹配字段的使用信息，按字段名（小写）合并
    field_usage_map = {}  # {field_lower: {roles: set, situations: set, source, details: []}}

    for rule in rules:
        parsed = parsed_map.get(rule.rule_code)
        if not parsed or parsed.parse_error:
            continue

        rule_idx = rules.index(rule)
        step_id = f"step_{rule_idx + 1}"
        rule_code = rule.rule_code or step_id  # 详情标签用规则编码（用户能对应到真实代码）
        is_final_step = not _is_intermediate(rule.target_table)

        # 别名→物理表映射（用于详情里的别名替换）
        alias_map = {}
        for j in parsed.source_tables:
            if j.alias and not j.source_table.startswith("(subquery:"):
                alias_map[j.alias] = j.source_table.split(".")[-1]  # 短名（不带 schema，避免太长）

        def _resolve_aliases(text):
            """把文本里的别名.field 替换成 物理表短名.field"""
            import re
            for alias, table in alias_map.items():
                text = re.sub(r'\b' + re.escape(alias) + r'\.', table + '.', text)
            return text

        # 1. SELECT 字段（写入/临时）
        for f in fields_list:
            if f.get("producing_step") != step_id:
                continue
            fname = f.get("target_field", "")
            matched = _match_field(fname, f, keywords_lower)
            if not matched:
                continue

            fl = fname.lower()
            if fl not in field_usage_map:
                field_usage_map[fl] = {"name": fname, "roles": [], "situations": [], "source": "", "details": []}

            entry = field_usage_map[fl]
            transform = f.get("transform_type", "direct")
            source = _trace_physical_source(f, parsed_map, rules) if is_final_step else ""

            if is_final_step:
                if "写入目标表" not in entry["roles"]:
                    entry["roles"].append("写入目标表")
                if not entry["source"]:
                    entry["source"] = source
                sit = _situation_label(transform)
                if sit not in entry["situations"]:
                    entry["situations"].append(sit)
                # 写入字段的详情：[来源/stepN] + [加工/stepN]
                if source and source != "-":
                    detail_src = f"[来源/{rule_code}] {source}"
                    if detail_src not in entry["details"]:
                        entry["details"].append(detail_src)
                raw_expr = _detail_label(f)
                if raw_expr and transform != "direct":
                    detail_expr = f"[加工/{rule_code}] {raw_expr}"
                    if detail_expr not in entry["details"]:
                        entry["details"].append(detail_expr)
            else:
                # 中间步骤的字段，只在详情里体现传递路径
                pass

        # 2. 辅助字段（关联键 + 过滤条件）
        seen_aux = set()
        for ju in parsed.join_usage:
            jf = ju.get("field", "")
            if not _match_keyword(jf, keywords_lower):
                continue
            aux_key = (jf.lower(), "关联键")
            if aux_key in seen_aux:
                continue
            seen_aux.add(aux_key)

            fl = jf.lower()
            if fl not in field_usage_map:
                field_usage_map[fl] = {"name": jf, "roles": [], "situations": [], "source": "", "details": []}
            entry = field_usage_map[fl]
            if "关联键" not in entry["roles"]:
                entry["roles"].append("关联键")
            cond = _resolve_aliases(ju.get("on_condition", ""))
            detail_join = f"[关联/{rule_code}] {cond}" if cond else ""
            if detail_join and detail_join not in entry["details"]:
                entry["details"].append(detail_join)

        for wu in parsed.where_usage:
            wf = wu.get("field", "")
            if not _match_keyword(wf, keywords_lower):
                continue
            aux_key = (wf.lower(), "过滤条件")
            if aux_key in seen_aux:
                continue
            seen_aux.add(aux_key)

            fl = wf.lower()
            if fl not in field_usage_map:
                field_usage_map[fl] = {"name": wf, "roles": [], "situations": [], "source": "", "details": []}
            entry = field_usage_map[fl]
            if "过滤条件" not in entry["roles"]:
                entry["roles"].append("过滤条件")
            cond = _resolve_aliases(wu.get("condition", ""))
            detail_where = f"[过滤/{rule_code}] {cond}" if cond else ""
            if detail_where and detail_where not in entry["details"]:
                entry["details"].append(detail_where)

    # 输出合并后的字段使用记录（目标表 = 最终目标表）
    for fl, entry in field_usage_map.items():
        if not entry["roles"]:
            continue
        all_usages.append(FieldUsage(
            target_table=final_target,
            field_name=entry["name"],
            role="/".join(entry["roles"]),
            situation="/".join(entry["situations"]) if entry["situations"] else "-",
            source=entry["source"] or "-",
            detail="；".join(entry["details"]) if entry["details"] else "-",
        ))


def _trace_physical_source(f, parsed_map, rules, depth=0, visited=None):
    """按需追溯字段的物理来源（轻量，不依赖 enrich）。

    沿 lineage 追，遇到中间表继续追，遇到 CTE 穿透到内部，物理源表停止。
    带深度限制（max 12）和循环保护。
    """
    if visited is None:
        visited = set()
    if depth > 25:
        return ""

    lineages = f.get("lineage", [])
    if not lineages:
        return ""

    first = lineages[0]
    src_alias = first.get("source_table", "")
    src_field = first.get("source_field", f.get("target_field", ""))

    # CTE 穿透信息
    cte_source_fields = first.get("cte_source_fields")
    cte_step = first.get("step", "")
    if cte_source_fields:
        step_ctes = _get_step_ctes(cte_step, parsed_map, rules)
        parts = _resolve_cte_sources(cte_source_fields, step_ctes)
        if parts:
            return "；".join(parts)

    # 解析别名 → 物理表
    src_table = _resolve_alias_to_table(src_alias, f.get("producing_step", ""), parsed_map, rules)
    if not src_table:
        return src_field

    # 物理源表 → 返回
    if not _is_intermediate(src_table):
        return f"{src_table}.{src_field}"

    # 中间表 → 继续追溯
    visit_key = (src_table.lower(), src_field.lower())
    if visit_key in visited:
        return f"{src_table}.{src_field}"
    visited.add(visit_key)

    for r2 in rules:
        if r2.target_table and r2.target_table.lower() == src_table.split(".")[-1].lower():
            from analyzer import build_field_mappings
            fm2 = build_field_mappings([r2], parsed_map, {}).get("fields", [])
            for f2 in fm2:
                if f2.get("target_field", "").lower() == src_field.lower():
                    result = _trace_physical_source(f2, parsed_map, rules, depth + 1, visited.copy())
                    if result:
                        return result
    return f"{src_table}.{src_field}"


def _resolve_alias_to_table(alias, step_id, parsed_map, rules):
    """解析别名到物理表名。"""
    if not alias:
        return ""
    step_to_rule = {f"step_{i+1}": r.rule_code for i, r in enumerate(rules)}
    rule_code = step_to_rule.get(step_id, "")
    p = parsed_map.get(rule_code)
    if not p:
        return alias
    for j in p.source_tables:
        if j.alias and j.alias.upper() == alias.upper():
            return j.source_table
    return alias


def _get_step_ctes(step_id, parsed_map, rules):
    """获取某步骤的 CTE 索引。"""
    step_to_rule = {f"step_{i+1}": r.rule_code for i, r in enumerate(rules)}
    rule_code = step_to_rule.get(step_id, "")
    p = parsed_map.get(rule_code)
    if not p:
        return {}
    result = {}
    for ct in p.ctes:
        cname = (ct.name or "").upper()
        if not cname:
            continue
        alias_to_table = {}
        for st in ct.source_tables:
            if st.get("alias") and st.get("name"):
                alias_to_table[st["alias"].upper()] = st["name"]
        fields_map = {}
        for cf in ct.fields:
            fname = (cf.get("name") or "").upper()
            if fname:
                fields_map[fname] = cf.get("source_fields", [])
        result[cname] = {"fields_map": fields_map, "alias_to_table": alias_to_table}
    return result


def _get_physical_source(f, parsed_map=None, rules=None):
    """从 field 的 physical_source（enrich 注入）取物理源表.字段。

    如果 physical_source 停在 CTE 别名（如 c.total），在该别名所在步骤
    （physical_source.step_id）的 CTE 里找内部物理来源。
    CTE 不跨步骤，只在该步骤的 CTE 定义里穿透。
    """
    ps_list = f.get("physical_source", [])
    lineages = f.get("lineage", [])

    # step_id → rule_code 映射
    step_to_rule = {}
    if rules:
        for i, r in enumerate(rules):
            step_to_rule[f"step_{i+1}"] = r.rule_code

    def _get_step_ctes(step_id):
        """获取某步骤的 CTE 索引 {cte_name(UPPER): {fields_map, alias_to_table}}"""
        rule_code = step_to_rule.get(step_id, "")
        p = parsed_map.get(rule_code) if parsed_map else None
        if not p:
            return {}
        result = {}
        for ct in p.ctes:
            cname = (ct.name or "").upper()
            if not cname:
                continue
            alias_to_table = {}
            for st in ct.source_tables:
                if st.get("alias") and st.get("name"):
                    alias_to_table[st["alias"].upper()] = st["name"]
            fields_map = {}
            for cf in ct.fields:
                fname = (cf.get("name") or "").upper()
                if fname:
                    fields_map[fname] = cf.get("source_fields", [])
            result[cname] = {"fields_map": fields_map, "alias_to_table": alias_to_table}
        return result

    # 当前 field lineage 里的 CTE 信息（同步骤的 CTE）
    cte_source_fields = None
    cte_step_id = ""
    for l in lineages:
        if l.get("cte_source_fields"):
            cte_source_fields = l["cte_source_fields"]
            cte_step_id = l.get("step", "")
            break

    if not ps_list:
        if lineages:
            src = lineages[0].get("source_table", "")
            field = lineages[0].get("source_field", "")
            if cte_source_fields:
                step_ctes = _get_step_ctes(cte_step_id)
                parts = _resolve_cte_sources(cte_source_fields, step_ctes)
                if parts:
                    return "；".join(parts)
            if src and field:
                return f"{src}.{field}"
        return ""

    parts = []
    for ps in ps_list:
        tbl = ps.get("table", "")
        fld = ps.get("field", "")
        ps_step = ps.get("step_id", "")  # CTE 别名所在步骤
        if tbl and "." not in tbl:
            # CTE 别名：在该步骤的 CTE 里找
            step_ctes = _get_step_ctes(ps_step) if ps_step else {}
            resolved = _resolve_cte_field(tbl, fld, step_ctes)
            if resolved:
                parts.extend(resolved)
                continue
            # 同步骤 CTE 穿透（lineage 有 cte_source_fields）
            if cte_source_fields:
                step_ctes2 = _get_step_ctes(cte_step_id)
                resolved2 = _resolve_cte_sources(cte_source_fields, step_ctes2)
                if resolved2:
                    parts.extend(resolved2)
                    continue
        if tbl and fld:
            parts.append(f"{tbl}.{fld}")
    return "；".join(parts) if parts else ""


def _resolve_cte_field(cte_alias, field_name, step_ctes):
    """在该步骤的 CTE 里穿透别名到物理表。"""
    for cname, ct_info in step_ctes.items():
        fields_map = ct_info["fields_map"]
        alias_to_table = ct_info["alias_to_table"]
        fld_upper = (field_name or "").upper()
        if fld_upper in fields_map:
            src_fields = fields_map[fld_upper]
            return [f"{alias_to_table.get((sf.get('alias') or '').upper(), sf.get('alias',''))}.{sf.get('field','')}"
                    for sf in src_fields]
    return None


def _resolve_cte_sources(cte_source_fields, step_ctes):
    """从 cte_source_fields 在该步骤 CTE 里解析物理来源。"""
    parts = []
    for csf in cte_source_fields:
        csf_alias = (csf.get("alias", "") or "").upper()
        csf_field = csf.get("field", "")
        for cname, ct_info in step_ctes.items():
            alias_to_table = ct_info["alias_to_table"]
            if csf_alias in alias_to_table:
                parts.append(f"{alias_to_table[csf_alias]}.{csf_field}")
                break
        else:
            parts.append(f"{csf_alias.lower()}.{csf_field}")
    return parts


def _is_intermediate(table_name):
    """判断是否中间表（复用 analyzer 逻辑）。"""
    import re
    short = (table_name or "").strip().lower().split(".")[-1]
    return bool(re.search(r"(?:^tmp\d*$|_tmp\d*$|^temp\d*$|_temp\d*$|^tmp_|_tmp_|^temp_|_temp_)", short))


def _match_field(fname, f, keywords_lower):
    """字段名或加工表达式匹配关键字。"""
    if _match_keyword(fname, keywords_lower):
        return True
    for l in f.get("lineage", []):
        if _match_keyword(l.get("source_field", ""), keywords_lower):
            return True
        if _match_keyword(l.get("raw_sql", ""), keywords_lower):
            return True
    return False


def _match_keyword(text, keywords_lower):
    """文本是否包含任一关键字（大小写不敏感）。"""
    if not text:
        return False
    text_lower = text.lower()
    return any(k in text_lower for k in keywords_lower)


def _situation_label(transform):
    """字段情况标签。"""
    tt = transform or "direct"
    labels = {
        "direct": "直取", "value": "赋值", "aggregate": "加工(聚合)",
        "expression": "加工", "case_when": "加工(条件)", "fallback": "加工(兜底)",
        "window": "加工(窗口)", "pivot": "加工(行转列)",
    }
    return labels.get(tt, tt)


def _detail_label(f):
    """详情标签（加工表达式 / 关联信息）。"""
    parts = []
    for l in f.get("lineage", []):
        src_field = l.get("source_field", "")
        transform = l.get("transform", "direct")
        raw = l.get("raw_sql", "")
        if transform != "direct" and raw:
            parts.append(raw)
        elif src_field:
            parts.append(src_field)
    return "；".join(parts) if parts else ""


def output_excel(usages: list, output_path: str) -> bool:
    """输出字段使用情况到 Excel（一个大 sheet）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[ERROR] 缺少 openpyxl", file=sys.stderr)
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "字段使用情况"

    headers = ["目标表", "字段名", "字段角色", "字段情况", "最初来源", "详情"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for u in usages:
        ws.append([u.target_table, u.field_name, u.role, u.situation, u.source, u.detail])

    col_widths = [25, 20, 14, 16, 30, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(output_path)
    return True


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="字段使用情况批量搜索（支持多个规则组 + 多关键字）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python run.py field_search --input execution_tasks.xlsx --keyword amount --output field_usage.xlsx
  python run.py field_search --input execution_tasks.xlsx --keyword "amount,user_id" --output field_usage.xlsx
""",
    )
    parser.add_argument("--input", required=True, help="execution_tasks.xlsx 文件路径（含多个规则组）")
    parser.add_argument("--keyword", required=True, help="搜索关键字，多个用逗号分隔（如 amount,user_id）")
    parser.add_argument("--output", required=True, help="输出 Excel 路径")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    keywords = [k.strip() for k in args.keyword.split(",") if k.strip()]
    if not keywords:
        print("错误: 至少提供一个关键字", file=sys.stderr)
        sys.exit(1)

    print(f"=== 字段使用情况搜索 ===")
    print(f"输入: {input_path}")
    print(f"关键字: {keywords}")
    print()

    print("Step 1: 解析 Excel（多规则组）...")
    usages = search_field_usage(str(input_path), keywords)
    print(f"  匹配到 {len(usages)} 条字段使用记录")

    print(f"\nStep 2: 输出 Excel...")
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    ok = output_excel(usages, str(output_path))
    if ok:
        print(f"  [OK] {output_path}")
        print(f"\n=== 完成 ===")
    else:
        print(f"\n=== 失败 ===", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
