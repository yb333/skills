#!/usr/bin/env python3
"""合成 execution_tasks.xlsx 辅助工具。

用 Python 字典定义生成符合制品包格式的 Excel，避免手写。

Usage:
    from _build_xlsx import build_xlsx
    build_xlsx("case_01_minimal/execution_tasks.xlsx", rules=[...], target_fields=[...])

或者命令行（单个 case）:
    python _build_xlsx.py --case case_01_minimal

支持的 case 定义格式:
    rules = [{
        "rule_code": "UR001",
        "rule_type": 1,            # 1=取数规则, 12=参数变量
        "exec_sequence": 1,
        "target_schema": "dws",
        "target_table": "dwb_test_f",
        "delete_mode": "1",        # 1=TRUNCATE+INSERT, 0=APPEND
        "query_sql": "SELECT ...",
        "rule_group_code": "GR001",
        "rule_name": "测试规则",
    }]
    target_fields = [{
        "rule_code": "UR001",
        "target_field": "product_id",
        "source_field": "product_id",
        "field_type": "bigint",
    }]
    group_variables = [{
        "rule_code": "UR001",
        "var_name": "P_CYCLE_ID",
        "default_value": "20260101",
    }]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("Error: openpyxl required", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════
# RULE sheet 列定义（关键列，其他列留空）
# ═══════════════════════════════════════════════════════════════

RULE_COLUMNS = [
    "租户ID", "组织英文简称", "类型", "项目编码", "项目中文名", "项目英文名",
    "项目描述", "子项目编码", "子项目中文名", "子项目英文名", "子项目描述",
    "规则组编码", "规则组中文名称", "规则组英文名称", "规则组业务责任人",
    "规则组描述", "规则组数据源", "规则编码", "规则中文名称", "规则英文名称",
    "创建方式", "规则类型", "数据源", "备注", "(生成的）查询语句1",
    "(生成的）查询语句2", "(生成的）查询语句3", "(生成的）查询语句4",
    "(生成的）查询语句5", "(生成的）查询语句6", "(生成的）查询语句7",
    "(生成的）查询语句8", "(生成的）查询语句9", "运行条件", "Select Hint语句",
    "执行序列", "源Schema", "目标Schema", "目标SCHEMA解析值", "目标表",
    "目标表解析", "是否去重", "删除模式", "删除条件", "业务责任人",
    "delete hint", "交换分区来源表", "目标表统计信息收集", "行迁移开关",
    "会话变量", "环境变量设置", "并行开关", "事前操作", "事后操作",
    "存储模式", "压缩比", "是否散列", "程序包名", "SP名称", "API参数",
    "更新索引", "循环变量", "规则循环并行调度标志", "循环分组设置",
    "循环优先级", "引用规则", "重试间隔", "重试次数", "不满足时",
    "数据库类型", "调度类型", "指定分区", "来源表统计分析收集",
    "统计分析来源表", "规则描述", "装载字段", "进程数", "运行内存",
    "线程数", "批量大小", "并发数", "spark数据源",
]

TF_COLUMNS = [
    "规则编码", "目标字段名称", "来源字段名称", "加密方式",
    "Merge模式数据源字段值", "别名", "字段类型", "备注",
]

GV_COLUMNS = [
    "规则编码", "动态参数/变量名", "字段类型", "字段定义类型",
    "字段值类型", "变量默认值", "是否校验通过", "数据类型", "描述", "是否必填",
]


# ═══════════════════════════════════════════════════════════════
# 核心函数
# ═══════════════════════════════════════════════════════════════

def build_xlsx(
    output_path: str,
    rules: list[dict],
    target_fields: list[dict] | None = None,
    group_variables: list[dict] | None = None,
    extra_sheets: bool = True,
) -> str:
    """生成 execution_tasks.xlsx。

    Args:
        output_path: 输出路径
        rules: 规则列表，每个 dict 支持 key:
            rule_code, rule_type(1/12), exec_sequence, target_schema,
            target_table, delete_mode, query_sql, rule_group_code,
            rule_name, project_code, data_source
        target_fields: TargetFields 列表，每个 dict 支持 key:
            rule_code, target_field, source_field, field_type, alias, remark
        group_variables: GroupVariables 列表，每个 dict 支持 key:
            rule_code, var_name, default_value
        extra_sheets: 是否创建空的其他 sheet (ModelRelations 等)

    Returns: 输出文件路径
    """
    wb = Workbook()

    # ── RULE sheet ──
    ws_rule = wb.active
    ws_rule.title = "RULE"
    ws_rule.append(RULE_COLUMNS)

    for r in rules:
        row = _build_rule_row(r)
        ws_rule.append(row)

    # ── GroupVariables sheet ──
    ws_gv = wb.create_sheet("GroupVariables")
    ws_gv.append(GV_COLUMNS)
    if group_variables:
        for gv in group_variables:
            ws_gv.append([
                gv.get("rule_code", ""),
                gv.get("var_name", ""),
                gv.get("field_type", ""),
                gv.get("field_def_type", ""),
                gv.get("value_type", ""),
                gv.get("default_value", ""),
                gv.get("validated", ""),
                gv.get("data_type", ""),
                gv.get("description", ""),
                gv.get("required", ""),
            ])

    # ── TargetFields sheet ──
    ws_tf = wb.create_sheet("TargetFields")
    ws_tf.append(TF_COLUMNS)
    if target_fields:
        for tf in target_fields:
            ws_tf.append([
                tf.get("rule_code", ""),
                tf.get("target_field", ""),
                tf.get("source_field", ""),
                tf.get("encryption", ""),
                tf.get("merge_source", ""),
                tf.get("alias", ""),
                tf.get("field_type", ""),
                tf.get("remark", ""),
            ])

    # ── 额外空 sheet（制品包标准结构）──
    if extra_sheets:
        for sheet_name, cols in [
            ("ModelRelations", ["规则编码", "左表schema", "左表名", "左表别名",
             "右表schema", "右表", "右表别名", "模型顺序号", "关联关系",
             "左表字段列表串", "右表字段列表串"]),
            ("ExtraFields", ["规则编码", "拓展字段名", "别名", "表达式",
             "字段类型", "生效", "统计标识"]),
            ("SPParams", ["规则编码", "规则参数名", "数据类型", "入参、出参",
             "变量默认值"]),
            ("Conditions", ["规则编码", "字段名称", "字段关系", "字段值1",
             "字段值2", "与下个条件的逻辑关系", "序号", "字段类型",
             "条件类型", "树形组件业务父类id", "树形组件业务id"]),
            ("MaintenanceParams", ["规则编码", "执行序列", "类型", "schema",
             "表名", "字段名", "分区表名"]),
            ("Extract", ["标签id", "规则编码", "数据库名称", "数据库类型",
             "标签名", "分区读写字段", "分区读写字段类型", "分区数量",
             "分区下界", "分区上界", "批量提取大小", "数据提取SQL",
             "运行SQL", "统计信息分析标识", "统计信息分析来源表信息"]),
            ("ExtractColumn", ["数据标签id", "规则编码", "解密字段", "字段类型",
             "解密类型"]),
        ]:
            ws = wb.create_sheet(sheet_name)
            ws.append(cols)

    # 保存
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)


def _build_rule_row(r: dict) -> list:
    """构建 RULE sheet 行数据（按 RULE_COLUMNS 顺序）。"""
    # 构建列名 → 值的映射
    val_map = {
        "类型": 3,
        "项目编码": r.get("project_code", "TEST_ETL"),
        "项目中文名": r.get("project_cn", "测试ETL"),
        "项目英文名": r.get("project_en", "TEST_ETL"),
        "子项目编码": r.get("sub_project", "待配置"),
        "子项目中文名": r.get("sub_project_cn", "待配置"),
        "子项目英文名": r.get("sub_project_en", "待配置"),
        "规则组编码": r.get("rule_group_code", "GR000001"),
        "规则组中文名称": r.get("rule_group_cn", r.get("rule_group_code", "测试")),
        "规则组英文名称": r.get("rule_group_en", r.get("rule_group_code", "TEST")),
        "规则组业务责任人": r.get("business_owner", "tester"),
        "规则组数据源": r.get("rule_group_ds", "TEST_DS"),
        "规则编码": r.get("rule_code", ""),
        "规则中文名称": r.get("rule_name", r.get("rule_code", "")),
        "规则英文名称": r.get("rule_name_en", r.get("rule_code", "")),
        "创建方式": 2,
        "规则类型": r.get("rule_type", 1),
        "数据源": r.get("data_source", "TEST_DS"),
        "备注": r.get("remark", ""),
        "(生成的）查询语句1": r.get("query_sql", ""),
        "运行条件": 0,
        "执行序列": r.get("exec_sequence", 0),
        "目标Schema": r.get("target_schema", ""),
        "目标表": r.get("target_table", ""),
        "删除模式": r.get("delete_mode", "1"),
        "删除条件": r.get("delete_condition", ""),
        "交换分区来源表": r.get("exchange_source_table", ""),
        "规则中文名称": r.get("rule_name", r.get("rule_code", "")),
        "业务责任人": r.get("owner", "tester"),
        "行迁移开关": 0,
        "并行开关": 0,
        "数据库类型": r.get("db_type", "GaussDB"),
        "调度类型": 0,
    }

    row = []
    for col in RULE_COLUMNS:
        row.append(val_map.get(col))
    return row


# ═══════════════════════════════════════════════════════════════
# CLI（用于单个 case 构建）
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="合成 execution_tasks.xlsx")
    parser.add_argument("--case", required=True, help="case 名称（如 case_01_minimal）")
    parser.add_argument("--output-dir", default=".", help="输出目录")
    args = parser.parse_args()

    # 延迟导入 case 定义
    try:
        case_module = __import__(f"cases.{args.case}", fromlist=["rules"])
    except ImportError:
        print(f"Error: case '{args.case}' not found in cases/", file=sys.stderr)
        sys.exit(1)

    output = Path(args.output_dir) / args.case / "execution_tasks.xlsx"
    result = build_xlsx(
        str(output),
        rules=getattr(case_module, "rules"),
        target_fields=getattr(case_module, "target_fields", None),
        group_variables=getattr(case_module, "group_variables", None),
    )
    print(f"生成: {result}")
