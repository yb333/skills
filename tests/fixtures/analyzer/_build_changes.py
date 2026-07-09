#!/usr/bin/env python3
"""合成变更清单 Excel 辅助工具（三 Sheet 模板）。

用 Python 对象定义生成符合变更清单模板格式的 Excel。

Usage:
    from _build_changes import build_changes_xlsx
    build_changes_xlsx("changes.xlsx",
        table_changes=[TableChange(...)],
        field_changes=[ChangeItem(...)],
    )
"""

from __future__ import annotations

from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise ImportError("openpyxl required")


def build_changes_xlsx(
    path: str,
    table_changes: list = None,
    field_changes: list = None,
    type_dict: dict = None,
    table_headers: list = None,
    field_headers: list = None,
):
    """生成变更清单 Excel（三 Sheet）。

    table_changes: list[TableChange]
    field_changes: list[ChangeItem]
    type_dict: {类型: 说明}
    table_headers / field_headers: 自定义列名（测试模糊匹配用）
    """
    table_changes = table_changes or []
    field_changes = field_changes or []
    type_dict = type_dict or {}

    wb = Workbook()

    # ── Sheet1: 表级 mapping ──
    ws1 = wb.active
    ws1.title = "源系统切换前后表级mapping"
    ws1.append(table_headers or [
        "切换前表名", "切换后表名", "是否平切", "切换说明", "表级变化类型",
    ])
    for tc in table_changes:
        ping_val = "Y" if tc.is_ping_cut else "N"
        ws1.append([tc.before_table, tc.after_table, ping_val, tc.note, tc.change_type])

    # ── Sheet2: 字段级 mapping ──
    ws2 = wb.create_sheet("源系统切换前后字段mapping")
    ws2.append(field_headers or [
        "序号", "切换前数据库", "切换前表schema", "切换前表名",
        "切换前表字段名", "切换前表字段中文名", "切换前表字段类型",
        "切换后数据库", "切换后表schema", "切换后表名",
        "切换后表字段名", "切换后表字段中文名", "切换后表字段类型",
        "字段变化类型", "是否可还原", "还原方案详细说明",
        "源端IT责任人", "源端业务责任人",
    ])
    for i, fc in enumerate(field_changes, 1):
        ws2.append([
            i,
            fc.before_db, fc.before_schema, fc.before_table,
            fc.before_field, fc.before_field_cn, fc.before_type,
            fc.after_db, fc.after_schema, fc.after_table,
            fc.after_field, fc.after_field_cn, fc.after_type,
            fc.change_type, fc.recoverable, fc.recovery_plan,
            fc.source_it_owner, fc.source_biz_owner,
        ])

    # ── Sheet3: 变动类型说明 ──
    ws3 = wb.create_sheet("源端变动类型")
    ws3.append(["类型", "说明"])
    for t, d in type_dict.items():
        ws3.append([t, d])

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
