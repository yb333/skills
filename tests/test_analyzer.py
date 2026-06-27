"""
pytest 测试：dws-pipeline-analyzer 案例库回归测试

测试模式：
- Golden File 对比（结构完整性）：每个 case 断言关键指标
- xfail 标记（已知能力缺口）

运行：
    pytest tests/test_analyzer.py -v
    pytest tests/test_analyzer.py -k case_02 -v  # 单个 case
"""

import sys
import json
from pathlib import Path

import pytest

# ── 路径配置 ──────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent  # skills 仓库根
ANALYZER_SKILL = PROJECT_ROOT / "dws-pipeline-analyzer" / "references"
FIXTURES_DIR = PROJECT_ROOT / "tests" / "fixtures" / "analyzer"

# 添加路径
sys.path.insert(0, str(ANALYZER_SKILL))
sys.path.insert(0, str(FIXTURES_DIR))
sys.path.insert(0, str(FIXTURES_DIR / "cases"))

from analyzer import (
    read_excel, detect_dialect, parse_single_sql,
    build_topology, build_data_flow, build_field_mappings, analyze_quality,
)
from _build_xlsx import build_xlsx


# ── Fixtures ──────────────────────────────────────────────

@pytest.fixture(scope="session")
def ensure_xlsx_generated():
    """确保所有 case 的 Excel 已生成（session 级缓存）。"""
    cases = [
        "case_01_minimal", "case_02_cte_basic", "case_03_cte_nested",
        "case_04_pivot", "case_05_window",
        "case_06_multi_step_parallel", "case_07_multi_step_serial",
        "case_08_self_reference", "case_09_same_target_writes", "case_10_view_step",
        "case_11_union_all", "case_12_oracle_dialect", "case_13_field_mismatch",
        "case_14_audit_fields", "case_15_comment_alias",
        "case_16_many_joins", "case_17_many_ctes", "case_18_many_case_when",
    "case_19_multi_scenario_2", "case_20_multi_scenario_3",
        "case_21_scenario_with_common", "case_22_scenario_chain",
        "case_23_exchange_partition",
        "case_25_table_case_dedup",
        "case_26_field_usage",
    ]
    for case_name in cases:
        xlsx_path = FIXTURES_DIR / case_name / "execution_tasks.xlsx"
        if not xlsx_path.exists():
            mod = __import__(case_name)
            build_xlsx(
                str(xlsx_path),
                rules=mod.rules,
                target_fields=getattr(mod, "target_fields", None),
                group_variables=getattr(mod, "group_variables", None),
            )
    yield


def run_analysis(case_name: str) -> dict:
    """运行 analyzer 并返回完整结果。"""
    xlsx_path = FIXTURES_DIR / case_name / "execution_tasks.xlsx"
    raw = read_excel(str(xlsx_path))
    rules = raw["rules"]
    sql_texts = [r.query_sql for r in rules if r.query_sql]
    dialect = detect_dialect(sql_texts)

    parsed_map = {}
    for rule in rules:
        parsed_map[rule.rule_code] = parse_single_sql(rule.query_sql, dialect)

    topology = build_topology(rules, parsed_map)
    data_flow = build_data_flow(rules, parsed_map)
    field_mappings = build_field_mappings(rules, parsed_map, raw["target_fields"])
    quality = analyze_quality(topology, data_flow, field_mappings, parsed_map)

    return {
        "raw": raw,
        "rules": rules,
        "dialect": dialect,
        "parsed_map": parsed_map,
        "topology": topology,
        "data_flow": data_flow,
        "field_mappings": field_mappings,
        "quality": quality,
    }


def get_transform_types(result: dict) -> dict:
    """提取 target_field → transform_type 映射。"""
    return {
        f["target_field"]: f.get("transform_type", "unknown")
        for f in result["field_mappings"]["fields"]
    }


# ═══════════════════════════════════════════════════════════════
# Tier 1: 基础覆盖
# ═══════════════════════════════════════════════════════════════

class TestTier1Basic:
    """Tier 1: 基础 happy path 验证。"""

    def test_case_01_minimal(self, ensure_xlsx_generated):
        """最简基线：1 step，5 字段全 direct，0 issues。"""
        r = run_analysis("case_01_minimal")
        assert len(r["topology"]["steps"]) == 1
        assert len(r["field_mappings"]["fields"]) == 5
        assert len(r["quality"]["issues"]) == 0

        tt = get_transform_types(r)
        for field, t in tt.items():
            assert t == "direct", f"{field} 应为 direct，实际 {t}"

    def test_case_02_cte_basic(self, ensure_xlsx_generated):
        """单层 CTE：1 个 CTE 被正确提取。"""
        r = run_analysis("case_02_cte_basic")
        ctes = r["data_flow"]["steps"][0]["ctes"]
        assert len(ctes) == 1, f"应有 1 个 CTE，实际 {len(ctes)}"
        assert ctes[0]["name"] == "agg"

        # CTE 内字段含 transform_type
        agg_fields = {f["name"]: f for f in ctes[0]["fields"]}
        assert "total" in agg_fields
        assert agg_fields["total"].get("transform_type") == "aggregate"

    def test_case_03_cte_nested(self, ensure_xlsx_generated):
        """嵌套 CTE：CTE_A 引用 CTE_B，穿透传播生效。"""
        r = run_analysis("case_03_cte_nested")
        ctes = r["data_flow"]["steps"][0]["ctes"]
        assert len(ctes) == 2, f"应有 2 个 CTE（base, agg），实际 {len(ctes)}"

        # 验证 total 字段穿透后 transform_type=aggregate（不是 direct）
        tt = get_transform_types(r)
        assert tt.get("total") == "aggregate", \
            f"total 应穿透为 aggregate（CTE 内 SUM），实际 {tt.get('total')}"

        # 验证穿透链：total 的 lineage 应有 cte_name
        total_field = next(
            (f for f in r["field_mappings"]["fields"] if f["target_field"] == "total"),
            None
        )
        assert total_field is not None
        lineages = total_field.get("lineage", [])
        cte_lineages = [l for l in lineages if l.get("cte_name")]
        assert len(cte_lineages) > 0, "total 字段应有 CTE 穿透信息"

    def test_case_04_pivot(self, ensure_xlsx_generated):
        """行转列：3 个字段 transform_type=pivot。"""
        r = run_analysis("case_04_pivot")
        tt = get_transform_types(r)
        for field in ("jan_amt", "feb_amt", "mar_amt"):
            assert tt.get(field) == "pivot", f"{field} 应为 pivot，实际 {tt.get(field)}"

        cw = r["quality"]["complexity_metrics"]["total_case_when_branches"]
        assert cw == 3, f"CASE WHEN 分支应为 3，实际 {cw}"

    def test_case_05_window(self, ensure_xlsx_generated):
        """窗口函数：rn 和 prev_login transform_type=window。"""
        r = run_analysis("case_05_window")
        tt = get_transform_types(r)
        assert tt.get("rn") == "window", f"rn 应为 window，实际 {tt.get('rn')}"
        assert tt.get("prev_login") == "window", f"prev_login 应为 window，实际 {tt.get('prev_login')}"


# ═══════════════════════════════════════════════════════════════
# Tier 2: 结构复杂度
# ═══════════════════════════════════════════════════════════════

class TestTier2Structure:
    """Tier 2: 多步骤拓扑结构验证。"""

    def test_case_06_multi_step_parallel(self, ensure_xlsx_generated):
        """多步骤并行：2 个 rule exec_sequence 相同。"""
        r = run_analysis("case_06_multi_step_parallel")
        assert len(r["topology"]["steps"]) == 2
        parallel = r["topology"]["schedule_plan"][0]["parallel_steps"]
        assert len(parallel) == 2, f"应有 2 个并行步骤，实际 {len(parallel)}"

    def test_case_07_multi_step_serial(self, ensure_xlsx_generated):
        """多步骤串行：跨步骤数据依赖。"""
        r = run_analysis("case_07_multi_step_serial")
        deps = r["topology"]["data_dependencies"]
        assert len(deps) >= 1, "应有跨步骤数据依赖"
        # step_1 → step_2
        assert any(d["from"] == "step_1" and d["to"] == "step_2" for d in deps), \
            f"应存在 step_1→step_2 依赖，实际 {deps}"

    def test_case_08_self_reference(self, ensure_xlsx_generated):
        """自引用：WHERE EXISTS(SELECT 1 FROM 目标表)。"""
        r = run_analysis("case_08_self_reference")
        self_refs = r["topology"]["self_references"]
        assert len(self_refs) >= 1, "应检测到自引用"
        assert "EXISTS" in self_refs[0].get("pattern", "").upper(), \
            f"自引用模式应含 EXISTS，实际 {self_refs[0].get('pattern')}"

    def test_case_09_same_target_writes(self, ensure_xlsx_generated):
        """同表多次写入：target_write_groups 非空。"""
        r = run_analysis("case_09_same_target_writes")
        groups = r["topology"]["target_write_groups"]
        assert len(groups) >= 1, "应检测到同表多次写入组"

    def test_case_10_view_step(self, ensure_xlsx_generated):
        """视图步骤：CREATE VIEW + 跨步骤依赖。"""
        r = run_analysis("case_10_view_step")
        assert len(r["topology"]["steps"]) == 2
        # step_2 (view) 应依赖 step_1 (table)
        deps = r["topology"]["data_dependencies"]
        assert any(d["from"] == "step_1" and d["to"] == "step_2" for d in deps), \
            "视图步骤应依赖事实表步骤"


# ═══════════════════════════════════════════════════════════════
# Tier 3: 边界场景
# ═══════════════════════════════════════════════════════════════

class TestTier3Boundary:
    """Tier 3: 边界场景和已知缺口。"""

    def test_case_11_union_all(self, ensure_xlsx_generated):
        """UNION ALL：两个分支的 source_tables 都应被提取。"""
        r = run_analysis("case_11_union_all")
        
        # source_tables 应包含两个分支的表
        step1 = r["topology"]["steps"][0]
        all_tables = step1.get("all_tables_from_sql", [])
        assert "ods.orders_a" in all_tables, f"应包含 orders_a，实际 {all_tables}"
        assert "ods.orders_b" in all_tables, f"应包含 orders_b，实际 {all_tables}"

        # 字段以第一个分支为准，3 列
        fields = r["field_mappings"]["fields"]
        field_names = {f["target_field"] for f in fields}
        assert "order_id" in field_names
        assert "source" in field_names
        assert "amount" in field_names

    def test_case_12_oracle_dialect(self, ensure_xlsx_generated):
        """Oracle 方言：NVL 检测。"""
        r = run_analysis("case_12_oracle_dialect")
        assert r["dialect"] == "oracle", f"应检测为 oracle 方言，实际 {r['dialect']}"

        tt = get_transform_types(r)
        # NVL(x, 0) 应识别为 fallback
        assert tt.get("price") == "fallback", f"price (NVL) 应为 fallback，实际 {tt.get('price')}"

    def test_case_13_field_mismatch(self, ensure_xlsx_generated):
        """字段别名不匹配：differences 应非空。"""
        r = run_analysis("case_13_field_mismatch")
        stats = r["field_mappings"]["statistics"]
        # 应有 only_in_excel（extra_field）或 only_in_sql
        diff_count = len(stats.get("only_in_sql", [])) + len(stats.get("only_in_excel", []))
        assert diff_count > 0, "应检测到字段差异"

    def test_case_14_audit_fields(self, ensure_xlsx_generated):
        """审计字段推断：'N'→del_flag, CURRENT_TIMESTAMP→dw_last_update_date。"""
        r = run_analysis("case_14_audit_fields")
        tt = get_transform_types(r)
        assert "del_flag" in tt, "应推断出 del_flag 字段名"
        assert "dw_last_update_date" in tt, "应推断出 dw_last_update_date 字段名"
        assert tt["del_flag"] == "value", f"del_flag 应为 value，实际 {tt['del_flag']}"

    def test_case_15_comment_alias(self, ensure_xlsx_generated):
        """注释别名提取：/* field_name */ 格式。"""
        r = run_analysis("case_15_comment_alias")
        tt = get_transform_types(r)
        # 注释中的字段名应被提取
        assert "del_flag" in tt, "应从注释提取 del_flag"
        assert "dw_last_update_date" in tt, "应从注释提取 dw_last_update_date"


# ═══════════════════════════════════════════════════════════════
# Tier 4: 性能阈值
# ═══════════════════════════════════════════════════════════════

class TestTier4Performance:
    """Tier 4: 性能阈值告警。"""

    def test_case_16_many_joins(self, ensure_xlsx_generated):
        """多表 JOIN（9张）：验证复杂度指标统计正确。"""
        r = run_analysis("case_16_many_joins")
        cm = r["quality"]["complexity_metrics"]
        assert cm["max_join_count"] >= 9, f"JOIN 数应 ≥9，实际 {cm['max_join_count']}"

    def test_case_17_many_ctes(self, ensure_xlsx_generated):
        """多 CTE（4个）：触发 medium 复杂度告警。"""
        r = run_analysis("case_17_many_ctes")
        cm = r["quality"]["complexity_metrics"]
        assert cm["max_cte_count"] >= 4, f"CTE 数应 ≥4，实际 {cm['max_cte_count']}"

        medium_issues = [i for i in r["quality"]["issues"] if i["severity"] == "medium"]
        assert len(medium_issues) >= 1, "应触发 medium 级 CTE 嵌套过深告警"

    def test_case_18_many_case_when(self, ensure_xlsx_generated):
        """多 CASE WHEN（21个）：触发 medium 复杂度告警。"""
        r = run_analysis("case_18_many_case_when")
        cm = r["quality"]["complexity_metrics"]
        assert cm["total_case_when_branches"] >= 21, \
            f"CASE WHEN 分支应 ≥21，实际 {cm['total_case_when_branches']}"

        medium_issues = [i for i in r["quality"]["issues"] if i["severity"] == "medium"]
        assert len(medium_issues) >= 1, "应触发 medium 级 CASE WHEN 过多告警"


# ═══════════════════════════════════════════════════════════════
# Tier 5: 多场景综合
# ═══════════════════════════════════════════════════════════════

class TestTier5MultiScenario:
    """多场景分区写入、场景+公共混合、串行依赖链。"""

    def test_case_19_multi_scenario_2(self, ensure_xlsx_generated):
        """2场景分区写入 + 公共步骤：场景数=3。"""
        r = run_analysis("case_19_multi_scenario_2")
        scenarios = r["topology"]["scenarios"]
        # 2个分区场景 + 1个公共步骤
        non_common = [s for s in scenarios if not s.get("is_common")]
        common = [s for s in scenarios if s.get("is_common")]
        assert len(non_common) == 2, f"应有2个分区场景，实际 {len(non_common)}"
        assert len(common) == 1, f"应有1个公共步骤，实际 {len(common)}"
        # 每个分区场景含2个规则
        for sc in non_common:
            assert sc["rule_count"] == 2, f"场景 {sc['name']} 应含2个规则"

    def test_case_20_multi_scenario_3(self, ensure_xlsx_generated):
        """3场景分区写入：场景数=3。"""
        r = run_analysis("case_20_multi_scenario_3")
        scenarios = r["topology"]["scenarios"]
        non_common = [s for s in scenarios if not s.get("is_common")]
        assert len(non_common) == 3, f"应有3个分区场景，实际 {len(non_common)}"

    def test_case_21_scenario_with_common(self, ensure_xlsx_generated):
        """场景+公共混合：2个分区场景 + 1个公共步骤。"""
        r = run_analysis("case_21_scenario_with_common")
        scenarios = r["topology"]["scenarios"]
        non_common = [s for s in scenarios if not s.get("is_common")]
        common = [s for s in scenarios if s.get("is_common")]
        assert len(non_common) == 2
        assert len(common) == 1
        # 验证公共步骤的删除模式
        common_step = next(s for s in r["topology"]["steps"] if s.get("is_common_step"))
        assert common_step["delete_mode_label"] == "TRUNCATE TABLE"

    def test_case_22_scenario_chain(self, ensure_xlsx_generated):
        """单场景3步串行：TRUNCATE → APPEND → APPEND。"""
        r = run_analysis("case_22_scenario_chain")
        steps = r["topology"]["steps"]
        assert len(steps) == 3
        # 第1步 TRUNCATE TABLE
        assert steps[0]["delete_mode_label"] == "TRUNCATE TABLE"
        # 第2/3步 NO DELETE
        assert steps[1]["delete_mode_label"] == "NO DELETE (追加)"
        assert steps[2]["delete_mode_label"] == "NO DELETE (追加)"
        # 同目标表写入组（3个规则写同一张表）
        write_groups = r["topology"]["target_write_groups"]
        assert len(write_groups) >= 1, "应检测到同表多次写入组"

    def test_case_23_exchange_partition(self, ensure_xlsx_generated):
        """分区交换：临时表 → 真正目标表。"""
        r = run_analysis("case_23_exchange_partition")
        steps = r["topology"]["steps"]
        assert len(steps) == 2

        # step_2 是分区交换，目标表应该是 exchange_source_table（不是临时表）
        exchange_step = steps[1]
        assert exchange_step.get("is_exchange") == True
        assert exchange_step["target_table"] == "dwl_real_f", \
            f"分区交换目标表应为 dwl_real_f，实际 {exchange_step['target_table']}"
        assert exchange_step.get("exchange_temp_table") == "dwl_temp_f"

        # 兜底描述应包含分区交换信息
        descs = r["topology"]["scenarios"]
        # step_2 应标注为分区交换类型
        assert exchange_step.get("rule_type") == 9


# ═══════════════════════════════════════════════════════════════
# Tier 6: Excel 读取健壮性
# ═══════════════════════════════════════════════════════════════

class TestExcelReadRobustness:
    """read_excel 必须对空 sheet / 缺失标题行优雅降级，不抛异常。

    生产环境用户导出的 xlsx 格式不可控（空 sheet、标题行缺失、
    被误清空的 sheet 都可能出现）。read_excel 用 next(ws.iter_rows(...))
    读取标题行——当前 read_only=False 模式下空 sheet 会返回全 None 行，
    不抛 StopIteration；但这是脆弱契约：一旦有人改成 read_only=True，
    空 sheet 就会抛 StopIteration 让整个分析崩溃。

    这组测试锁定"空 sheet 优雅降级"的行为契约，作为防御性回归。
    """

    @staticmethod
    def _make_xlsx(path, sheet_def):
        """生成 xlsx。sheet_def: dict[sheet_name, rows]；空 rows = 完全空 sheet。"""
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认 sheet
        for sheet_name, rows in sheet_def.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        wb.save(str(path))
        wb.close()
        return str(path)

    @staticmethod
    def _clear_sheet(path, sheet_name):
        """清空指定 sheet 的所有行（含标题行）。"""
        import openpyxl
        wb = openpyxl.load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row or 1)
        wb.save(path)
        wb.close()

    def test_empty_rule_sheet_does_not_crash(self, tmp_path):
        """RULE sheet 存在但完全空（无标题行）：优雅降级返回空 rules"""
        xlsx = self._make_xlsx(tmp_path / "empty_rule.xlsx", {"RULE": []})
        # 关键：调用不抛异常
        raw = read_excel(xlsx)
        assert raw["rules"] == [], "空 RULE sheet 应返回空 rules"

    def test_empty_target_fields_sheet(self, tmp_path):
        """TargetFields sheet 空但 RULE 正常：不应抛异常"""
        from _build_xlsx import build_xlsx
        from case_01_minimal import rules as minimal_rules
        xlsx = str(tmp_path / "tf_empty.xlsx")
        build_xlsx(xlsx, rules=minimal_rules)
        self._clear_sheet(xlsx, "TargetFields")
        # 不抛异常
        raw = read_excel(xlsx)
        assert len(raw["rules"]) >= 1, "RULE 应正常解析"
        assert raw["target_fields"] == {} or all(
            not v for v in raw["target_fields"].values()
        ), "空 TargetFields 应无字段"

    def test_empty_group_variables_sheet(self, tmp_path):
        """GroupVariables sheet 空但 RULE 正常：不应抛异常"""
        from _build_xlsx import build_xlsx
        from case_01_minimal import rules as minimal_rules
        xlsx = str(tmp_path / "gv_empty.xlsx")
        build_xlsx(xlsx, rules=minimal_rules)
        self._clear_sheet(xlsx, "GroupVariables")
        # 不抛异常
        raw = read_excel(xlsx)
        assert len(raw["rules"]) >= 1, "RULE 应正常解析"
        assert raw["group_variables"] == {}, "空 GV 应返回空 dict"


# ═══════════════════════════════════════════════════════════════
# Tier 7: exec_sequence 解析健壮性
# ═══════════════════════════════════════════════════════════════

class TestExecSequenceParsing:
    """exec_sequence 必须正确处理数值/字符串/浮点字符串格式。

    Bug: int(exec_seq_str) 对字符串 "1.0" 会 ValueError → 塌缩为 0，
    导致调度图扁平化、串行依赖丢失。制品包 Excel 可能以文本格式存储执行序列。
    """

    def test_float_string_exec_sequence(self, tmp_path):
        """执行序列存为字符串 '2.0' 应正确解析为 2，不塌缩为 0"""
        import openpyxl
        from _build_xlsx import build_xlsx, RULE_COLUMNS
        from case_01_minimal import rules as minimal_rules

        xlsx = str(tmp_path / "float_seq.xlsx")
        build_xlsx(xlsx, rules=minimal_rules)
        # 把执行序列列改成字符串 "2.0"
        wb = openpyxl.load_workbook(xlsx)
        ws = wb["RULE"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1))]
        if "执行序列" in headers:
            col_idx = headers.index("执行序列") + 1
            for row in ws.iter_rows(min_row=2):
                if row[col_idx - 1].value is not None:
                    row[col_idx - 1].value = "2.0"  # 字符串浮点
        wb.save(xlsx)
        wb.close()

        raw = read_excel(xlsx)
        assert len(raw["rules"]) >= 1
        # 关键：字符串 "2.0" 应解析为 2，不是 0
        assert raw["rules"][0].exec_sequence == 2, \
            f"字符串 '2.0' 应解析为 2，实际 {raw['rules'][0].exec_sequence}"

    def test_integer_exec_sequence_unchanged(self, tmp_path):
        """整数执行序列不受影响"""
        from _build_xlsx import build_xlsx
        from case_01_minimal import rules as minimal_rules
        xlsx = str(tmp_path / "int_seq.xlsx")
        build_xlsx(xlsx, rules=minimal_rules)
        raw = read_excel(xlsx)
        # case_01 的 exec_sequence 默认值（int 格式，不受 bug 影响）
        assert isinstance(raw["rules"][0].exec_sequence, int)


# ═══════════════════════════════════════════════════════════════
# Tier 8: 子查询统计正确性
# ═══════════════════════════════════════════════════════════════

class TestSubqueryCount:
    """subquery_count 应只数子查询本身，不数子查询内部的物理表。

    Bug: 原实现按 join_type 含 "subquery" 计数，把 FROM_SUBQUERY_MAIN/
    FROM_SUBQUERY/JOIN_SUBQUERY_INNER（子查询内部物理表）也算进去，
    导致一个含 2 表的子查询被记成 2 个子查询。
    """

    def test_single_subquery_counted_as_one(self):
        """1 个 FROM 子查询（内部 2 表）应记为 1 个子查询"""
        import tempfile
        sql = "SELECT t.x FROM (SELECT a.x FROM ods.t1 a LEFT JOIN ods.t2 b ON a.id=b.id) t"
        rules = [{"rule_code":"R1","rule_type":1,"exec_sequence":0,
                  "target_schema":"dws","target_table":"t_f","delete_mode":"1",
                  "query_sql":sql,"rule_name":"t"}]
        xlsx = tempfile.mktemp(suffix=".xlsx")
        build_xlsx(xlsx, rules=rules)
        raw = read_excel(xlsx)
        pm = {r.rule_code: parse_single_sql(r.query_sql, "dws") for r in raw["rules"]}
        topo = build_topology(raw["rules"], pm)
        df = build_data_flow(raw["rules"], pm)
        fm = build_field_mappings(raw["rules"], pm, {})
        q = analyze_quality(topo, df, fm, pm)
        assert q["complexity_metrics"]["max_subquery_count"] == 1, \
            f"1 个子查询应记为 1，实际 {q['complexity_metrics']['max_subquery_count']}"
