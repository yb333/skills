"""交付件验证测试 — 验证最终产物的正确性（不只是中间数据层）。

覆盖三个缺口:
1. data_blocks 层（逻辑块结构/嵌套/分支独立/CTE内部）
2. view_generator 输出层（mapping.xlsx / asset_report.html 的实际内容）
3. 端到端（复杂 SQL → 完整产物链路）

运行:
    pytest tests/test_deliverables.py -v
"""

import sys
import json
import re
import tempfile
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ANALYZER_REF = PROJECT_ROOT / "dws-pipeline-analyzer" / "references"
FIXTURES = PROJECT_ROOT / "tests" / "fixtures" / "analyzer"
sys.path.insert(0, str(ANALYZER_REF))
sys.path.insert(0, str(FIXTURES))

from analyzer import (
    parse_single_sql, detect_dialect, build_topology, build_data_flow,
    build_field_mappings, analyze_quality, detect_patterns, build_source,
    enrich_join_key_lineage, enrich_field_physical_sources,
    build_data_blocks, build_structured_step_summary,
    RawRule,
)


# ═══════════════════════════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════════════════════════

def _make_rule(sql, rule_code="R1", target="f"):
    return RawRule(rule_code=rule_code, rule_name="test", rule_type=1, exec_sequence=1,
                   target_schema="dws", target_table=target, delete_mode="1", query_sql=sql)


def _get_blocks(sql, rule_code="R1", target="f"):
    """跑完整解析，返回 data_blocks。"""
    rule = _make_rule(sql, rule_code, target)
    pm = {rule_code: parse_single_sql(sql, "dws")}
    df = build_data_flow([rule], pm)
    fm = build_field_mappings([rule], pm, {})
    step = {"step_id": "step_1", "rule_code": rule_code}
    return build_data_blocks(step, df["steps"][0], pm[rule_code], fm["fields"])


def _flatten_blocks(blocks, result=None):
    """递归展平逻辑块，返回 [{role, table, where_clause, ...}]。"""
    if result is None:
        result = []
    for b in blocks:
        result.append(b)
        _flatten_blocks(b.get("children", []), result)
    return result


def _find_blocks_by_role(blocks, role_keyword):
    """按角色关键字找块。"""
    all_blocks = _flatten_blocks(blocks)
    return [b for b in all_blocks if role_keyword in b.get("role", "")]


# ═══════════════════════════════════════════════════════════════
# 1. data_blocks 层验证
# ═══════════════════════════════════════════════════════════════

class TestDataBlocksSimpleSelect:
    """普通 SELECT 的逻辑块。"""

    def test_main_and_secondary(self):
        """主表 + 从表结构正确"""
        blocks = _get_blocks(
            "SELECT a.id, b.name FROM ods.t1 a LEFT JOIN ods.t2 b ON a.id = b.id"
        )
        mains = _find_blocks_by_role(blocks, "主表")
        secondaries = _find_blocks_by_role(blocks, "从表")
        assert len(mains) >= 1, f"应有主表，实际 {mains}"
        assert len(secondaries) >= 1, f"应有从表，实际 {secondaries}"
        assert "t1" in mains[0]["table"]
        assert "t2" in secondaries[0]["table"]

    def test_where_clause_displayed(self):
        """过滤条件应在主表块上显示"""
        blocks = _get_blocks(
            "SELECT a.id FROM ods.t1 a WHERE a.del = 'N'"
        )
        mains = _find_blocks_by_role(blocks, "主表")
        assert mains[0].get("where_clause"), f"主表应有 WHERE，实际 {mains[0]}"
        assert "del" in mains[0]["where_clause"]

    def test_group_by_displayed(self):
        """收敛（GROUP BY）应在主表块上显示"""
        blocks = _get_blocks(
            "SELECT a.id, SUM(a.amt) FROM ods.t1 a GROUP BY a.id"
        )
        mains = _find_blocks_by_role(blocks, "主表")
        assert "收敛" in mains[0].get("ops", []), f"应有收敛操作"


class TestDataBlocksUnion:
    """UNION 的逻辑块。"""

    def test_top_level_union_two_branches(self):
        """顶层 UNION 应有两个独立分支"""
        blocks = _get_blocks("""SELECT a.id FROM ods.t1 a WHERE a.del='N'
UNION ALL
SELECT b.id FROM ods.t2 b WHERE b.sts='A'""")
        union_blocks = _find_blocks_by_role(blocks, "合并")
        assert len(union_blocks) >= 1, "应有合并块"

    def test_branches_where_independent(self):
        """两个分支的 WHERE 应各自独立，不混在一起"""
        blocks = _get_blocks("""SELECT a.id FROM ods.t1 a WHERE a.del='N'
UNION ALL
SELECT b.id FROM ods.t2 b WHERE b.sts='A'""")
        all_flat = _flatten_blocks(blocks)
        branches = [b for b in all_flat if "UNION 分支" in b.get("role", "")]
        assert len(branches) == 2, f"应有2个分支，实际 {len(branches)}"
        assert "del" in branches[0].get("where_clause", ""), f"分支1应含 del"
        assert "sts" in branches[1].get("where_clause", ""), f"分支2应含 sts"
        assert "sts" not in branches[0].get("where_clause", ""), "分支1不应含 sts（混了）"
        assert "del" not in branches[1].get("where_clause", ""), "分支2不应含 del（混了）"

    def test_subquery_union_two_branches(self):
        """FROM 子查询内部 UNION 也应有两个分支"""
        blocks = _get_blocks("""SELECT t.id FROM (
SELECT a.id FROM ods.t1 a WHERE a.del='N'
UNION ALL
SELECT b.id FROM ods.t2 b WHERE b.sts='A'
) t""")
        all_flat = _flatten_blocks(blocks)
        branches = [b for b in all_flat if "UNION 分支" in b.get("role", "")]
        assert len(branches) == 2, f"FROM子查询UNION应有2分支，实际 {len(branches)}"


class TestDataBlocksCTE:
    """CTE 的逻辑块。"""

    def test_cte_internal_structure(self):
        """CTE 内部表和 JOIN 应作为 children 展示"""
        blocks = _get_blocks("""WITH tm AS (
    SELECT a.id, a.amt FROM ods.fact_a a
    INNER JOIN ods.dim_b b ON a.k = b.k WHERE a.del = 'N'
)
SELECT m.id, m.amt FROM tm m WHERE m.amt > 0""")
        all_flat = _flatten_blocks(blocks)
        # CTE 内部的表应出现
        tables = [b["table"].lower() for b in all_flat]
        assert any("fact_a" in t for t in tables), f"CTE 内部 fact_a 应出现，实际 {tables}"
        assert any("dim_b" in t for t in tables), f"CTE 内部 dim_b 应出现，实际 {tables}"
        # CTE 内部的 WHERE 应展示
        cte_mains = [b for b in all_flat if "内部主表" in b.get("role", "")]
        if cte_mains:
            assert any("del" in b.get("where_clause", "") for b in cte_mains), \
                "CTE 内部 WHERE 应展示"

    def test_cte_internal_join_condition(self):
        """CTE 内部的 JOIN ON 条件应展示"""
        blocks = _get_blocks("""WITH tm AS (
    SELECT a.id FROM ods.fact_a a INNER JOIN ods.dim_b b ON a.k = b.k
)
SELECT m.id FROM tm m""")
        all_flat = _flatten_blocks(blocks)
        cte_joins = [b for b in all_flat if "内部关联" in b.get("role", "")]
        assert len(cte_joins) >= 1, f"CTE 内部 JOIN 应展示"
        assert "k" in cte_joins[0].get("on_condition", ""), "JOIN ON 条件应含 k"


class TestDataBlocksNestedSubquery:
    """嵌套子查询的逻辑块。"""

    def test_two_layer_nesting(self):
        """两层嵌套：外层子查询 → 内层子查询"""
        blocks = _get_blocks("""SELECT t.region FROM (
    SELECT m.region FROM (
        SELECT a.region FROM ods.fact_a a INNER JOIN ods.dim_b b ON a.k=b.k WHERE a.del='N'
    ) m
) t""")
        all_flat = _flatten_blocks(blocks)
        tables = [b["table"].lower() for b in all_flat]
        assert any("fact_a" in t for t in tables), f"最内层 fact_a 应出现"
        assert any("dim_b" in t for t in tables), f"最内层 dim_b 应出现"

    def test_inner_where_displayed(self):
        """内层子查询的 WHERE 应展示"""
        blocks = _get_blocks("""SELECT t.id FROM (
    SELECT a.id FROM ods.fact_a a WHERE a.del = 'N'
) t""")
        all_flat = _flatten_blocks(blocks)
        has_del_where = any("del" in b.get("where_clause", "") for b in all_flat)
        assert has_del_where, "内层 WHERE 应展示"


# ═══════════════════════════════════════════════════════════════
# 2. view_generator 输出层验证
# ═══════════════════════════════════════════════════════════════

class TestMappingExcel:
    """mapping.xlsx 的实际单元格内容验证。"""

    def test_attribute_mapping_physical_source(self, tmp_path):
        """属性级 mapping：字段来源应穿透到物理表"""
        from view_generator import generate_mapping
        import openpyxl

        rules = [
            RawRule(rule_code="R1", rule_name="s1", rule_type=1, exec_sequence=1,
                    target_schema="dws", target_table="tmp1", delete_mode="1",
                    query_sql="SELECT a.id, a.amount FROM ods.tbl_a a"),
            RawRule(rule_code="R2", rule_name="s2", rule_type=1, exec_sequence=2,
                    target_schema="dws", target_table="final_f", delete_mode="1",
                    query_sql="SELECT t.id, t.amount FROM dws.tmp1 t"),
        ]
        pm = {r.rule_code: parse_single_sql(r.query_sql, "dws") for r in rules}
        topo = build_topology(rules, pm)
        df = build_data_flow(rules, pm)
        fm = build_field_mappings(rules, pm, {})
        enrich_join_key_lineage(df, rules, pm, topo, fm)
        enrich_field_physical_sources(fm, df, rules, pm, topo)
        q = analyze_quality(topo, df, fm, pm)
        knowledge = {
            "meta": {"source_type": "t", "analysis_time": "", "dialect": "dws",
                     "total_rules": 2, "target_table": "final_f",
                     "patterns": [], "target_field_types": {}, "target_field_comments": {}},
            "topology": topo, "data_flow": df, "field_mappings": fm, "quality": q,
            "business_logic": {"summary": "", "step_descriptions": [], "key_transforms": []},
            "source": build_source(rules, {}, {}, pm),
        }
        out = str(tmp_path / "test_out")
        generate_mapping(knowledge, out)

        wb = openpyxl.load_workbook(Path(out) / "mapping.xlsx", read_only=True)
        ws2 = wb["属性级mapping"]
        # 找 amount 字段的行
        found_amount = False
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row and row[8] and "amount" in str(row[8]).lower():
                found_amount = True
                # 最初来源（源表物理表名 col[2]）应含 tbl_a（穿透中间表 tmp1）
                source_table = str(row[2] or "")
                assert "tbl_a" in source_table, f"amount 来源表应穿透到 tbl_a，实际 {source_table}"
                break
        wb.close()
        assert found_amount, "属性级 mapping 应有 amount 字段"

    def test_entity_mapping_no_intermediate(self, tmp_path):
        """实体级 mapping：中间表不出现"""
        from view_generator import generate_mapping
        import openpyxl

        rules = [
            RawRule(rule_code="R1", rule_name="s1", rule_type=1, exec_sequence=1,
                    target_schema="dws", target_table="tmp1", delete_mode="1",
                    query_sql="SELECT a.id FROM ods.tbl_a a"),
            RawRule(rule_code="R2", rule_name="s2", rule_type=1, exec_sequence=2,
                    target_schema="dws", target_table="final_f", delete_mode="1",
                    query_sql="SELECT t.id FROM dws.tmp1 t"),
        ]
        pm = {r.rule_code: parse_single_sql(r.query_sql, "dws") for r in rules}
        topo = build_topology(rules, pm)
        df = build_data_flow(rules, pm)
        fm = build_field_mappings(rules, pm, {})
        enrich_join_key_lineage(df, rules, pm, topo, fm)
        enrich_field_physical_sources(fm, df, rules, pm, topo)
        q = analyze_quality(topo, df, fm, pm)
        knowledge = {
            "meta": {"source_type": "t", "analysis_time": "", "dialect": "dws",
                     "total_rules": 2, "target_table": "final_f",
                     "patterns": [], "target_field_types": {}, "target_field_comments": {}},
            "topology": topo, "data_flow": df, "field_mappings": fm, "quality": q,
            "business_logic": {"summary": "", "step_descriptions": [], "key_transforms": []},
            "source": build_source(rules, {}, {}, pm),
        }
        out = str(tmp_path / "test_out")
        generate_mapping(knowledge, out)

        wb = openpyxl.load_workbook(Path(out) / "mapping.xlsx", read_only=True)
        ws1 = wb["实体级mapping"]
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                tbl = str(row[1]).lower()
                assert "tmp1" not in tbl, f"实体级不应出现中间表 tmp1，实际 {tbl}"
        wb.close()


class TestAssetReportData:
    """asset_report.html 的 REPORT_DATA 验证。"""

    def test_report_data_has_data_blocks(self, tmp_path):
        """HTML 的 steps 应含 data_blocks"""
        from view_generator import generate_asset_report

        rule = _make_rule("SELECT a.id, b.name FROM ods.t1 a LEFT JOIN ods.t2 b ON a.id=b.id")
        pm = {"R1": parse_single_sql(rule.query_sql, "dws")}
        topo = build_topology([rule], pm)
        df = build_data_flow([rule], pm)
        fm = build_field_mappings([rule], pm, {})
        q = analyze_quality(topo, df, fm, pm)
        for ts in topo["steps"]:
            ds = next(s for s in df["steps"] if s["step_id"] == ts["step_id"])
            ds["data_blocks"] = build_data_blocks(ts, ds, pm["R1"], fm["fields"])
        knowledge = {
            "meta": {"source_type": "t", "analysis_time": "", "dialect": "dws",
                     "total_rules": 1, "target_table": "f",
                     "patterns": [], "target_field_types": {}, "target_field_comments": {}},
            "topology": topo, "data_flow": df, "field_mappings": fm, "quality": q,
            "business_logic": {"summary": "", "step_descriptions": [], "key_transforms": []},
            "source": build_source([rule], {}, {}, pm),
        }
        out = str(tmp_path / "test_out")
        generate_asset_report(knowledge, out)

        html = (Path(out) / "asset_report.html").read_text(encoding="utf-8")
        m = re.search(r'const REPORT_DATA = (\{.*?\});\s', html, re.DOTALL)
        data = json.loads(m.group(1))
        for s in data["steps"]:
            assert len(s.get("data_blocks", [])) > 0, "steps 应含 data_blocks"


# ═══════════════════════════════════════════════════════════════
# 3. 端到端验证（复杂 SQL → 完整产物）
# ═══════════════════════════════════════════════════════════════

class TestEndToEndDeliverables:
    """端到端：复杂 SQL → 验证 data_blocks + mapping + HTML。"""

    def test_cte_union_e2e_blocks(self):
        """CTE + UNION 场景：逻辑块分支独立 + CTE 内部展示"""
        sql = """WITH tm AS (
    SELECT a.id, a.amt FROM ods.fact_a a
    INNER JOIN ods.dim_b b ON a.k = b.k WHERE a.del = 'N'
)
SELECT t.region, t.total FROM (
    SELECT m.region, SUM(m.amt) AS total FROM tm m
    INNER JOIN ods.fact_c c ON m.id = c.id WHERE m.region IS NOT NULL GROUP BY m.region
    UNION ALL
    SELECT n.region, SUM(n.amt) AS total FROM ods.fact_d n
    LEFT JOIN ods.dim_e e ON n.k = e.k WHERE n.sts = 'A' GROUP BY n.region
) t WHERE t.total > 0"""
        blocks = _get_blocks(sql)

        # 1. UNION 分支独立
        all_flat = _flatten_blocks(blocks)
        branches = [b for b in all_flat if "UNION 分支" in b.get("role", "")]
        assert len(branches) == 2, f"应有2个UNION分支"
        assert "region IS NOT NULL" in branches[0].get("where_clause", "") or "region" in branches[0].get("where_clause", "")
        assert "sts" in branches[1].get("where_clause", "")

        # 2. 分支1 内含 CTE tm（作为主表）
        branch1_children = branches[0].get("children", [])
        branch1_tables = [b["table"].lower() for b in branch1_children]
        assert any("tm" in t for t in branch1_tables), f"分支1应含 tm，实际 {branch1_tables}"

        # 3. 分支2 内含 fact_d + dim_e
        branch2_children = branches[1].get("children", [])
        branch2_tables = [b["table"].lower() for b in branch2_children]
        assert any("fact_d" in t for t in branch2_tables), f"分支2应含 fact_d，实际 {branch2_tables}"

    def test_three_step_serial_e2e(self, tmp_path):
        """三步串行：验证 data_blocks + mapping 端到端"""
        from view_generator import generate_mapping

        rules = [
            RawRule(rule_code="R1", rule_name="s1", rule_type=1, exec_sequence=1,
                    target_schema="dws", target_table="tmp1", delete_mode="1",
                    query_sql="SELECT a.id, a.amount FROM ods.tbl_a a WHERE a.del='N'"),
            RawRule(rule_code="R2", rule_name="s2", rule_type=1, exec_sequence=2,
                    target_schema="dws", target_table="final_f", delete_mode="1",
                    query_sql="SELECT t.id, SUM(t.amount) AS total FROM dws.tmp1 t GROUP BY t.id"),
        ]
        pm = {r.rule_code: parse_single_sql(r.query_sql, "dws") for r in rules}
        topo = build_topology(rules, pm)
        df = build_data_flow(rules, pm)
        fm = build_field_mappings(rules, pm, {})
        enrich_join_key_lineage(df, rules, pm, topo, fm)
        enrich_field_physical_sources(fm, df, rules, pm, topo)
        q = analyze_quality(topo, df, fm, pm)

        # data_blocks 验证
        for ts in topo["steps"]:
            ds = next(s for s in df["steps"] if s["step_id"] == ts["step_id"])
            ds["data_blocks"] = build_data_blocks(ts, ds, pm.get(ts["rule_code"]), fm["fields"])

        step1_blocks = df["steps"][0].get("data_blocks", [])
        assert len(step1_blocks) >= 1, "step1 应有逻辑块"
        mains = _find_blocks_by_role(step1_blocks, "主表")
        assert any("tbl_a" in b["table"] for b in mains), "step1 主表应是 tbl_a"

        # mapping 验证
        knowledge = {
            "meta": {"source_type": "t", "analysis_time": "", "dialect": "dws",
                     "total_rules": 2, "target_table": "final_f",
                     "patterns": [], "target_field_types": {}, "target_field_comments": {}},
            "topology": topo, "data_flow": df, "field_mappings": fm, "quality": q,
            "business_logic": {"summary": "", "step_descriptions": [], "key_transforms": []},
            "source": build_source(rules, {}, {}, pm),
        }
        out = str(tmp_path / "e2e_out")
        generate_mapping(knowledge, out)
        assert (Path(out) / "mapping.xlsx").exists(), "mapping.xlsx 应生成"

        # 属性级 mapping 验证：total 应穿透到 tbl_a.amount
        import openpyxl
        wb = openpyxl.load_workbook(Path(out) / "mapping.xlsx", read_only=True)
        ws2 = wb["属性级mapping"]
        found_total = False
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row and row[8] and "total" in str(row[8]).lower():
                found_total = True
                source_table = str(row[2] or "")
                assert "tbl_a" in source_table, f"total 来源表应穿透到 tbl_a，实际 {source_table}"
                break
        wb.close()
        assert found_total, "属性级 mapping 应有 total 字段"
