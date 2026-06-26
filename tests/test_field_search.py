"""字段使用情况批量搜索测试。

验证多规则组解析、多关键字匹配、字段角色判断、来源追溯。

运行:
    pytest tests/test_field_search.py -v
"""

import sys
import tempfile
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ANALYZER_REF = PROJECT_ROOT / "dws-pipeline-analyzer" / "references"
FIXTURES = PROJECT_ROOT / "tests" / "fixtures" / "analyzer"
sys.path.insert(0, str(ANALYZER_REF))
sys.path.insert(0, str(FIXTURES))

from field_search import search_field_usage, output_excel, read_excel_grouped
from _build_xlsx import build_xlsx


def _make_multi_group_xlsx(path):
    """构造含两个规则组的 Excel。

    组1（GR001, final_f）：amount 字段在 SELECT + 关联 + 过滤
    组2（GR002, other_f）：user_id 字段在 SELECT
    """
    rules = [
        # 组1: amount 的各种用法
        {"rule_code": "R1", "rule_type": 1, "exec_sequence": 1,
         "target_schema": "dws", "target_table": "tmp1", "delete_mode": "1",
         "query_sql": "SELECT a.id, a.amount FROM ods.tbl_a a WHERE a.amount > 0",
         "rule_name": "源头", "rule_group_code": "GR001", "rule_group_en": "FINAL_F"},
        {"rule_code": "R2", "rule_type": 1, "exec_sequence": 2,
         "target_schema": "dws", "target_table": "final_f", "delete_mode": "1",
         "query_sql": "SELECT t.id, t.amount, b.region FROM dws.tmp1 t LEFT JOIN ods.dim_b b ON t.amount = b.amount",
         "rule_name": "关联", "rule_group_code": "GR001", "rule_group_en": "FINAL_F"},
        # 组2: user_id 的用法
        {"rule_code": "R3", "rule_type": 1, "exec_sequence": 1,
         "target_schema": "dws", "target_table": "other_f", "delete_mode": "1",
         "query_sql": "SELECT a.user_id, a.name FROM ods.users a",
         "rule_name": "用户表", "rule_group_code": "GR002", "rule_group_en": "OTHER_F"},
    ]
    build_xlsx(str(path), rules=rules)
    return str(path)


@pytest.fixture
def multi_group_xlsx(tmp_path):
    path = tmp_path / "multi.xlsx"
    return _make_multi_group_xlsx(path)


class TestReadExcelGrouped:
    """多规则组解析。"""

    def test_groups_count(self, multi_group_xlsx):
        """两个规则组应正确分组"""
        groups = read_excel_grouped(multi_group_xlsx)
        assert len(groups) == 2, f"应有 2 个规则组，实际 {len(groups)}"
        codes = {g["rule_group_code"] for g in groups}
        assert codes == {"GR001", "GR002"}

    def test_group_rules(self, multi_group_xlsx):
        """每个规则组的规则数量正确"""
        groups = read_excel_grouped(multi_group_xlsx)
        for g in groups:
            if g["rule_group_code"] == "GR001":
                assert len(g["rules"]) == 2
            elif g["rule_group_code"] == "GR002":
                assert len(g["rules"]) == 1


class TestSearchFieldUsage:
    """字段搜索 + 角色判断。"""

    def test_search_amount_write_field(self, multi_group_xlsx):
        """amount 关键字：找到写入目标表的字段（角色含"写入目标表"）"""
        usages = search_field_usage(multi_group_xlsx, ["amount"])
        write_fields = [u for u in usages if "写入目标表" in u.role and "amount" in u.field_name.lower()]
        assert len(write_fields) >= 1, f"应找到 amount 写入字段，实际 {usages}"

    def test_search_amount_only_final_target(self, multi_group_xlsx):
        """amount 关键字：目标表只显示最终表，不显示中间表"""
        usages = search_field_usage(multi_group_xlsx, ["amount"])
        tables = {u.target_table for u in usages}
        assert "dws.tmp1" not in tables, f"不应出现中间表 tmp1，实际 {tables}"
        assert "dws.final_f" in tables, f"应有最终表 final_f，实际 {tables}"

    def test_search_amount_aux_field(self, multi_group_xlsx):
        """amount 关键字：角色含关联键/过滤条件（合并到一行）"""
        usages = search_field_usage(multi_group_xlsx, ["amount"])
        amount_usages = [u for u in usages if "amount" in u.field_name.lower()]
        # 至少有一个 amount 的角色含关联键或过滤条件
        has_aux = any("关联键" in u.role or "过滤条件" in u.role for u in amount_usages)
        assert has_aux, f"应含关联键/过滤条件角色，实际 roles: {[u.role for u in amount_usages]}"

    def test_search_multiple_keywords(self, multi_group_xlsx):
        """多关键字：amount + user_id 同时搜索"""
        usages = search_field_usage(multi_group_xlsx, ["amount", "user_id"])
        has_amount = any("amount" in u.field_name.lower() for u in usages)
        has_user_id = any("user_id" in u.field_name.lower() for u in usages)
        assert has_amount, "应匹配到 amount"
        assert has_user_id, "应匹配到 user_id"

    def test_search_no_match(self, multi_group_xlsx):
        """不存在的关键字：无匹配"""
        usages = search_field_usage(multi_group_xlsx, ["nonexistent_xyz"])
        assert len(usages) == 0

    def test_source_traceback(self, multi_group_xlsx):
        """写入字段应追溯到最初来源（物理源表）"""
        usages = search_field_usage(multi_group_xlsx, ["amount"])
        write_amounts = [u for u in usages if u.role == "写入目标表" and "amount" in u.field_name.lower()]
        if write_amounts:
            # 至少有一个的 source 含 tbl_a（物理源表）
            has_source = any("tbl_a" in u.source for u in write_amounts)
            assert has_source, f"应追溯到 tbl_a，实际 source: {[u.source for u in write_amounts]}"

    def test_grouped_by_target_table(self, multi_group_xlsx):
        """结果按目标表分组"""
        usages = search_field_usage(multi_group_xlsx, ["amount", "user_id"])
        tables = [u.target_table for u in usages]
        # 应包含两个目标表
        assert "dws.final_f" in tables
        assert "dws.other_f" in tables


class TestOutputExcel:
    """Excel 输出。"""

    def test_output_excel_structure(self, multi_group_xlsx, tmp_path):
        """输出 Excel 应有大 sheet + 表头"""
        usages = search_field_usage(multi_group_xlsx, ["amount"])
        out = str(tmp_path / "field_usage.xlsx")
        ok = output_excel(usages, out)
        assert ok

        import openpyxl
        wb = openpyxl.load_workbook(out, read_only=True)
        assert "字段使用情况" in wb.sheetnames
        ws = wb["字段使用情况"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "目标表" in headers
        assert "字段名" in headers
        assert "字段角色" in headers
        assert "最初来源" in headers
        assert "详情" in headers
        wb.close()
