"""impact_analyzer 测试：关联影响分析（单资产 MVP）。

覆盖:
  - 输入解析（三 Sheet，容错）
  - 三层过滤（平切短路 / 字段命中 / 未命中）
  - 逐跳传播
  - 判定层映射表

运行:
    pytest tests/test_impact_analyzer.py -v
"""

import sys
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ANALYZER_REF = PROJECT_ROOT / "dws-pipeline-analyzer" / "references"
FIXTURES = PROJECT_ROOT / "tests" / "fixtures" / "analyzer"
sys.path.insert(0, str(ANALYZER_REF))
sys.path.insert(0, str(FIXTURES))

from impact_analyzer import (
    TableChange, ChangeItem, Hop, ImpactPath, AnalysisResult,
    read_changes, filter_and_propagate, assess_severity,
    _norm_table, _build_table_index, _parse_bool, _parse_sheet1, _parse_sheet2,
)
from _build_changes import build_changes_xlsx


# ═══════════════════════════════════════════════════════════════
# 辅助：构造测试用 knowledge
# ═══════════════════════════════════════════════════════════════

def _make_knowledge(source_tables=None, fields=None, target_table="dwb_test_f"):
    """构造最小 knowledge 结构，模拟 analyze_pipeline 产出。

    source_tables: [{step_id, rule_code, tables: ["ods.src_a", ...]}]
    fields: [{target_field, rule_code, lineage: [{step, source_field, source_table, raw_sql}]}]
    """
    source_tables = source_tables or []
    fields = fields or []

    steps = []
    raw_sql = {}
    for st in source_tables:
        steps.append({
            "step_id": st["step_id"],
            "rule_code": st["rule_code"],
            "exec_sequence": st.get("exec_sequence", 1),
            "target_table": st.get("target_table", target_table),
            "source_tables_from_sql": st["tables"],
            "is_view_step": False,
            "is_exchange": False,
        })

    return {
        "meta": {"target_table": target_table},
        "topology": {"steps": steps},
        "field_mappings": {"fields": fields},
        "source": {"raw_sql": raw_sql},
        "data_flow": {"blocks": []},
    }


def _field(target_field, rule_code, lineage_entries):
    """构造 field_mappings.fields[] 一项"""
    lineage = []
    for e in lineage_entries:
        lineage.append({
            "step": e["step"],
            "source_table": e.get("source_table", ""),
            "source_field": e.get("source_field", ""),
            "raw_sql": e.get("raw_sql", e.get("source_field", "")),
        })
    return {"target_field": target_field, "rule_code": rule_code, "lineage": lineage}


# ═══════════════════════════════════════════════════════════════
# 测试：输入解析
# ═══════════════════════════════════════════════════════════════

class TestReadChanges:
    """变更清单 Excel 解析（三 Sheet + 容错）。"""

    def test_parse_table_level_sheet(self, tmp_path):
        """Sheet1 表级 mapping 解析。"""
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[
                TableChange(before_table="ods.user_src", after_table="ods.user_new",
                            is_ping_cut=True, note="用户表平切"),
                TableChange(before_table="ods.old_log", after_table="",
                            is_ping_cut=False, note="日志表下线"),
            ],
            field_changes=[],
        )
        tc_list, fc_list, td = read_changes(str(xlsx_path))

        assert len(tc_list) == 2
        assert tc_list[0].before_table == "ods.user_src"
        assert tc_list[0].after_table == "ods.user_new"
        assert tc_list[0].is_ping_cut is True
        assert tc_list[1].is_table_dropped is True

    def test_parse_field_level_sheet(self, tmp_path):
        """Sheet2 字段级 mapping 解析。"""
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[],
            field_changes=[
                ChangeItem(before_table="ods.src_a", before_field="user_id",
                           before_type="varchar(20)", after_table="ods.src_a",
                           after_field="user_id", after_type="varchar(50)",
                           change_type="1:1数据类型/长度变化"),
                ChangeItem(before_table="ods.src_a", before_field="old_col",
                           before_type="int", after_table="", after_field="",
                           after_type="", change_type="1:0废弃字段"),
            ],
        )
        tc_list, fc_list, td = read_changes(str(xlsx_path))

        assert len(fc_list) == 2
        assert fc_list[0].before_field == "user_id"
        assert fc_list[0].change_type == "1:1数据类型/长度变化"
        assert fc_list[1].derived_change_type == "1:0废弃字段"

    def test_column_name_fuzzy_match(self, tmp_path):
        """列名模糊匹配：空格/全角括号差异应容错。"""
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[TableChange(before_table="t1", after_table="t2")],
            field_changes=[],
            # 故意用带空格的列名
            table_headers=["切换前 表名", "切换后表名", "是否平切", "切换说明"],
        )
        tc_list, _, _ = read_changes(str(xlsx_path))
        assert len(tc_list) == 1
        assert tc_list[0].before_table == "t1"

    def test_empty_rows_skipped(self, tmp_path):
        """空行跳过不报错。"""
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[
                TableChange(before_table="", after_table=""),  # 空行
                TableChange(before_table="real", after_table="new"),
            ],
            field_changes=[
                ChangeItem(before_table="", before_field=""),  # 空行
            ],
        )
        tc_list, fc_list, _ = read_changes(str(xlsx_path))
        assert len(tc_list) == 1
        assert tc_list[0].before_table == "real"
        assert len(fc_list) == 0

    def test_missing_file_graceful(self, tmp_path):
        """文件不存在不抛异常，返回空。"""
        tc_list, fc_list, td = read_changes(str(tmp_path / "nonexist.xlsx"))
        assert tc_list == []
        assert fc_list == []


# ═══════════════════════════════════════════════════════════════
# 测试：三层过滤
# ═══════════════════════════════════════════════════════════════

class TestFiltering:
    """三层过滤逻辑。"""

    def test_ping_cut_short_circuit(self):
        """平切表 → 表级短路，字段层跳过。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.user_src"]}],
            fields=[_field("uid", "R001", [{"step": "step_1", "source_field": "user_id"}])],
        )
        result = filter_and_propagate(
            [TableChange(before_table="ods.user_src", after_table="ods.user_new",
                         is_ping_cut=True)],
            [ChangeItem(before_table="ods.user_src", before_field="user_id",
                        change_type="1:1完全一致")],
            knowledge, {},
        )
        # 表级影响有平切
        assert len(result.table_level_impacts) == 1
        assert result.table_level_impacts[0]["type"] == "平切"
        # 字段级不进主表
        assert len(result.field_level_impacts) == 0

    def test_table_not_in_asset_filtered(self):
        """表不在资产源表中 → 未命中。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_a"]}],
        )
        result = filter_and_propagate(
            [],
            [ChangeItem(before_table="ods.unrelated", before_field="col",
                        change_type="1:1数据内容变化")],
            knowledge, {},
        )
        assert len(result.field_level_impacts) == 0
        assert len(result.filtered_out) == 1
        assert "未命中" in result.filtered_out[0]["status"]

    def test_field_not_referenced_filtered(self):
        """表命中但字段没用 → 未命中。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_a"]}],
            fields=[_field("uid", "R001", [{"step": "step_1", "source_field": "user_id"}])],
        )
        result = filter_and_propagate(
            [],
            [ChangeItem(before_table="ods.src_a", before_field="unused_col",
                        change_type="1:1数据内容变化")],
            knowledge, {},
        )
        assert len(result.field_level_impacts) == 0
        assert len(result.filtered_out) == 1
        assert "未命中" in result.filtered_out[0]["status"]

    def test_table_dropped_lists_all_fields(self):
        """整表下线 → 表级影响，列出所有受波及字段。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_drop"]}],
            fields=[
                _field("f1", "R001", [{"step": "step_1", "source_field": "a"}]),
                _field("f2", "R001", [{"step": "step_1", "source_field": "b"}]),
            ],
        )
        result = filter_and_propagate(
            [TableChange(before_table="ods.src_drop", after_table="",
                         is_ping_cut=False)],
            [], knowledge, {},
        )
        assert len(result.table_level_impacts) == 1
        assert result.table_level_impacts[0]["type"] == "表/视图下线"
        assert result.table_level_impacts[0]["status"] == "🔴有影响"
        assert len(result.table_level_impacts[0]["touched_fields"]) == 2


# ═══════════════════════════════════════════════════════════════
# 测试：逐跳传播
# ═══════════════════════════════════════════════════════════════

class TestPropagation:
    """逐跳传播逻辑。"""

    def test_field_propagation_hits_target(self):
        """字段变更传播到目标字段。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_a"]}],
            fields=[_field("uid", "R001", [
                {"step": "step_1", "source_field": "user_id",
                 "raw_sql": "cast(a.user_id as bigint)"}])],
        )
        result = filter_and_propagate(
            [],
            [ChangeItem(before_table="ods.src_a", before_field="user_id",
                        before_type="varchar(20)", after_type="varchar(50)",
                        change_type="1:1数据类型/长度变化")],
            knowledge, {},
        )
        # 有 cast → 待确认
        assert len(result.field_level_impacts) == 1
        row = result.field_level_impacts[0]
        assert row["target_field"] == "uid"
        assert row["status"] == "🟡待确认"

    def test_multi_source_field(self):
        """多源字段（COALESCE）：每源一行。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_a", "ods.src_b"]}],
            fields=[_field("uid", "R001", [
                {"step": "step_1", "source_table": "a", "source_field": "uid",
                 "raw_sql": "coalesce(a.uid, b.uid)"},
                {"step": "step_1", "source_table": "b", "source_field": "uid",
                 "raw_sql": "coalesce(a.uid, b.uid)"}])],
        )
        # 两个源表都变了
        result = filter_and_propagate(
            [],
            [
                ChangeItem(before_table="ods.src_a", before_field="uid",
                           change_type="1:1数据内容变化"),
                ChangeItem(before_table="ods.src_b", before_field="uid",
                           change_type="1:1数据内容变化"),
            ],
            knowledge, {},
        )
        # uid 被两个源命中 → 两行
        uid_rows = [r for r in result.field_level_impacts if r["target_field"] == "uid"]
        assert len(uid_rows) == 2


# ═══════════════════════════════════════════════════════════════
# 测试：判定层
# ═══════════════════════════════════════════════════════════════

class TestSeverity:
    """变化类型 → 状态映射。"""

    def test_deprecated_field_is_high(self):
        """1:0废弃字段 → 🔴有影响。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="1:0废弃字段")
        path = ImpactPath(target_field="f", change=fc)
        assess_severity(path, fc, _make_knowledge(), {})
        assert path.status == "🔴有影响"

    def test_identical_is_no_impact(self):
        """1:1完全一致 → 🟢无影响。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="1:1完全一致")
        path = ImpactPath(target_field="f", change=fc)
        assess_severity(path, fc, _make_knowledge(), {})
        assert path.status == "🟢无影响"

    def test_content_change_is_uncertain(self):
        """1:1数据内容变化 → 🟡待确认。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="1:1数据内容变化")
        path = ImpactPath(target_field="f", change=fc)
        assess_severity(path, fc, _make_knowledge(), {})
        assert path.status == "🟡待确认"

    def test_select_star_marks_uncertain(self):
        """SELECT * 断链 → 🟡待确认，原因说明。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="1:1数据内容变化")
        path = ImpactPath(
            target_field="f", change=fc,
            uncertain_reason="步骤 step_1 使用了 SELECT *",
            hops=[Hop(step="step_1")],
        )
        assess_severity(path, fc, _make_knowledge(), {})
        assert path.status == "🟡待确认"
        assert "SELECT *" in path.reason

    # ── 数据初始化联动 load_strategy（新能力）──

    def test_init_without_ts_incremental_is_high(self):
        """不刷时间戳 + 增量 → 🔴有影响（隐蔽高风险）。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="字段数据初始化（不刷时间戳）")
        knowledge = _make_knowledge()
        knowledge["meta"]["load_strategy"] = "incremental"
        path = ImpactPath(target_field="f", change=fc)
        assess_severity(path, fc, knowledge, {})
        assert path.status == "🔴有影响"
        assert "增量" in path.reason

    def test_init_without_ts_full_is_no_impact(self):
        """不刷时间戳 + 全量 → 🟢无影响（每次全量拉）。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="字段数据初始化（不刷时间戳）")
        knowledge = _make_knowledge()
        knowledge["meta"]["load_strategy"] = "full"
        path = ImpactPath(target_field="f", change=fc)
        assess_severity(path, fc, knowledge, {})
        assert path.status == "🟢无影响"

    def test_init_with_ts_incremental_is_uncertain(self):
        """刷时间戳 + 增量 → 🟡待确认（会触发重拉）。"""
        fc = ChangeItem(before_table="t", before_field="c",
                        change_type="字段数据初始化（刷时间戳）")
        knowledge = _make_knowledge()
        knowledge["meta"]["load_strategy"] = "incremental"
        path = ImpactPath(target_field="f", change=fc)
        assess_severity(path, fc, knowledge, {})
        assert path.status == "🟡待确认"


# ═══════════════════════════════════════════════════════════════
# 测试：表级新变化类型
# ═══════════════════════════════════════════════════════════════

class TestTableChangeTypes:
    """表级 9 种变化类型的映射判定。"""

    def test_table_offline(self):
        """表/视图下线 → 🔴有影响。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_t"]}],
            fields=[_field("f1", "R001", [{"step": "step_1", "source_field": "x"}])],
        )
        result = filter_and_propagate(
            [TableChange(before_table="ods.src_t", change_type="表/视图下线")],
            [], knowledge, {},
        )
        assert len(result.table_level_impacts) == 1
        assert result.table_level_impacts[0]["status"] == "🔴有影响"

    def test_table_revoke_permission(self):
        """表/视图取消权限 → 🔴有影响。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_t"]}],
        )
        result = filter_and_propagate(
            [TableChange(before_table="ods.src_t", change_type="表/视图取消权限")],
            [], knowledge, {},
        )
        assert result.table_level_impacts[0]["status"] == "🔴有影响"

    def test_table_schema_change(self):
        """表/视图名称或schema变化 → 🟡待确认。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_t"]}],
        )
        result = filter_and_propagate(
            [TableChange(before_table="ods.src_t", after_table="ods2.src_t",
                         change_type="表/视图名称或者schema变化")],
            [], knowledge, {},
        )
        assert result.table_level_impacts[0]["status"] == "🟡待确认"

    def test_table_init_without_ts_incremental(self):
        """表级数据初始化不刷时间戳 + 增量 → 🔴有影响。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_t"]}],
        )
        knowledge["meta"]["load_strategy"] = "incremental"
        result = filter_and_propagate(
            [TableChange(before_table="ods.src_t",
                         change_type="表/视图初始化（不刷时间戳）")],
            [], knowledge, {},
        )
        assert result.table_level_impacts[0]["status"] == "🔴有影响"

    def test_table_init_with_ts_full(self):
        """表级数据初始化刷时间戳 + 全量 → 🟢无影响。"""
        knowledge = _make_knowledge(
            source_tables=[{"step_id": "step_1", "rule_code": "R001",
                            "tables": ["ods.src_t"]}],
        )
        knowledge["meta"]["load_strategy"] = "full"
        result = filter_and_propagate(
            [TableChange(before_table="ods.src_t",
                         change_type="表/视图数据初始化（刷时间戳）")],
            [], knowledge, {},
        )
        assert result.table_level_impacts[0]["status"] == "🟢无影响"

    def test_table_change_type_column_parsed(self, tmp_path):
        """Sheet1 新增「表级变化类型」列能被正确解析。"""
        from _build_changes import build_changes_xlsx
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[
                TableChange(before_table="ods.src_t", change_type="表/视图数据归档"),
            ],
            field_changes=[],
        )
        tc_list, _, _ = read_changes(str(xlsx_path))
        assert len(tc_list) == 1
        assert tc_list[0].change_type == "表/视图数据归档"


# ═══════════════════════════════════════════════════════════════
# 测试：渲染
# ═══════════════════════════════════════════════════════════════

class TestRender:
    """Excel 渲染。"""

    def test_render_produces_valid_xlsx(self, tmp_path):
        """渲染产出有效 xlsx，含四 Sheet。"""
        result = AnalysisResult(
            table_level_impacts=[{
                "status": "🟡待确认", "type": "平切",
                "source_table": "ods.t", "new_table": "ods.t2",
                "note": "test", "steps": ["step_1"], "rule_codes": ["R001"],
                "touched_fields": [],
            }],
            field_level_impacts=[{
                "status": "🔴有影响", "severity": "high",
                "target_table": "f", "target_field": "x",
                "source_table": "ods.t", "source_field": "y",
                "change_type": "1:0废弃字段",
                "before_type": "int", "after_type": "",
                "reason": "test", "hops": "R001/step_1: y",
                "steps": "step_1", "rule_codes": "R001", "recovery_plan": "",
            }],
            filtered_out=[{
                "status": "⚪未命中", "source_table": "ods.t",
                "source_field": "z", "change_type": "0:1新增字段",
                "reason": "test", "target_field": "",
            }],
            summary={"impacted": 1, "uncertain": 0, "no_impact": 0,
                     "not_hit": 1, "table_level": 1,
                     "total_field_changes": 2, "total_table_changes": 1},
        )
        from impact_analyzer import render_excel
        out = tmp_path / "impact.xlsx"
        render_excel(result, str(out), "TEST_ASSET")
        assert out.exists()
        # 验证 Sheet 数量
        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "统计摘要" in wb.sheetnames
        assert "影响清单" in wb.sheetnames
        assert "表级影响" in wb.sheetnames
        assert "过滤摘要" in wb.sheetnames
        wb.close()


# ═══════════════════════════════════════════════════════════════
# 测试：端到端
# ═══════════════════════════════════════════════════════════════

class TestEndToEnd:
    """完整流程：Excel 变更清单 + knowledge → 分析 → Excel 报告。"""

    def test_full_pipeline(self, tmp_path):
        """端到端：构造变更清单 + knowledge，跑完整分析。"""
        knowledge = _make_knowledge(
            source_tables=[
                {"step_id": "step_1", "rule_code": "R001", "tables": ["ods.src_a"]},
                {"step_id": "step_2", "rule_code": "R002", "tables": ["ods.src_b"]},
            ],
            fields=[
                _field("uid", "R001", [
                    {"step": "step_1", "source_field": "user_id",
                     "raw_sql": "cast(a.user_id as bigint)"}]),
                _field("amount", "R002", [
                    {"step": "step_2", "source_field": "amt",
                     "raw_sql": "b.amt"}]),
            ],
        )
        # 变更清单
        xlsx_path = tmp_path / "changes.xlsx"
        build_changes_xlsx(str(xlsx_path),
            table_changes=[],
            field_changes=[
                ChangeItem(before_table="ods.src_a", before_field="user_id",
                           before_type="varchar(20)", after_type="varchar(50)",
                           change_type="1:1数据类型/长度变化"),
                ChangeItem(before_table="ods.src_b", before_field="amt",
                           before_type="int", after_type="bigint",
                           change_type="1:1数据类型/长度变化"),
                ChangeItem(before_table="ods.unrelated", before_field="x",
                           change_type="1:1数据内容变化"),
            ],
        )
        tc_list, fc_list, td = read_changes(str(xlsx_path))
        result = filter_and_propagate(tc_list, fc_list, knowledge, td)

        # src_a.user_id 有 cast → 待确认，进主表
        # src_b.amt 无 cast，int→bigint 直传，需看 DDL（无 DDL → 待确认）
        assert len(result.field_level_impacts) == 2
        statuses = {r["status"] for r in result.field_level_impacts}
        assert all(s == "🟡待确认" for s in statuses)
        # unrelated 未命中
        assert result.summary["not_hit"] >= 1

        # 渲染
        out = tmp_path / "impact.xlsx"
        from impact_analyzer import render_excel
        render_excel(result, str(out), "E2E_TEST")
        assert out.exists()
