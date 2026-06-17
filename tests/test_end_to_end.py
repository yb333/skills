"""
端到端用例 — 合成 Excel → 完整 analyzer + view_generator → 三视图产物验证

用 A 类用例的 case 定义动态生成 Excel，不依赖外部文件。

运行:
    pytest tests/test_end_to_end.py -v
"""

import sys
import json
import shutil
import tempfile
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ANALYZER_REF = PROJECT_ROOT / "dws-pipeline-analyzer" / "references"
FIXTURES = PROJECT_ROOT / "tests" / "fixtures" / "analyzer"
sys.path.insert(0, str(ANALYZER_REF))
sys.path.insert(0, str(FIXTURES))
sys.path.insert(0, str(FIXTURES / "cases"))

from analyzer import (
    read_excel, detect_dialect, parse_single_sql,
    build_topology, build_data_flow, build_field_mappings, analyze_quality,
    detect_patterns, build_source, generate_step_description,
)
from view_generator import generate_mapping, generate_asset_report, generate_tech_design
from _build_xlsx import build_xlsx


def run_full_analysis(xlsx_path, output_dir):
    """完整分析流程"""
    from datetime import datetime
    raw = read_excel(xlsx_path)
    rules = raw["rules"]
    dialect = detect_dialect([r.query_sql for r in rules if r.query_sql])
    parsed_map = {r.rule_code: parse_single_sql(r.query_sql, dialect) for r in rules}
    topology = build_topology(rules, parsed_map)
    data_flow = build_data_flow(rules, parsed_map)
    field_mappings = build_field_mappings(rules, parsed_map, raw["target_fields"])
    quality = analyze_quality(topology, data_flow, field_mappings, parsed_map)
    patterns = detect_patterns(parsed_map, topology)
    scenarios = topology.get("scenarios", [])
    auto_step_desc = []
    for rule in rules:
        parsed = parsed_map.get(rule.rule_code)
        desc = generate_step_description(rule, parsed, scenarios, rules)
        step = next((s for s in topology["steps"] if s["rule_code"] == rule.rule_code), None)
        auto_step_desc.append({"step_id": step["step_id"] if step else "", "rule_code": rule.rule_code, "purpose": desc["purpose"], "logic": desc["logic"]})
    knowledge = {"meta": {"source_type": "execution_tasks.xlsx", "analysis_time": datetime.now().isoformat(), "dialect": dialect, "total_rules": len(rules), "target_table": rules[0].target_table or "", "patterns": patterns, "target_field_types": {}, "target_field_comments": {}}, "topology": topology, "data_flow": data_flow, "field_mappings": field_mappings, "quality": quality, "business_logic": {"summary": "", "step_descriptions": auto_step_desc, "key_transforms": []}, "source": build_source(rules, raw["target_fields"], raw["group_variables"], parsed_map)}
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    (out / "knowledge_final.json").write_text(json.dumps(knowledge, ensure_ascii=False, indent=2))
    results = {}
    results["mapping"] = generate_mapping(knowledge, str(out))
    results["asset"] = generate_asset_report(knowledge, str(out))
    results["techspec"] = generate_tech_design(knowledge, str(out))
    return knowledge, results


class TestEndToEnd:

    @pytest.fixture
    def tmp_output(self):
        d = tempfile.mkdtemp()
        yield d
        shutil.rmtree(d, ignore_errors=True)

    def _make_xlsx(self, case_name, tmp_dir):
        mod = __import__(case_name)
        xlsx_path = Path(tmp_dir) / f"{case_name}.xlsx"
        build_xlsx(str(xlsx_path), rules=mod.rules, target_fields=getattr(mod, "target_fields", None), group_variables=getattr(mod, "group_variables", None))
        return str(xlsx_path)

    def test_cte_pivot_e2e(self, tmp_output):
        """CTE场景：三视图全部生成成功。"""
        xlsx = self._make_xlsx("case_02_cte_basic", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        assert knowledge is not None
        assert results["mapping"] == True
        assert results["asset"] == True
        assert results["techspec"] == True
        assert (Path(tmp_output) / "mapping.xlsx").exists()
        assert (Path(tmp_output) / "asset_report.html").exists()
        assert (Path(tmp_output) / "tech_design.md").exists()
        html = (Path(tmp_output) / "asset_report.html").read_text()
        assert "REPORT_DATA" in html
        assert len(html) > 5000

    def test_multi_scenario_e2e(self, tmp_output):
        """多场景：场景分组正确，三视图生成。"""
        xlsx = self._make_xlsx("case_19_multi_scenario_2", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        assert knowledge is not None
        assert results["mapping"] == True
        assert results["asset"] == True
        assert results["techspec"] == True
        scenarios = knowledge["topology"].get("scenarios", [])
        non_common = [s for s in scenarios if not s.get("is_common")]
        assert len(non_common) >= 2
        html = (Path(tmp_output) / "asset_report.html").read_text()
        assert "is_multi_scenario" in html

    def test_mapping_xlsx_structure(self, tmp_output):
        """mapping.xlsx 结构验证。"""
        xlsx = self._make_xlsx("case_02_cte_basic", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        assert results["mapping"] == True
        import openpyxl
        wb = openpyxl.load_workbook(Path(tmp_output) / "mapping.xlsx", read_only=True)
        assert "实体级mapping" in wb.sheetnames
        assert "属性级mapping" in wb.sheetnames
        assert wb["实体级mapping"].max_row >= 2
        assert wb["属性级mapping"].max_row >= 2
        wb.close()

    def test_tech_design_structure(self, tmp_output):
        """tech_design.md 结构验证。"""
        xlsx = self._make_xlsx("case_04_pivot", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        assert results["techspec"] == True
        md = (Path(tmp_output) / "tech_design.md").read_text()
        assert len(md) > 500

    def test_lineage_is_dict_not_list(self, tmp_output):
        """回归: REPORT_DATA.lineage 必须是 dict（布局对象），不是 list（字段 lineage）。
        防止变量名覆盖导致数据流图不渲染。"""
        xlsx = self._make_xlsx("case_02_cte_basic", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        html = (Path(tmp_output) / "asset_report.html").read_text()
        import re, json
        m = re.search(r'const REPORT_DATA = ({.*?});\s', html, re.DOTALL)
        assert m, "REPORT_DATA 未找到"
        data = json.loads(m.group(1))
        assert "lineage" in data, "lineage 缺失"
        assert isinstance(data["lineage"], dict), \
            f"lineage 应为 dict（布局对象），实际是 {type(data['lineage']).__name__}"
        assert "nodes" in data["lineage"], "lineage 缺少 nodes"
        assert len(data["lineage"]["nodes"]) > 0, "lineage nodes 为空"

    def test_target_table_is_max_sequence(self, tmp_output):
        """回归: target_table 取最大 exec_sequence 的步骤，不是 steps[0]。"""
        xlsx = self._make_xlsx("case_19_multi_scenario_2", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        html = (Path(tmp_output) / "asset_report.html").read_text()
        import re, json
        m = re.search(r'const REPORT_DATA = ({.*?});\s', html, re.DOTALL)
        data = json.loads(m.group(1))
        target = data["summary"]["target_table"]
        # case_19 的 max seq=2 是 step_5，目标表是 dwl_inv_summary_f
        assert "summary" in target.lower(), \
            f"目标表应取 max seq 步骤(dwl_inv_summary_f)，实际取到 {target}"

    def test_html_js_syntax_valid(self, tmp_output):
        """回归: 生成的 HTML 里 JS 括号必须匹配，防止语法错误导致页面空白。"""
        xlsx = self._make_xlsx("case_19_multi_scenario_2", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        html = (Path(tmp_output) / "asset_report.html").read_text()
        import re
        scripts = re.findall(r'<script[^>]*>(.*?)</script>', html, re.DOTALL)
        for i, s in enumerate(scripts):
            if not s.strip():
                continue
            opens = s.count('{')
            closes = s.count('}')
            assert opens == closes, \
                f"script {i}: JS 大括号不匹配 ({opens} {{ vs {closes} }})，页面会空白"

    def test_html_div_tags_balanced(self, tmp_output):
        """回归: 生成的 HTML 的 div 开闭标签必须平衡（静态模板层级）。"""
        template = Path(__file__).resolve().parent.parent / "dws-pipeline-analyzer" / "references" / "templates" / "asset_report.html"
        html = template.read_text(encoding="utf-8")
        # 只检查静态 HTML 部分（script 标签前）
        static = html.split("<script>")[0]
        opens = static.count("<div")
        closes = static.count("</div>")
        assert opens == closes, \
            f"静态 HTML div 标签不平衡 ({opens} 开 vs {closes} 闭)，会导致布局层级错误"

    def test_last_topology_node_is_target(self, tmp_output):
        """回归: 数据流图拓扑序的最后一个节点必须是目标表，不是步骤。"""
        xlsx = self._make_xlsx("case_23_exchange_partition", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        html = (Path(tmp_output) / "asset_report.html").read_text()
        import re, json
        m = re.search(r'const REPORT_DATA = ({.*?});\s', html, re.DOTALL)
        data = json.loads(m.group(1))
        lineage = data.get("lineage", {})
        assert isinstance(lineage, dict), "lineage 应为 dict"
        nodes = lineage.get("nodes", [])
        assert len(nodes) > 0, "无节点"
        max_layer = max(n["layer"] for n in nodes)
        last_nodes = [n for n in nodes if n["layer"] == max_layer]
        for n in last_nodes:
            assert n["type"] == "target", \
                f"最后一层(layer={max_layer})应为目标表，实际有 {n['type']}: {n['name']}"

    def test_exchange_partition_data_flow(self, tmp_output):
        """回归: 分区交换步骤在 data_flow 里有正确的目标表和步骤详情。"""
        xlsx = self._make_xlsx("case_23_exchange_partition", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        steps = knowledge["data_flow"]["steps"]
        # step_2 是交换分区，target 应该是 dwl_real_f
        exchange_step = next(s for s in steps if "exchange" in s.get("write_mode", "").lower() or s["step_id"] == "step_2")
        assert "dwl_real_f" in exchange_step["target_table"].lower(), \
            f"交换分区目标表应为 dwl_real_f，实际 {exchange_step['target_table']}"
        # dwl_real_f 应该出现在 tables 里
        table_names = [t["name"].lower() for t in knowledge["data_flow"]["tables"]]
        assert "dwl_real_f" in table_names, f"dwl_real_f 应在 tables 里，实际 {table_names}"

    def test_table_case_dedup(self, tmp_output):
        """回归: 表名大小写不一致时 data_flow.tables 不重复，
        data_dependencies 不丢失。"""
        xlsx = self._make_xlsx("case_25_table_case_dedup", tmp_output)
        knowledge, results = run_full_analysis(xlsx, tmp_output)
        tables = knowledge["data_flow"]["tables"]
        # 按归一化名检查重复
        seen = set()
        for t in tables:
            norm = f"{t['schema']}.{t['name']}".lower()
            assert norm not in seen, f"表重复(大小写): {norm} 出现多次，tables={[t['name'] for t in tables]}"
            seen.add(norm)
        # 依赖不应为空（大小写不应导致 target_writers 匹配失败）
        deps = knowledge["topology"]["data_dependencies"]
        assert len(deps) >= 1, f"应有数据依赖(step_1→step_2)，实际 {len(deps)}"
